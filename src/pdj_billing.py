from __future__ import annotations

import datetime as dt
import json
import re
import uuid
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# PDF (factures simplifiées par site)
import io
import zipfile

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas


# -----------------------------------------------------------------------------
# Objectif
# -----------------------------------------------------------------------------
# Module "Facturation PDJ" (Petit-déjeuner / goûter)
# - Permet d'enregistrer plusieurs bons de commande (lignes produit / quantités)
# - Normalise le site de facturation (24 ter + 24 simple => Internat)
# - Gère des prix unitaires par produit (optionnellement spécifiques à un site)
# - Autorise des ajustements manuels (consommations supplémentaires) et des avoirs
# - Exporte un classeur Excel mensuel (détail + synthèse)


# -----------------------------------------------------------------------------
# Produits par défaut (issus du modèle de bon PDJ fourni)
# -----------------------------------------------------------------------------

DEFAULT_PRODUCTS: List[str] = [
    "Lait demi - écrémé",
    "Lait entier",
    "Céréales",
    "Biscotte",
    "Sucre en sachet",
    "Sucre en morceau",
    "Beurre, plaquettes de 250g",
    "Chocolat en poudre",
    "Brioche",
    "Bledine arome chocolat",
    "Bledine arome vanille",
    "Confiture en carton",
    "Thé en boite",
    "Café en carton",
    "Jus d'orange",
    "Jus de pomme",
    "Jus de raisin",
    "Mayonnaise",
    "Ketchup",
    "Sel",
    "Poivre",
    "Fromage blanc pot de 5kg",
    "Yaourt Nature",
]


# -----------------------------------------------------------------------------
# Storage
# -----------------------------------------------------------------------------


def _data_dir() -> Path:
    base = Path(__file__).resolve().parent.parent
    d = base / "data" / "facturation_pdj"
    try:
        d.mkdir(parents=True, exist_ok=True)
        return d
    except PermissionError:
        # Fallback (environnements en lecture seule)
        d2 = Path.home() / ".gestion_cuisine" / "facturation_pdj"
        d2.mkdir(parents=True, exist_ok=True)
        return d2


def _records_path() -> Path:
    return _data_dir() / "pdj_records.csv"


def _prices_path() -> Path:
    return _data_dir() / "pdj_unit_prices.csv"


def _money_adj_path() -> Path:
    return _data_dir() / "pdj_money_adjustments.csv"


def _meta_path() -> Path:
    return _data_dir() / "meta.json"


def _exports_dir() -> Path:
    """Répertoire de stockage des exports PDJ (persistants)."""
    d = _data_dir() / "exports"
    d.mkdir(parents=True, exist_ok=True)
    return d


def exports_for_month_dir(month: str) -> Path:
    """Sous-dossier des exports pour un mois (YYYY-MM)."""
    d = _exports_dir() / str(month)
    d.mkdir(parents=True, exist_ok=True)
    return d


def list_saved_exports(month: Optional[str] = None) -> List[Path]:
    """Liste les exports enregistrés. Si month est fourni, filtre sur ce mois."""
    base = _exports_dir()
    if month:
        d = base / str(month)
        if not d.exists():
            return []
        return sorted([p for p in d.iterdir() if p.is_file()])
    out: List[Path] = []
    for d in sorted([p for p in base.iterdir() if p.is_dir()]):
        out.extend(sorted([p for p in d.iterdir() if p.is_file()]))
    return out


def delete_saved_export(path: str) -> bool:
    """Supprime un export PDJ enregistré (sécurisé : uniquement dans le dossier exports)."""
    try:
        p = Path(path).resolve()
        root = _exports_dir().resolve()
        if root not in p.parents:
            return False
        if p.exists() and p.is_file():
            p.unlink()
            return True
    except Exception:
        return False
    return False


def _read_meta() -> dict:
    p = _meta_path()
    if not p.exists():
        return {}
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _write_meta(meta: dict) -> None:
    _meta_path().write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")


# -----------------------------------------------------------------------------
# Normalisation
# -----------------------------------------------------------------------------


def _norm(s: str) -> str:
    return (str(s or "")).strip().lower()


def norm_site_facturation(site: str) -> str:
    """Règle métier: '24 ter' + '24 simple' => colonne unique 'Internat'."""
    s0 = str(site or "").strip()
    sN = _norm(s0)

    if re.fullmatch(r"24\s*(ter|simple)", sN):
        return "Internat"
    if sN in {"24ter", "24simple"}:
        return "Internat"
    if "24" in sN and ("ter" in sN or "simple" in sN):
        return "Internat"

    return s0


def norm_product(p: str) -> str:
    # On conserve la casse d'origine la plupart du temps, mais on nettoie les espaces.
    return " ".join(str(p or "").strip().split())


# -----------------------------------------------------------------------------
# Chargement / sauvegarde
# -----------------------------------------------------------------------------


def load_pdj_records() -> pd.DataFrame:
    p = _records_path()
    if not p.exists():
        return pd.DataFrame(
            columns=[
                "date",
                "month",
                "site",
                "product",
                "qty",
                "kind",
                "comment",
                "source",
            ]
        )
    df = pd.read_csv(p, parse_dates=["date"])
    df["date"] = pd.to_datetime(df["date"]).dt.date
    df["month"] = df["month"].astype(str)
    df["site"] = df["site"].astype(str)
    df["product"] = df["product"].astype(str)
    df["qty"] = pd.to_numeric(df["qty"], errors="coerce").fillna(0)
    df["kind"] = df["kind"].astype(str)
    df["comment"] = df.get("comment", "").astype(str)
    df["source"] = df.get("source", "").astype(str)
    # Compat rétro: anciens fichiers sans identifiant
    if "record_id" not in df.columns:
        df["record_id"] = [str(uuid.uuid4()) for _ in range(len(df))]
        df.to_csv(p, index=False)
    else:
        df["record_id"] = df["record_id"].astype(str)
    return df


def add_pdj_records(
    records: pd.DataFrame,
    *,
    source_filename: str = "",
) -> int:
    """Ajoute (append) des lignes PDJ.

    records attend au minimum: date, site, product, qty.
    Optionnel: kind (commande|manuel|avoir_qty)
    """
    if records is None or records.empty:
        return 0

    df = records.copy()
    if "date" not in df.columns:
        raise ValueError("La colonne 'date' est obligatoire")
    if "site" not in df.columns:
        raise ValueError("La colonne 'site' est obligatoire")
    if "product" not in df.columns:
        raise ValueError("La colonne 'product' est obligatoire")
    if "qty" not in df.columns:
        raise ValueError("La colonne 'qty' est obligatoire")

    df["date"] = pd.to_datetime(df["date"]).dt.date
    df["month"] = df["date"].map(lambda d: f"{d.year:04d}-{d.month:02d}")
    df["site"] = df["site"].astype(str).map(norm_site_facturation)
    df["product"] = df["product"].astype(str).map(norm_product)
    df["qty"] = pd.to_numeric(df["qty"], errors="coerce").fillna(0)
    if "kind" not in df.columns:
        df["kind"] = "commande"
    df["kind"] = df["kind"].astype(str)

    # Sécurité métier : un avoir en quantités doit être négatif.
    # Beaucoup d'utilisateurs saisissent "5" au lieu de "-5".
    try:
        mask_avoir = df["kind"].str.lower().eq("avoir_qty")
        if mask_avoir.any():
            df.loc[mask_avoir, "qty"] = -df.loc[mask_avoir, "qty"].abs()
    except Exception:
        pass
    if "comment" not in df.columns:
        df["comment"] = ""
    df["comment"] = df["comment"].astype(str)
    df["source"] = source_filename
    df["record_id"] = [str(uuid.uuid4()) for _ in range(len(df))]

    df = df[["record_id", "date", "month", "site", "product", "qty", "kind", "comment", "source"]]

    p = _records_path()
    if p.exists():
        old = pd.read_csv(p, parse_dates=["date"])
        old["date"] = pd.to_datetime(old["date"]).dt.date
    else:
        old = pd.DataFrame(columns=df.columns)

    merged = pd.concat([old, df], ignore_index=True)
    merged.to_csv(p, index=False)

    meta = _read_meta()
    meta.setdefault("last_update", dt.datetime.now().isoformat(timespec="seconds"))
    _write_meta(meta)
    return int(len(df))


def delete_pdj_records(
    *,
    month: Optional[str] = None,
    site: Optional[str] = None,
    source: Optional[str] = None,
    kind: Optional[str] = None,
    record_ids: Optional[List[str]] = None,
) -> int:
    """Supprime des lignes PDJ stockées.

    - Si record_ids est fourni: supprime exactement ces lignes.
    - Sinon, supprime selon filtres (mois/site/source/kind).
    Retourne le nombre de lignes supprimées.
    """
    p = _records_path()
    if not p.exists():
        return 0

    df = load_pdj_records()
    before = len(df)

    mask = pd.Series([True] * len(df))
    if record_ids:
        ids = {str(x) for x in record_ids}
        mask &= df["record_id"].astype(str).isin(ids)
    else:
        if month:
            mask &= df["month"].astype(str) == str(month)
        if site and str(site).strip():
            mask &= df["site"].astype(str) == norm_site_facturation(site)
        if source and str(source).strip():
            mask &= df["source"].astype(str) == str(source)
        if kind and str(kind).strip():
            mask &= df["kind"].astype(str) == str(kind)

    to_delete = df.loc[mask]
    if to_delete.empty:
        return 0

    df2 = df.loc[~mask].copy()
    df2.to_csv(p, index=False)
    return int(before - len(df2))


def load_unit_prices() -> pd.DataFrame:
    """Tarifs unitaires.

    Colonnes:
    - product
    - site (optionnel; '__default__' si commun)
    - unit_price
    - unit (optionnel)
    """
    p = _prices_path()
    if not p.exists():
        # Initialise une grille avec les produits par défaut
        return pd.DataFrame(
            {
                "product": DEFAULT_PRODUCTS,
                "site": "__default__",
                "unit_price": 0.0,
                "unit": "",
            }
        )
    df = pd.read_csv(p)
    df["product"] = df["product"].astype(str).map(norm_product)
    df["site"] = df.get("site", "__default__").astype(str)
    df["unit_price"] = pd.to_numeric(df.get("unit_price", 0.0), errors="coerce").fillna(0.0)
    df["unit"] = df.get("unit", "").astype(str)
    return df[["product", "site", "unit_price", "unit"]]


def save_unit_prices(prices_df: pd.DataFrame) -> int:
    if prices_df is None or prices_df.empty:
        return 0
    df = prices_df.copy()
    if "product" not in df.columns:
        raise ValueError("La colonne 'product' est obligatoire")
    df["product"] = df["product"].astype(str).map(norm_product)
    if "site" not in df.columns:
        df["site"] = "__default__"
    df["site"] = df["site"].astype(str)
    df["unit_price"] = pd.to_numeric(df.get("unit_price", 0.0), errors="coerce").fillna(0.0)
    if "unit" not in df.columns:
        df["unit"] = ""
    df["unit"] = df["unit"].astype(str)
    df = df[["product", "site", "unit_price", "unit"]]

    _prices_path().write_text(df.to_csv(index=False), encoding="utf-8")
    return int(len(df))


def load_money_adjustments() -> pd.DataFrame:
    """Ajustements monétaires (avoirs / frais / corrections) en euros.

    Colonnes:
    - date
    - month
    - site
    - label
    - amount_eur (positif ou négatif)
    - comment
    """
    p = _money_adj_path()
    if not p.exists():
        return pd.DataFrame(columns=["adj_id", "date", "month", "site", "label", "amount_eur", "comment"])
    df = pd.read_csv(p, parse_dates=["date"])
    df["date"] = pd.to_datetime(df["date"]).dt.date
    df["month"] = df["month"].astype(str)
    df["site"] = df["site"].astype(str)
    df["label"] = df.get("label", "Ajustement").astype(str)
    df["amount_eur"] = pd.to_numeric(df.get("amount_eur", 0.0), errors="coerce").fillna(0.0)
    df["comment"] = df.get("comment", "").astype(str)

    # Compat rétro: anciens fichiers sans identifiant
    if "adj_id" not in df.columns:
        df["adj_id"] = [str(uuid.uuid4()) for _ in range(len(df))]
        df.to_csv(p, index=False)
    else:
        df["adj_id"] = df["adj_id"].astype(str)

    return df[["adj_id", "date", "month", "site", "label", "amount_eur", "comment"]]


def add_money_adjustments(adj: pd.DataFrame) -> int:
    if adj is None or adj.empty:
        return 0
    df = adj.copy()
    if "date" not in df.columns:
        raise ValueError("La colonne 'date' est obligatoire")
    if "site" not in df.columns:
        raise ValueError("La colonne 'site' est obligatoire")
    if "amount_eur" not in df.columns:
        raise ValueError("La colonne 'amount_eur' est obligatoire")
    if "label" not in df.columns:
        df["label"] = "Ajustement"
    if "comment" not in df.columns:
        df["comment"] = ""

    df["date"] = pd.to_datetime(df["date"]).dt.date
    df["month"] = df["date"].map(lambda d: f"{d.year:04d}-{d.month:02d}")
    df["site"] = df["site"].astype(str).map(norm_site_facturation)
    df["label"] = df["label"].astype(str)
    df["amount_eur"] = pd.to_numeric(df["amount_eur"], errors="coerce").fillna(0.0)
    df["comment"] = df["comment"].astype(str)
    df["adj_id"] = [str(uuid.uuid4()) for _ in range(len(df))]
    df = df[["adj_id", "date", "month", "site", "label", "amount_eur", "comment"]]

    p = _money_adj_path()
    if p.exists():
        old = pd.read_csv(p, parse_dates=["date"])
        old["date"] = pd.to_datetime(old["date"]).dt.date
    else:
        old = pd.DataFrame(columns=df.columns)
    merged = pd.concat([old, df], ignore_index=True)
    merged.to_csv(p, index=False)
    return int(len(df))


def delete_money_adjustments(
    *,
    month: Optional[str] = None,
    site: Optional[str] = None,
    adj_ids: Optional[List[str]] = None,
) -> int:
    """Supprime des ajustements en euros.

    - Si adj_ids est fourni: supprime exactement ces lignes.
    - Sinon, supprime selon filtres (mois/site).
    """
    p = _money_adj_path()
    if not p.exists():
        return 0

    df = load_money_adjustments()
    before = len(df)

    mask = pd.Series([True] * len(df))
    if adj_ids:
        ids = {str(x) for x in adj_ids}
        mask &= df["adj_id"].astype(str).isin(ids)
    else:
        if month:
            mask &= df["month"].astype(str) == str(month)
        if site and str(site).strip():
            mask &= df["site"].astype(str) == norm_site_facturation(site)

    if not df.loc[mask].any().any():
        return 0

    df2 = df.loc[~mask].copy()
    df2.to_csv(p, index=False)
    return int(before - len(df2))


# -----------------------------------------------------------------------------
# Calcul facture
# -----------------------------------------------------------------------------


def _pick_unit_price(prices: pd.DataFrame, product: str, site: str) -> Tuple[float, str]:
    """Retourne (prix, unité).

    Priorité:
    1) (product, site)
    2) (product, '__default__')
    3) 0.0
    """
    if prices is None or prices.empty:
        return 0.0, ""
    p = norm_product(product)
    s = str(site)
    sub = prices[(prices["product"] == p) & (prices["site"] == s)]
    if not sub.empty:
        r = sub.iloc[0]
        return float(r.get("unit_price", 0.0)), str(r.get("unit", "") or "")
    sub = prices[(prices["product"] == p) & (prices["site"] == "__default__")]
    if not sub.empty:
        r = sub.iloc[0]
        return float(r.get("unit_price", 0.0)), str(r.get("unit", "") or "")
    return 0.0, ""


def compute_monthly_pdj(month: str) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Retourne (synthèse_site, détail_lignes, ajustements_monetaires) pour un mois YYYY-MM."""
    rec = load_pdj_records()
    rec = rec[rec["month"] == month].copy()
    rec["site"] = rec["site"].astype(str).map(norm_site_facturation)
    rec["product"] = rec["product"].astype(str).map(norm_product)
    prices = load_unit_prices()

    if rec.empty:
        detail = pd.DataFrame(
            columns=["date", "site", "product", "qty", "unit_price", "unit", "amount_eur", "kind", "comment", "source"]
        )
    else:
        ups = []
        units = []
        for _, r in rec.iterrows():
            up, u = _pick_unit_price(prices, r["product"], r["site"])
            ups.append(up)
            units.append(u)
        rec["unit_price"] = ups
        rec["unit"] = units
        rec["amount_eur"] = rec["qty"].astype(float) * rec["unit_price"].astype(float)
        detail = rec[["date", "site", "product", "qty", "unit_price", "unit", "amount_eur", "kind", "comment", "source"]].copy()

    # Pivot synthèse
    if detail.empty:
        synth = pd.DataFrame(columns=["site", "total_eur"])
    else:
        synth = (
            detail.groupby("site", as_index=False)["amount_eur"].sum().sort_values("amount_eur", ascending=False)
        )
        synth = synth.rename(columns={"amount_eur": "total_eur"})

    # Ajustements €
    adj = load_money_adjustments()
    adj = adj[adj["month"] == month].copy()
    if not adj.empty:
        # Intègre aux totaux
        add = adj.groupby("site", as_index=False)["amount_eur"].sum().rename(columns={"amount_eur": "adj_eur"})
        synth = synth.merge(add, on="site", how="outer")
        synth["total_eur"] = pd.to_numeric(synth.get("total_eur", 0.0), errors="coerce").fillna(0.0)
        synth["adj_eur"] = pd.to_numeric(synth.get("adj_eur", 0.0), errors="coerce").fillna(0.0)
        synth["total_facture_eur"] = synth["total_eur"] + synth["adj_eur"]
    else:
        synth["total_facture_eur"] = synth.get("total_eur", 0.0)

    return synth, detail, adj


# -----------------------------------------------------------------------------
# Export Excel
# -----------------------------------------------------------------------------


def export_monthly_pdj_workbook(month: str, out_path: str) -> str:
    """Exporte le classeur de facturation PDJ pour le mois YYYY-MM."""
    synth, detail, adj = compute_monthly_pdj(month)
    prices = load_unit_prices()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Facturation PDJ"

    # Styles
    header_fill = PatternFill("solid", fgColor="2F5597")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style="thin", color="999999"),
        right=Side(style="thin", color="999999"),
        top=Side(style="thin", color="999999"),
        bottom=Side(style="thin", color="999999"),
    )
    money_fmt = "#,##0.00 €"

    ws["A1"].value = f"Facturation PDJ — {month}"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:D1")

    # Synthèse
    start = 3
    ws.cell(row=start, column=1, value="Site")
    ws.cell(row=start, column=2, value="Total lignes (€/produits)")
    ws.cell(row=start, column=3, value="Ajustements (€)")
    ws.cell(row=start, column=4, value="Total facturé (€)")
    for c in range(1, 5):
        cell = ws.cell(row=start, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    if synth.empty:
        ws.cell(row=start + 1, column=1, value="(Aucune donnée)")
    else:
        for i, r in enumerate(synth.itertuples(index=False), start=1):
            row = start + i
            ws.cell(row=row, column=1, value=getattr(r, "site", ""))
            ws.cell(row=row, column=2, value=float(getattr(r, "total_eur", 0.0) or 0.0))
            ws.cell(row=row, column=3, value=float(getattr(r, "adj_eur", 0.0) or 0.0))
            ws.cell(row=row, column=4, value=float(getattr(r, "total_facture_eur", 0.0) or 0.0))
            for c in range(1, 5):
                cell = ws.cell(row=row, column=c)
                cell.border = border
                if c in (2, 3, 4):
                    cell.number_format = money_fmt

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18

    # Détail
    ws2 = wb.create_sheet("Détail lignes")
    cols = ["date", "site", "product", "qty", "unit", "unit_price", "amount_eur", "kind", "comment", "source"]
    ws2.append(cols)
    for c in range(1, len(cols) + 1):
        cell = ws2.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    if not detail.empty:
        for r in detail[cols].itertuples(index=False):
            ws2.append(list(r))
        for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=1, max_col=len(cols)):
            for cell in row:
                cell.border = border
                if cell.column_letter in {"F", "G"}:
                    cell.number_format = money_fmt
                if cell.column_letter == "D":
                    cell.number_format = "0.00"
    ws2.freeze_panes = "A2"
    widths = {
        "A": 12,
        "B": 20,
        "C": 28,
        "D": 10,
        "E": 10,
        "F": 12,
        "G": 14,
        "H": 12,
        "I": 28,
        "J": 22,
    }
    for k, w in widths.items():
        ws2.column_dimensions[k].width = w

    # Ajustements monétaires
    ws3 = wb.create_sheet("Ajustements €")
    ws3.append(["date", "site", "label", "amount_eur", "comment"])
    for c in range(1, 6):
        cell = ws3.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    if not adj.empty:
        for r in adj[["date", "site", "label", "amount_eur", "comment"]].itertuples(index=False):
            ws3.append(list(r))
        for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row, min_col=1, max_col=5):
            for cell in row:
                cell.border = border
                if cell.column_letter == "D":
                    cell.number_format = money_fmt
    ws3.column_dimensions["A"].width = 12
    ws3.column_dimensions["B"].width = 20
    ws3.column_dimensions["C"].width = 22
    ws3.column_dimensions["D"].width = 14
    ws3.column_dimensions["E"].width = 28

    # Tarifs
    ws4 = wb.create_sheet("Tarifs")
    ws4.append(["product", "site", "unit_price", "unit"])
    for c in range(1, 5):
        cell = ws4.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    if prices is not None and not prices.empty:
        for r in prices[["product", "site", "unit_price", "unit"]].itertuples(index=False):
            ws4.append(list(r))
        for row in ws4.iter_rows(min_row=2, max_row=ws4.max_row, min_col=1, max_col=4):
            for cell in row:
                cell.border = border
                if cell.column_letter == "C":
                    cell.number_format = money_fmt
    ws4.column_dimensions["A"].width = 32
    ws4.column_dimensions["B"].width = 18
    ws4.column_dimensions["C"].width = 14
    ws4.column_dimensions["D"].width = 10

    wb.save(out_path)

    # Sauvegarde persistante (historique) pour permettre suppression / traçabilité
    try:
        persist_dir = exports_for_month_dir(month)
        persist_path = persist_dir / f"Facturation_PDJ_{month}.xlsx"
        # On ne déplace pas le fichier temporaire (streamlit en a besoin) : on copie.
        persist_path.write_bytes(Path(out_path).read_bytes())
    except Exception:
        # On n'échoue pas l'export si la copie persistante ne marche pas.
        pass
    return out_path


# -----------------------------------------------------------------------------
# Export PDF simplifié : 1 facture par site (mois)
# -----------------------------------------------------------------------------


def _safe_filename(s: str) -> str:
    s = re.sub(r"[^A-Za-z0-9._ -]+", "_", str(s or "")).strip()
    s = re.sub(r"\s+", " ", s)
    return s or "site"


def build_site_invoice_tables(month: str, site: str) -> Tuple[pd.DataFrame, pd.DataFrame, float]:
    """Retourne (lignes_produits, ajustements, total_facture).

    lignes_produits: product, qty, unit, unit_price, line_total
    ajustements: label, amount_eur
    """
    synth, detail, adj = compute_monthly_pdj(month)
    siteN = norm_site_facturation(site)

    d = detail[detail["site"] == siteN].copy()
    if d.empty:
        lines = pd.DataFrame(columns=["product", "qty", "unit", "unit_price", "line_total"])
    else:
        lines = (
            d.groupby(["product", "unit", "unit_price"], as_index=False)["qty"].sum()
            .sort_values("product")
        )
        lines["line_total"] = lines["qty"].astype(float) * lines["unit_price"].astype(float)

    a = adj[adj["site"] == siteN].copy()
    if a.empty:
        adj_tbl = pd.DataFrame(columns=["label", "amount_eur"])
        adj_sum = 0.0
    else:
        adj_tbl = a.groupby("label", as_index=False)["amount_eur"].sum().sort_values("label")
        adj_sum = float(adj_tbl["amount_eur"].sum())

    total = float(lines["line_total"].sum() if not lines.empty else 0.0) + adj_sum
    return lines, adj_tbl, total


def export_site_invoice_pdf(month: str, site: str, out_path: str) -> str:
    """Génère un PDF A4 (simple, sans colonnes superflues) pour 1 site."""
    lines, adj_tbl, total = build_site_invoice_tables(month, site)
    siteN = norm_site_facturation(site)

    c = canvas.Canvas(out_path, pagesize=A4)
    w, h = A4
    left = 18 * mm
    right = w - 18 * mm
    y = h - 18 * mm

    def txt(s: str, size: int = 11, bold: bool = False):
        nonlocal y
        c.setFont("Helvetica-Bold" if bold else "Helvetica", size)
        c.drawString(left, y, s)
        y -= (size + 4)

    # En-tête
    txt(f"FACTURATION PDJ — {month}", size=16, bold=True)
    txt(f"Site : {siteN}", size=12, bold=True)
    txt(f"Généré le : {dt.date.today().strftime('%d/%m/%Y')}", size=10)
    y -= 4

    # Tableau
    # Colonnes (x) : on réserve assez d'espace pour éviter que "PU" et "Total" se chevauchent
    # quand les montants deviennent longs (ex: 1 470,85 €). Les anciennes positions (right-20)
    # étaient trop proches et donnaient un rendu "pâté".
    qty_x = right - 120
    pu_x = right - 65
    total_x = right

    c.setFont("Helvetica-Bold", 10)
    c.drawString(left, y, "Produit")
    c.drawRightString(qty_x, y, "Qté")
    c.drawRightString(pu_x, y, "PU")
    c.drawRightString(total_x, y, "Total")
    y -= 8
    c.line(left, y, right, y)
    y -= 12

    def money(v: float) -> str:
        return f"{v:,.2f} €".replace(",", " ").replace(".", ",")

    c.setFont("Helvetica", 10)
    if lines.empty:
        c.drawString(left, y, "(Aucune consommation enregistrée)")
        y -= 14
    else:
        for r in lines.itertuples(index=False):
            prod = str(getattr(r, "product", ""))
            qty = float(getattr(r, "qty", 0.0) or 0.0)
            pu = float(getattr(r, "unit_price", 0.0) or 0.0)
            lt = float(getattr(r, "line_total", 0.0) or 0.0)
            # Saut de page si besoin
            if y < 35 * mm:
                c.showPage()
                y = h - 18 * mm
                c.setFont("Helvetica", 10)
            c.drawString(left, y, prod[:65])
            c.drawRightString(qty_x, y, f"{qty:g}")
            c.drawRightString(pu_x, y, money(pu))
            c.drawRightString(total_x, y, money(lt))
            y -= 14

    # Ajustements
    if not adj_tbl.empty:
        y -= 6
        c.setFont("Helvetica-Bold", 10)
        c.drawString(left, y, "Ajustements")
        y -= 10
        c.setFont("Helvetica", 10)
        for r in adj_tbl.itertuples(index=False):
            if y < 35 * mm:
                c.showPage()
                y = h - 18 * mm
            lab = str(getattr(r, "label", "Ajustement"))
            amt = float(getattr(r, "amount_eur", 0.0) or 0.0)
            c.drawString(left, y, lab[:65])
            c.drawRightString(right, y, money(amt))
            y -= 14

    # Total
    y -= 4
    c.line(left, y, right, y)
    y -= 18
    c.setFont("Helvetica-Bold", 14)
    c.drawString(left, y, "TOTAL MENSUEL À RÉGLER")
    c.drawRightString(right, y, money(total))

    c.showPage()
    c.save()
    return out_path


def export_monthly_invoices_zip(month: str) -> bytes:
    """Retourne un ZIP en mémoire contenant 1 PDF par site présent sur le mois."""
    synth, detail, adj = compute_monthly_pdj(month)
    sites = set(detail["site"].dropna().astype(str).tolist()) | set(adj["site"].dropna().astype(str).tolist())
    sites = sorted([s for s in sites if str(s).strip()])

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as z:
        if not sites:
            # zip vide + README
            z.writestr("README.txt", f"Aucune donnée PDJ pour {month}.")
        for s in sites:
            pdf_bytes = io.BytesIO()
            # reportlab écrit sur un chemin ou file-like : on passe un buffer.
            export_site_invoice_pdf(month, s, pdf_bytes)  # type: ignore[arg-type]
            name = f"Facture_PDJ_{month}_{_safe_filename(s)}.pdf"
            z.writestr(name, pdf_bytes.getvalue())
    data = buf.getvalue()
    # Sauvegarde persistante (historique)
    try:
        persist_dir = exports_for_month_dir(month)
        persist_path = persist_dir / f"Factures_PDJ_{month}.zip"
        persist_path.write_bytes(data)
    except Exception:
        pass
    return data


# -----------------------------------------------------------------------------
# Import helpers (Excel)
# -----------------------------------------------------------------------------


def parse_pdj_excel(path_or_buffer) -> pd.DataFrame:
    """Tente d'extraire une table (product, qty) depuis un fichier Excel.

    Cette fonction est volontairement "tolérante" : elle essaie quelques heuristiques
    sans casser si le format varie.
    """
    def _try_read(*, header=None, engine=None) -> pd.DataFrame:
        try:
            return pd.read_excel(path_or_buffer, sheet_name=0, header=header, engine=engine)
        except Exception:
            return pd.DataFrame()

    # 1) Lecture standard (table avec en-têtes)
    df = _try_read(header=0)
    if df.empty:
        # Tentative xls via xlrd si dispo
        df = _try_read(header=0, engine="xlrd")

    if df is None or df.empty:
        # 2) Lecture "template" (bons PDJ avec cellules fusionnées / sans en-têtes utiles)
        df_raw = _try_read(header=None)
        if df_raw.empty:
            df_raw = _try_read(header=None, engine="xlrd")
        out2 = _parse_pdj_template_excel(df_raw)
        return out2

    # Normalise en string
    cols = [str(c) for c in df.columns]
    df.columns = cols

    # Cherche colonnes candidates
    def _find_col(patterns: List[str]) -> Optional[str]:
        for c in cols:
            cn = _norm(c)
            if any(p in cn for p in patterns):
                return c
        return None

    col_prod = _find_col(["produit", "libell", "designation", "article"]) or cols[0]
    col_qty = _find_col(["qt", "quant", "qte", "commande", "consomm"]) 

    # Si pas trouvé, essaye la 2e colonne numérique
    if col_qty is None and len(cols) >= 2:
        col_qty = cols[1]

    out = pd.DataFrame(
        {
            "product": df[col_prod].astype(str),
            "qty": pd.to_numeric(df[col_qty], errors="coerce") if col_qty in df.columns else 0,
        }
    )
    out["product"] = out["product"].map(norm_product)
    out["qty"] = pd.to_numeric(out["qty"], errors="coerce").fillna(0.0)
    out = out[out["product"].astype(str).str.strip() != ""].copy()
    out = out[out["qty"] != 0].copy()
    out = out[["product", "qty"]]
    # Si la lecture standard ne donne rien, tente le mode template.
    if out.empty:
        df_raw = _try_read(header=None)
        if df_raw.empty:
            df_raw = _try_read(header=None, engine="xlrd")
        return _parse_pdj_template_excel(df_raw)
    return out


def _parse_pdj_template_excel(df_raw: pd.DataFrame) -> pd.DataFrame:
    """Parse les bons PDJ "template" (souvent sans en-têtes tabulaires).

    On cherche la zone "Ingrédients à commander" puis on lit:
    - colonne A = produit
    - colonne B = quantité
    - colonne C = commentaire (optionnel)
    """
    if df_raw is None or df_raw.empty:
        return pd.DataFrame(columns=["product", "qty"])

    # Repère la ligne de départ
    start_row = None
    for i in range(len(df_raw)):
        row = " ".join([str(x) for x in df_raw.iloc[i].fillna("").tolist()])
        if "ingr" in row.lower() and "commander" in row.lower():
            start_row = i + 1
            break
    if start_row is None:
        # fallback : tente depuis le haut
        start_row = 0

    rows = []
    empty_streak = 0
    for i in range(start_row, len(df_raw)):
        prod = df_raw.iloc[i, 0] if df_raw.shape[1] > 0 else None
        qty = df_raw.iloc[i, 1] if df_raw.shape[1] > 1 else None

        prod_s = "" if pd.isna(prod) else str(prod).strip()
        if not prod_s and (qty is None or (isinstance(qty, float) and pd.isna(qty))):
            empty_streak += 1
            if empty_streak >= 6:
                break
            continue
        empty_streak = 0

        if prod_s.lower().startswith("autres"):
            break
        if not prod_s:
            continue

        q = pd.to_numeric(qty, errors="coerce")
        if pd.isna(q) or float(q) == 0.0:
            continue

        rows.append({"product": norm_product(prod_s), "qty": float(q)})

    if not rows:
        return pd.DataFrame(columns=["product", "qty"])
    return pd.DataFrame(rows)[["product", "qty"]]


def _detect_site_from_text(text: str) -> str:
    """Détection best-effort du site depuis un texte (Excel/PDF)."""
    t = re.sub(r"\s+", " ", str(text or "")).strip().lower()
    if not t:
        return ""
    # Quelques heuristiques simples (tu peux enrichir en ajoutant des mots-clés)
    if "toulouse" in t and "lautrec" in t:
        return "MAS TOULOUSE LAUTREC"
    if "rosa" in t and "bonheur" in t:
        return "24 ter"
    if "léonard" in t and "vinci" in t:
        return 'internat simple "Léonard de Vinci"'
    # IME ...
    m = re.search(r"\bime\s*([^\n\r]+)", t, flags=re.I)
    if m:
        return m.group(1).strip()
    return ""


def parse_pdj_document(path_or_buffer) -> Tuple[str, pd.DataFrame]:
    """Parse un bon PDJ (Excel/PDF) et renvoie (site_guess, df(product,qty)).

    Le résultat doit rester **modifiable manuellement** dans l'UI.
    """
    # Détection extension si on a un chemin
    site_guess = ""
    ext = ""
    try:
        if hasattr(path_or_buffer, "name"):
            ext = str(path_or_buffer.name).lower()
        elif isinstance(path_or_buffer, (str, Path)):
            ext = str(path_or_buffer).lower()
    except Exception:
        ext = ""

    if ext.endswith(".pdf"):
        df = parse_pdj_pdf(path_or_buffer)
        # Tente d'extraire du texte si possible
        try:
            import pdfplumber

            with pdfplumber.open(path_or_buffer) as pdf:
                text = "\n".join([p.extract_text() or "" for p in pdf.pages[:2]])
            site_guess = _detect_site_from_text(text)
        except Exception:
            site_guess = ""
        return site_guess, df
    else:
        # Excel
        df = parse_pdj_excel(path_or_buffer)
        # détection site dans les cellules hautes si possible
        try:
            df_raw = pd.read_excel(path_or_buffer, sheet_name=0, header=None)
            header_text = " ".join([str(x) for x in df_raw.head(25).stack().dropna().tolist()])
            site_guess = _detect_site_from_text(header_text)
        except Exception:
            site_guess = ""
        return site_guess, df


def parse_pdj_pdf(path_or_buffer) -> pd.DataFrame:
    """Extraction *best-effort* (product, qty) depuis un PDF.

    ⚠️ Beaucoup de bons PDJ sont scannés (image) : il n'existe pas de texte exploitable.
    On tente donc une OCR légère, puis on cherche les produits connus et un nombre
    présent sur la même ligne. Résultat à **vérifier/corriger** dans l'interface.
    """
    try:
        import pdfplumber
        import pytesseract
    except Exception:
        # Pas de dépendances : on ne bloque pas
        return pd.DataFrame(columns=["product", "qty"])

    # OCR page 1 (souvent suffisant)
    try:
        with pdfplumber.open(path_or_buffer) as pdf:
            if not pdf.pages:
                return pd.DataFrame(columns=["product", "qty"])
            img = pdf.pages[0].to_image(resolution=250).original
    except Exception:
        return pd.DataFrame(columns=["product", "qty"])

    # Pré-traitement image (contraste) pour améliorer l'OCR sur scans.
    # Si Pillow n'est pas dispo, on fait sans.
    try:
        from PIL import ImageOps, ImageFilter

        g = img.convert("L")
        g = ImageOps.autocontrast(g)
        # léger renforcement
        g = g.filter(ImageFilter.SHARPEN)
        img_for_ocr = g
    except Exception:
        img_for_ocr = img

    # OCR en français (meilleur pour vos bons)
    try:
        raw = pytesseract.image_to_string(img_for_ocr, lang="fra", config="--psm 6")
    except Exception:
        # fallback si 'fra' non installé
        try:
            raw = pytesseract.image_to_string(img_for_ocr, lang="eng", config="--psm 6")
        except Exception:
            return pd.DataFrame(columns=["product", "qty"])

    lines = [" ".join(l.strip().split()) for l in (raw or "").splitlines() if l.strip()]
    if not lines:
        return pd.DataFrame(columns=["product", "qty"])

    # Indexe par ligne normalisée
    ln_norm = [re.sub(r"[^a-z0-9' ]+", " ", l.lower()).strip() for l in lines]

    out_rows = []
    for prod in DEFAULT_PRODUCTS:
        p_norm = re.sub(r"[^a-z0-9' ]+", " ", prod.lower()).strip()
        # match approximatif: contient au moins 2 mots du produit
        tokens = [t for t in p_norm.split() if len(t) >= 3]
        if not tokens:
            continue
        best_i = None
        best_score = 0
        for i, l in enumerate(ln_norm):
            score = sum(1 for t in tokens if t in l)
            if score > best_score:
                best_score = score
                best_i = i
        if best_i is None or best_score < max(1, min(2, len(tokens))):
            continue

        original_line = lines[best_i]
        # Cherche un nombre (ex: 12, 12.5, 12,5) sur la ligne
        m = re.findall(r"-?\d+[\.,]?\d*", original_line)
        qty = 0.0
        if m:
            # prend le dernier nombre (souvent à droite)
            try:
                qty = float(m[-1].replace(",", "."))
            except Exception:
                qty = 0.0

        if qty != 0:
            out_rows.append({"product": norm_product(prod), "qty": qty})

    if not out_rows:
        return pd.DataFrame(columns=["product", "qty"])
    return pd.DataFrame(out_rows)[["product", "qty"]]

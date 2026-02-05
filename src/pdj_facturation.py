from __future__ import annotations

import io
import os
import re
import unicodedata
from dataclasses import dataclass
from typing import Dict, List, Tuple, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# PDF -> image
try:
    import fitz  # pymupdf
    PYMUPDF_AVAILABLE = True
except Exception:
    fitz = None
    PYMUPDF_AVAILABLE = False
from PIL import Image, ImageOps, ImageEnhance
# OCR (requis pour PDF scanné)
import pytesseract


# Sur Streamlit Cloud, tesseract est à /usr/bin/tesseract si tu as packages.txt
# En local Windows/Mac, ça dépend (PATH). On ne force QUE si le chemin existe.
try:
    if os.path.exists("/usr/bin/tesseract"):
        pytesseract.pytesseract.tesseract_cmd = "/usr/bin/tesseract"
except Exception:
    pass


# --- Configuration modèle économat (colonnes sites) ---
SITE_COL_NAMES = [
    "FAM TL",
    "Bussière",
    "Bruyères",
    "FO",
    "MAS",
    "ESAT",
    "FM",
    "René Lalouette",
    "Internat",
]

# Quelques normalisations de libellés (tu peux enrichir)
ALIASES = {
    "yaourt nature": "Yaourt",
    "lait demi": "Lait 1/2 écrémé",
    "lait demi-écrémé": "Lait 1/2 écrémé",
    "lait 1/2": "Lait 1/2 écrémé",
    "jus de pomme": "Jus Pomme",
    "jus pomme": "Jus Pomme",
    "jus d'orange": "Jus Orange",
    "jus orange": "Jus Orange",
    "jus de raisin": "Jus Raisin",
    "jus raisin": "Jus Raisin",
    "tablettes chocolat": "Chocolat",
    "nesquick": "Chocolat",
    "sucre en poudre": "Sucre",
    "sucre en poudre (sachets)": "Sucre",
}

# Stop-words / bruit OCR
BAD_TOKENS = {
    "quantité",
    "livré",
    "reçu",
    "commandé",
    "commande",
    "total",
    "produit",
    "unité",
    "prix",
}


@dataclass
class ParsedOrder:
    filename: str
    site: Optional[str]
    # liste (produit_normalisé, quantité_commandée)
    items: List[Tuple[str, float]]


# ----------------- Utilitaire lecture fichier (UploadedFile OU chemin str) -----------------

def _read_any_file(f) -> Tuple[str, bytes]:
    """
    Accepte:
    - UploadedFile Streamlit (read + name)
    - chemin str vers un fichier
    - file-like (read) sans name
    Retour: (filename, bytes)
    """
    # Cas: chemin string
    if isinstance(f, str):
        filename = os.path.basename(f)
        with open(f, "rb") as fh:
            return filename, fh.read()

    # Cas: UploadedFile / file-like
    filename = getattr(f, "name", "bon")
    if hasattr(f, "read"):
        return filename, f.read()

    raise TypeError(f"Type de fichier non supporté: {type(f)}")


# ----------------- Détection site -----------------

def _detect_site_from_text(text: str, filename: str = "") -> Optional[str]:
    t = (text or "").lower() + " " + (filename or "").lower()

    # Ajuste tes mots-clés ici
    if "toulouse lautrec" in t or "lautrec" in t:
        return "MAS"
    if "mas" in t and "toulouse" in t:
        return "MAS"
    # Internat: accepte plusieurs variantes OCR ("24 ter", "24T", "24ter", "24 simple", etc.)
    if (
        "internat" in t
        or "24t" in t
        or "24 t" in t
        or "24 ter" in t
        or "24ter" in t
        or "24 simple" in t
        or "24simple" in t
        or re.search(r"\b24\s*s\b", t) is not None
    ):
        return "Internat"

    return None


# ----------------- Normalisation produit -----------------

def _simplify_text(s: str) -> str:
    """Normalisation soft pour matching (minuscules, sans accents, ponctuation->espace)."""
    s = (s or "").lower()
    s = s.replace("’", "'")
    # enlever accents
    s = "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
    # ponctuation -> espace
    s = re.sub(r"[^a-z0-9]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def _normalize_product_name(raw: str) -> str:
    s = (raw or "").strip()
    s = re.sub(r"\s+", " ", s).strip()
    simp = _simplify_text(s)

    mappings = {
        "yaourt nature": "Yaourt",
        "yaourt": "Yaourt",
        "jus de pomme": "Jus Pomme",
        "jus pomme": "Jus Pomme",
        "jus d orange": "Jus Orange",
        "jus de orange": "Jus Orange",
        "jus orange": "Jus Orange",
        "jus d ananas": "Jus Ananas",
        "jus ananas": "Jus Ananas",
        "jus de raisin": "Jus Raisin",
        "jus raisin": "Jus Raisin",
        "tablettes chocolat": "Chocolat",
        "tablette chocolat": "Chocolat",
        "chocolat": "Chocolat",
        "sucre en poudre sachets": "Sucre",
        "sucre en poudre": "Sucre",
        "sucre": "Sucre",
        "compotes pruneaux": "Compotes",
        "compote pruneaux": "Compotes",
        "compotes": "Compotes",
        "beurre": "Micro-beurres",
        "micro beurres": "Micro-beurres",
        "biscotte": "Pain de mie complet",
        "pain de mie complet": "Pain de mie complet",
        "pain de mie": "Pain de mie complet",
        "cereales": "Céréales",
        "bledine": "Blédine",
        "gateaux": "Gateaux",
    }

    for k, v in mappings.items():
        if k in simp:
            return v

    # lait
    if "lait" in simp and ("1 2" in simp or "demi" in simp or "ecreme" in simp):
        return "Lait 1/2 écrémé"
    if "lait" in simp and "entier" in simp:
        return "Lait entier"

    return s


def _pdf_to_images(pdf_bytes: bytes, dpi_scale: float = 2.0) -> List[Image.Image]:
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    images: List[Image.Image] = []
    for page in doc:
        mat = fitz.Matrix(dpi_scale, dpi_scale)
        pix = page.get_pixmap(matrix=mat)
        img = Image.open(io.BytesIO(pix.tobytes("png")))
        images.append(img)
    return images


def _ocr_text_from_pdf_bytes(pdf_bytes: bytes) -> str:
    """OCR multi-pass pour PDF scannés (tableaux)."""
    images = _pdf_to_images(pdf_bytes, dpi_scale=3.0)
    out = []
    for img in images:
        try:
            out.append(pytesseract.image_to_string(img, lang="fra"))
        except Exception:
            pass
        try:
            w, h = img.size
            table = img.crop((0, int(h * 0.18), w, int(h * 0.95)))
            gray = ImageOps.grayscale(table)
            gray = ImageEnhance.Contrast(gray).enhance(3.0)
            gray = ImageEnhance.Sharpness(gray).enhance(2.0)
            bw = gray.point(lambda x: 0 if x < 175 else 255, "1")
            out.append(pytesseract.image_to_string(bw, lang="fra", config="--psm 6"))
            out.append(
                pytesseract.image_to_string(
                    bw,
                    lang="fra",
                    config="--psm 6 -c tessedit_char_whitelist=0123456789GOU,./",
                )
            )
        except Exception:
            pass
    return "\n".join(out)

def _parse_items_from_pdf_table(pdf_bytes: bytes) -> List[Tuple[str, float]]:
    """
    Lecture robuste du bon MAS (scanné) :
    - OCR avec positions (image_to_data) sur la zone tableau
    - reconstruction du libellé (colonne gauche) par lignes
    - association de la quantité (colonne droite) par proximité Y

    Cette approche est nettement plus fiable que l'OCR texte pour les quantités manuscrites.
    """
    images = _pdf_to_images(pdf_bytes, dpi_scale=3.4)
    if not images:
        return []

    img = images[0]
    w, h = img.size

    # Zone tableau (sans l'entête)
    y0 = int(h * 0.17)
    y1 = int(h * 0.93)
    x0 = int(w * 0.05)
    x1 = int(w * 0.96)

    table = img.crop((x0, y0, x1, y1))
    tw, th = table.size

    gray = ImageOps.grayscale(table)
    gray = ImageEnhance.Contrast(gray).enhance(3.0)
    gray = ImageEnhance.Sharpness(gray).enhance(2.0)
    bw = gray.point(lambda x: 0 if x < 175 else 255, "1")

    try:
        data = pytesseract.image_to_data(
            bw,
            lang="fra",
            config="--psm 6",
            output_type=pytesseract.Output.DICT,
        )
    except Exception:
        return []

    n = len(data.get("text", []))
    if n == 0:
        return []

    def _qty_from_token(tok: str) -> Optional[float]:
        t = (tok or "").strip()
        if not t:
            return None
        # "/" => vide
        if t in {"/", "|"}:
            return 0.0
        if t.upper() == "GO":
            return 30.0
        if t.upper() == "U":
            return 4.0
        if re.fullmatch(r"\d+", t):
            try:
                return float(t)
            except Exception:
                return None
        return None

    # Split gauche/droite (colonne quantité à droite)
    split_x = int(tw * 0.62)

    left_words = []   # (y, x, word)
    qty_tokens = []   # (y, x, qty)

    for i in range(n):
        txt = (data["text"][i] or "").strip()
        if not txt:
            continue
        try:
            conf = float(data.get("conf", [0] * n)[i] or 0)
        except Exception:
            conf = 0.0
        if conf < 35:
            continue

        x = int(data["left"][i])
        y = int(data["top"][i])
        w0 = int(data["width"][i])
        h0 = int(data["height"][i])
        xc = x + w0 / 2.0
        yc = y + h0 / 2.0

        if xc >= split_x:
            q = _qty_from_token(txt)
            if q is None:
                continue
            if q <= 0:
                continue
            qty_tokens.append((yc, xc, float(q)))
        else:
            low = txt.lower()
            if low in {"produit", "unite", "quantite", "commandee", "commande", "prix", "bon"}:
                continue
            left_words.append((yc, xc, txt))

    if not left_words or not qty_tokens:
        return []

    left_words.sort(key=lambda t: (t[0], t[1]))
    qty_tokens.sort(key=lambda t: t[0])

    # Regroupement des mots par lignes (proximité Y)
    y_thresh = max(10.0, th * 0.012)
    lines: List[Tuple[float, List[Tuple[float, str]]]] = []

    for yc, xc, word in left_words:
        placed = False
        for j in range(len(lines)):
            y_ref, arr = lines[j]
            if abs(yc - y_ref) <= y_thresh:
                arr.append((xc, word))
                lines[j] = (y_ref, arr)
                placed = True
                break
        if not placed:
            lines.append((yc, [(xc, word)]))

    items: List[Tuple[str, float]] = []

    for y_ref, arr in lines:
        arr.sort(key=lambda t: t[0])
        prod_raw = " ".join([w for _, w in arr]).strip()
        if not prod_raw:
            continue
        low = prod_raw.lower()
        if any(k in low for k in ["date", "livraison", "toulouse", "lautrec", "mas"]):
            continue

        # qty la plus proche sur la même ligne (Y)
        best_q = 0.0
        best_d = 1e9
        for yq, xq, q in qty_tokens:
            d = abs(yq - y_ref)
            if d < best_d and d <= (y_thresh * 1.6):
                best_d = d
                best_q = q

        if best_q > 0:
            items.append((_normalize_product_name(prod_raw), float(best_q)))

    # Dédoublonnage : garder le max par produit
    merged: Dict[str, float] = {}
    for p, q in items:
        merged[p] = max(merged.get(p, 0.0), float(q))

    return [(p, q) for p, q in merged.items()]


def _ocr_header_text_from_pdf_bytes(pdf_bytes: bytes) -> str:
    """OCR ciblé sur l'entête de la 1ère page (utile pour détecter le site sur scans pâles)."""
    try:
        images = _pdf_to_images(pdf_bytes, dpi_scale=3.0)
        if not images:
            return ""
        img = images[0]
        w, h = img.size
        header = img.crop((0, 0, w, int(h * 0.22)))

        gray = ImageOps.grayscale(header)
        gray = ImageEnhance.Contrast(gray).enhance(2.5)
        gray = ImageEnhance.Sharpness(gray).enhance(2.0)
        bw = gray.point(lambda x: 0 if x < 180 else 255, "1")

        return pytesseract.image_to_string(bw, lang="fra", config="--psm 6")
    except Exception:
        return ""



def _parse_items_from_ocr_text(text: str) -> List[Tuple[str, float]]:
    """Parse robuste pour bons PDJ scannés (tableau)."""
    items: List[Tuple[str, float]] = []
    lines = [l.strip() for l in (text or "").splitlines() if l.strip()]

    last_product: Optional[str] = None
    pending_qty: Optional[float] = None

    def _to_qty(token: str) -> Optional[float]:
        t = (token or "").strip()
        if not t:
            return None
        if t.upper() == "GO":
            return 30.0
        if t.upper() == "U":
            return 4.0
        if re.fullmatch(r"\d+(?:[.,]\d+)?", t):
            try:
                return float(t.replace(",", "."))
            except Exception:
                return None
        return None

    def _looks_like_weight(line_low: str) -> bool:
        return bool(re.search(r"\b\d+\s*(?:kg|g)\b", line_low))

    def _is_bad_product(name_low: str) -> bool:
        return any(x in name_low for x in ["sel", "poivre"])

    for l in lines:
        low = l.lower()
        if any(tok in low for tok in BAD_TOKENS):
            continue
        if _looks_like_weight(low):
            continue

        qty_only = _to_qty(l)
        if qty_only is not None:
            if last_product and not _is_bad_product(last_product.lower()):
                if 0 < qty_only <= 200:
                    items.append((_normalize_product_name(last_product), float(qty_only)))
                last_product = None
                pending_qty = None
            else:
                pending_qty = float(qty_only)
                last_product = None
            continue

        m = re.match(r"^(.*?)(?:\s+)([A-Za-z]{1,2}|\d+(?:[.,]\d+)?)\s*$", l)
        if m:
            name = m.group(1).strip(" -:\t")
            token = m.group(2).strip()
            qty = _to_qty(token)
            if name and qty is not None and 0 < qty <= 200:
                nlow = name.lower()
                if "date" not in nlow and "commande" not in nlow:
                    norm = _normalize_product_name(name)
                    if norm in {"Ketchup", "Mayonnaise", "Sel", "Poivre"} and 20 <= qty <= 31:
                        continue
                    items.append((norm, float(qty)))
                last_product = None
                pending_qty = None
                continue

        if not re.search(r"\d", l):
            if pending_qty is not None:
                norm = _normalize_product_name(l)
                if 0 < pending_qty <= 200:
                    items.append((norm, float(pending_qty)))
                pending_qty = None
                last_product = None
                continue
            if len(l) > 2 and not any(x in low for x in ["date", "commande", "livraison"]):
                last_product = l
            continue

    merged: Dict[str, float] = {}
    for p, q in items:
        if q <= 0:
            continue
        k = p.strip()
        if not k:
            continue
        merged[k] = max(merged.get(k, 0.0), float(q))
    return [(k, v) for k, v in merged.items()]


def _parse_excel(file_bytes: bytes, filename: str) -> ParsedOrder:
    bio = io.BytesIO(file_bytes)

    # 1) Lecture brute sans header pour pouvoir retrouver la vraie ligne d'en-tête
    try:
        df_raw = pd.read_excel(bio, header=None)
    except Exception:
        # fallback classique
        bio.seek(0)
        df_raw = pd.read_excel(bio)

    # Détection site à partir du nom + des premières cellules
    head_sample = ""
    try:
        sample_vals = []
        for r in range(min(8, len(df_raw))):
            row = df_raw.iloc[r].tolist()
            for v in row[:10]:
                if v is None:
                    continue
                s = str(v).strip()
                if s:
                    sample_vals.append(s)
        head_sample = " ".join(sample_vals)
    except Exception:
        head_sample = ""
    site = _detect_site_from_text(head_sample, filename)

    # 2) Trouver la ligne d'en-tête contenant "Produit"/"Désignation" etc.
    header_row_idx = None
    header_candidates = {"produit", "désignation", "designation", "article", "libellé", "libelle"}
    qty_candidates = {"qté", "qte", "quantité", "quantite", "commande", "commandé", "commandee"}

    try:
        for i in range(min(40, len(df_raw))):
            row = [str(x).strip().lower() for x in df_raw.iloc[i].tolist() if str(x).strip() != "nan"]
            if not row:
                continue
            if any(h in cell for cell in row for h in header_candidates) and any(
                q in cell for cell in row for q in qty_candidates
            ):
                header_row_idx = i
                break
    except Exception:
        header_row_idx = None

    items: List[Tuple[str, float]] = []

    if header_row_idx is not None:
        # Recharger avec cette ligne comme header
        bio.seek(0)
        df = pd.read_excel(bio, header=header_row_idx)

        # Détection site sur header si possible
        try:
            header_text = " ".join([str(c) for c in df.columns])
            site = site or _detect_site_from_text(header_text, filename)
        except Exception:
            pass

        col_prod = None
        for cand in ["Produit", "Désignation", "Designation", "Article", "Libellé", "Libelle"]:
            if cand in df.columns:
                col_prod = cand
                break

        col_qty = None
        for cand in ["Qté", "Qte", "Quantité", "Quantite", "Commande", "Commandé", "Commandee"]:
            if cand in df.columns:
                col_qty = cand
                break

        if col_prod and col_qty:
            for _, r in df.iterrows():
                prod_raw = str(r.get(col_prod, "")).strip()
                if not prod_raw:
                    continue
                low = prod_raw.lower()
                if any(tok in low for tok in BAD_TOKENS):
                    continue
                # quantité
                try:
                    qty = float(str(r.get(col_qty, "0")).replace(",", "."))
                except Exception:
                    continue
                if qty <= 0:
                    continue
                items.append((_normalize_product_name(prod_raw), float(qty)))
            return ParsedOrder(filename=filename, site=site, items=items)

    # 3) Fallback générique pour tableaux "Détail Déj-gouter 24T" :
    #    - produit = première colonne texte
    #    - quantité = somme des colonnes numériques de la ligne
    try:
        df = df_raw.copy()
        # repérer colonnes majoritairement numériques
        num_cols = []
        for c in df.columns:
            s = df[c]
            # convert to numeric where possible
            sn = pd.to_numeric(s, errors="coerce")
            if sn.notna().sum() >= max(3, int(0.2 * len(sn))):
                num_cols.append(c)

        text_cols = [c for c in df.columns if c not in num_cols]

        prod_col = text_cols[0] if text_cols else df.columns[0]

        for _, row in df.iterrows():
            prod_raw = row.get(prod_col, "")
            if prod_raw is None:
                continue
            prod_raw = str(prod_raw).strip()
            if not prod_raw:
                continue
            low = prod_raw.lower()
            if any(tok in low for tok in BAD_TOKENS):
                continue

            qty = 0.0
            for c in num_cols:
                v = row.get(c, 0)
                try:
                    fv = float(str(v).replace(",", "."))
                except Exception:
                    fv = 0.0
                if fv and not pd.isna(fv):
                    qty += fv
            if qty > 0:
                items.append((_normalize_product_name(prod_raw), float(qty)))
    except Exception:
        pass

    return ParsedOrder(filename=filename, site=site, items=items)

# ----------------- API attendue par app.py -----------------

def parse_pdj_order_file(uploaded_file) -> ParsedOrder:
    """
    Fonction attendue par ton app.py.
    Accepte UploadedFile Streamlit OU chemin str.
    """
    name, file_bytes = _read_any_file(uploaded_file)

    if name.lower().endswith(".pdf"):
        text = _ocr_text_from_pdf_bytes(file_bytes)
        site = _detect_site_from_text(text, name)

        if not site:
            header_text = _ocr_header_text_from_pdf_bytes(file_bytes)
            site = _detect_site_from_text(header_text, name)
        items = _parse_items_from_ocr_text(text)

        if site == "MAS":
            try:
                items_table = _parse_items_from_pdf_table(file_bytes)
                d: Dict[str, float] = {}
                for p, q in items + items_table:
                    if q <= 0:
                        continue
                    d[p] = max(d.get(p, 0.0), float(q))
                items = [(p, q) for p, q in d.items()]
            except Exception:
                pass
        return ParsedOrder(filename=name, site=site, items=items)

    if name.lower().endswith((".xls", ".xlsx", ".xlsm")):
        return _parse_excel(file_bytes, name)

    return ParsedOrder(filename=name, site=None, items=[])


def update_facturation_economat(
    modele_xlsx,
    order_files: List,
    mois: Optional[str] = None,
) -> bytes:
    """
    Fonction attendue par ton app.py.
    - modele_xlsx: UploadedFile Streamlit (xlsx) OU chemin str
    - order_files: liste de fichiers uploadés (pdf/xls/xlsx) OU chemins str
    - mois: 'YYYY-MM' optionnel
    Retour: bytes du fichier xlsx généré (pour st.download_button)
    """

    # 1) Charger modèle en openpyxl (conserver styles)
    if isinstance(modele_xlsx, str):
        wb = load_workbook(modele_xlsx)
    else:
        _, model_bytes = _read_any_file(modele_xlsx)
        wb = load_workbook(io.BytesIO(model_bytes))

    ws = wb.active

    # 2) Trouver ligne en-tête (celle qui contient "Produit")
    header_row = None
    for r in range(1, 50):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and v.strip().lower() == "produit":
            header_row = r
            break
    if header_row is None:
        header_row = 4

    # 3) Identifier colonnes
    col_by_name: Dict[str, int] = {}
    max_col = ws.max_column
    for c in range(1, max_col + 1):
        val = ws.cell(header_row, c).value
        if isinstance(val, str):
            col_by_name[val.strip()] = c

    col_prix = col_by_name.get("Prix unitaire (€)", 3)
    col_produit = col_by_name.get("Produit", 1)

    site_cols: Dict[str, int] = {}
    for s in SITE_COL_NAMES:
        if s in col_by_name:
            site_cols[s] = col_by_name[s]

    col_total_produit = col_by_name.get("Total produit (€)")
    if not col_total_produit:
        col_total_produit = ws.max_column

    # 4) Dictionnaire produit -> ligne
    product_row: Dict[str, int] = {}
    first_data_row = header_row + 1

    total_site_row = None
    for r in range(first_data_row, ws.max_row + 1):
        v = ws.cell(r, col_produit).value
        if isinstance(v, str) and "total site" in v.lower():
            total_site_row = r
            break

    last_product_row = (total_site_row - 1) if total_site_row else ws.max_row

    for r in range(first_data_row, last_product_row + 1):
        v = ws.cell(r, col_produit).value
        if isinstance(v, str) and v.strip():
            product_row[v.strip().lower()] = r

    product_row_simplified: Dict[str, int] = {}
    for k, r in product_row.items():
        product_row_simplified[_simplify_text(k)] = r

    # 5) Parser tous les bons + agréger (site, produit) = somme
    parsed: List[ParsedOrder] = [parse_pdj_order_file(f) for f in order_files]

    agg: Dict[Tuple[str, str], float] = {}
    for po in parsed:
        if not po.site:
            continue
        for prod, qty in po.items:
            if qty <= 0:
                continue
            key = (po.site, prod)
            agg[key] = agg.get(key, 0.0) + float(qty)

    def find_row_for_product(prod_name: str) -> Optional[int]:
        norm = _normalize_product_name(prod_name)
        low = norm.strip().lower()
        if low in product_row:
            return product_row[low]

        simp = _simplify_text(norm)
        if simp in product_row_simplified:
            return product_row_simplified[simp]

        # fallback contient / commence par
        for k, r in product_row.items():
            if low in k or k in low:
                return r
        for k, r in product_row_simplified.items():
            if simp in k or k in simp:
                return r
        return None

    # 6) Écrire les quantités (valeurs)
    for (site, prod), qty in agg.items():
        if site not in site_cols:
            continue
        r = find_row_for_product(_normalize_product_name(prod))
        if not r:
            # produit non trouvé => ignoré (pour ne pas casser le modèle)
            continue
        c = site_cols[site]
        cur = ws.cell(r, c).value
        try:
            cur_num = float(str(cur).replace(",", ".")) if cur not in (None, "") else 0.0
        except Exception:
            cur_num = 0.0
        ws.cell(r, c).value = cur_num + qty

    # 7) Écrire le mois si demandé
    if mois:
        for r in range(1, 15):
            for c in range(1, 6):
                v = ws.cell(r, c).value
                if isinstance(v, str) and "mois" in v.lower():
                    ws.cell(r, c + 1).value = mois
                    break

    # 8) Formules totaux (modifiable)
    site_col_indices = [site_cols[s] for s in SITE_COL_NAMES if s in site_cols]
    if site_col_indices:
        first_site_col = min(site_col_indices)
        last_site_col = max(site_col_indices)

        for r in range(first_data_row, last_product_row + 1):
            v = ws.cell(r, col_produit).value
            if not isinstance(v, str) or not v.strip():
                continue
            prix_cell = f"{get_column_letter(col_prix)}{r}"
            sum_range = (
                f"{get_column_letter(first_site_col)}{r}:{get_column_letter(last_site_col)}{r}"
            )
            ws.cell(r, col_total_produit).value = f"={prix_cell}*SUM({sum_range})"

        if total_site_row:
            prix_range = (
                f"{get_column_letter(col_prix)}{first_data_row}:{get_column_letter(col_prix)}{last_product_row}"
            )
            for s, c in site_cols.items():
                qty_range = (
                    f"{get_column_letter(c)}{first_data_row}:{get_column_letter(c)}{last_product_row}"
                )
                ws.cell(total_site_row, c).value = f"=SUMPRODUCT({prix_range},{qty_range})"

            total_range = (
                f"{get_column_letter(col_total_produit)}{first_data_row}:{get_column_letter(col_total_produit)}{last_product_row}"
            )
            ws.cell(total_site_row, col_total_produit).value = f"=SUM({total_range})"

    # 9) Retour bytes
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
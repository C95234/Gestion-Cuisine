
from __future__ import annotations

from pathlib import Path
from typing import Optional, Dict, Set, List

import pandas as pd
from openpyxl import load_workbook

from .config import ALLERGEN_COLUMNS
from .utils import normalize_key, normalize_space

ALLERGEN_LETTERS = list("CDEFGHIJKLMNOPQR")  # C -> R

def _truthy(v) -> bool:
    s = normalize_space(str(v or "")).strip().lower()
    return s in ("x", "1", "oui", "vrai", "true", "yes")



def read_overrides_sheet(path: str, sheet_name: str = "Overrides") -> pd.DataFrame:
    """Lit la feuille Overrides si elle existe.
    Format attendu (ligne 1 = en-têtes): Plat + colonnes allergènes.
    Valeurs: +1 (forcer), -1 (interdire), ou vide.
    """
    try:
        wb = load_workbook(path, data_only=True)
    except Exception:
        return pd.DataFrame(columns=["Plat"] + ALLERGEN_COLUMNS)

    ws = None
    for name in wb.sheetnames:
        if normalize_key(name) == normalize_key(sheet_name):
            ws = wb[name]
            break
    if ws is None:
        return pd.DataFrame(columns=["Plat"] + ALLERGEN_COLUMNS)

    # en-têtes
    header = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        header.append(normalize_space(str(v or "")).strip())

    # colonne plat
    plat_idx = None
    for i, h in enumerate(header):
        nk = normalize_key(h)
        if nk in ("plat", "plats", "denree", "produit", "libelle", "intitule"):
            plat_idx = i + 1
            break
    if plat_idx is None:
        plat_idx = 1

    norm_to_real = {normalize_key(x): x for x in ALLERGEN_COLUMNS}
    col_to_allergen = {}
    for i, h in enumerate(header):
        nk = normalize_key(h)
        if nk in norm_to_real:
            col_to_allergen[i + 1] = norm_to_real[nk]

    rows = []
    for r in range(2, ws.max_row + 1):
        dish = normalize_space(str(ws.cell(row=r, column=plat_idx).value or "")).strip()
        if not dish or dish.lower() == "nan":
            continue
        rec = {"Plat": dish}
        for a in ALLERGEN_COLUMNS:
            rec[a] = 0
        has_any = False
        for col_idx, allergen_name in col_to_allergen.items():
            v = ws.cell(row=r, column=col_idx).value
            if v is None or str(v).strip() == "":
                continue
            try:
                iv = int(float(str(v).strip()))
            except Exception:
                continue
            if iv not in (-1, 1):
                continue
            rec[allergen_name] = iv
            has_any = True
        if has_any:
            rows.append(rec)

    if not rows:
        return pd.DataFrame(columns=["Plat"] + ALLERGEN_COLUMNS)

    df = pd.DataFrame(rows)
    df["__k"] = df["Plat"].apply(normalize_key)
    out_rows = []
    for k, g in df.groupby("__k", dropna=False):
        if not k:
            continue
        base = {"Plat": g.iloc[-1]["Plat"]}
        for a in ALLERGEN_COLUMNS:
            vals = [int(x) for x in g[a].tolist() if int(x) in (-1, 1)]
            base[a] = vals[-1] if vals else 0
        out_rows.append(base)
    out = pd.DataFrame(out_rows)
    return out[["Plat"] + ALLERGEN_COLUMNS]

def apply_overrides_df(ref_df: pd.DataFrame, overrides_df: pd.DataFrame) -> pd.DataFrame:
    """Applique Overrides (-1/+1) sur un DataFrame de référence (X/vides)."""
    if ref_df is None or ref_df.empty or overrides_df is None or overrides_df.empty:
        return ref_df
    ref = ref_df.copy()
    ref["__k"] = ref["Plat"].apply(normalize_key)
    ov = overrides_df.copy()
    ov["__k"] = ov["Plat"].apply(normalize_key)
    ov_map = {k: ov[ov["__k"] == k].iloc[0] for k in ov["__k"].tolist() if k}
    for i, k in enumerate(ref["__k"].tolist()):
        if not k or k not in ov_map:
            continue
        row = ov_map[k]
        for a in ALLERGEN_COLUMNS:
            v = int(row.get(a, 0) or 0)
            if v == 1:
                ref.at[i, a] = "X"
            elif v == -1:
                ref.at[i, a] = ""
    return ref.drop(columns=["__k"], errors="ignore")

def update_overrides_from_corrections(
    effective_master_df: pd.DataFrame,
    filled_df: pd.DataFrame,
    overrides_df: pd.DataFrame,
) -> pd.DataFrame:
    """Met à jour Overrides en capturant les suppressions de X (et optionnellement les ajouts).
    Règle: si master(effectif) a X mais le classeur corrigé ne l'a plus => override -1.
    """
    if effective_master_df is None or effective_master_df.empty or filled_df is None or filled_df.empty:
        return overrides_df if overrides_df is not None else pd.DataFrame(columns=["Plat"] + ALLERGEN_COLUMNS)

    eff = effective_master_df.copy()
    eff["__k"] = eff["Plat"].apply(normalize_key)
    filled = filled_df.copy()
    filled["__k"] = filled["Plat"].apply(normalize_key)

    # map filled par clé
    filled_map = {k: filled[filled["__k"] == k].iloc[0] for k in filled["__k"].tolist() if k}

    ov = overrides_df.copy() if overrides_df is not None else pd.DataFrame(columns=["Plat"] + ALLERGEN_COLUMNS)
    for a in ALLERGEN_COLUMNS:
        if a not in ov.columns:
            ov[a] = 0
    if "Plat" not in ov.columns:
        ov["Plat"] = ""
    ov["__k"] = ov["Plat"].apply(normalize_key) if not ov.empty else []
    ov_idx = {k: i for i, k in enumerate(ov.get("__k", pd.Series([], dtype=str)).tolist()) if k}

    for _, row in eff.iterrows():
        k = row.get("__k", "")
        if not k or k not in filled_map:
            continue
        fr = filled_map[k]
        for a in ALLERGEN_COLUMNS:
            master_has = _truthy(row.get(a, ""))
            filled_has = _truthy(fr.get(a, ""))
            if master_has and not filled_has:
                # suppression => mémoriser -1
                if k in ov_idx:
                    i = ov_idx[k]
                    ov.at[i, a] = -1
                else:
                    new_row = {"Plat": row.get("Plat", fr.get("Plat", ""))}
                    for aa in ALLERGEN_COLUMNS:
                        new_row[aa] = 0
                    new_row[a] = -1
                    ov_idx[k] = len(ov)
                    ov.loc[len(ov)] = {**new_row, "__k": k}

    if ov is None or ov.empty:
        return pd.DataFrame(columns=["Plat"] + ALLERGEN_COLUMNS)
    ov = ov.drop(columns=["__k"], errors="ignore")
    # garder uniquement lignes avec au moins un override non nul
    mask = ov[ALLERGEN_COLUMNS].apply(lambda s: any(int(x) in (-1, 1) for x in s.tolist()), axis=1)
    ov = ov[mask].copy()
    return ov[["Plat"] + ALLERGEN_COLUMNS]

def extract_reference_from_filled_allergen_workbook(filled_path: str) -> pd.DataFrame:
    """Extrait un référentiel (Plat + colonnes allergènes) depuis un classeur 'Allergène ...' rempli (avec des X)."""
    wb = load_workbook(filled_path, data_only=True)
    rows: List[Dict[str, object]] = []

    for ws in wb.worksheets:
        # zone plats : B4:B27, allergènes : C4:R27
        for r in range(4, 28):
            dish = normalize_space(str(ws[f"B{r}"].value or "")).strip()
            if not dish or dish == "—":
                continue
            rec: Dict[str, object] = {"Plat": dish}
            for i, a in enumerate(ALLERGEN_COLUMNS):
                col = ALLERGEN_LETTERS[i]
                rec[a] = "X" if _truthy(ws[f"{col}{r}"].value) else ""
            rows.append(rec)

    if not rows:
        return pd.DataFrame(columns=["Plat"] + ALLERGEN_COLUMNS)

    df = pd.DataFrame(rows)
    # dédoublonne par plat normalisé, OR sur allergènes
    df["__k"] = df["Plat"].apply(normalize_key)
    agg = []
    for k, g in df.groupby("__k", dropna=False):
        if not k:
            continue
        base = {"Plat": g.iloc[0]["Plat"]}
        for a in ALLERGEN_COLUMNS:
            base[a] = "X" if any(_truthy(x) for x in g[a].tolist()) else ""
        agg.append(base)
    out = pd.DataFrame(agg)
    return out[["Plat"] + ALLERGEN_COLUMNS]



def write_reference_workbook(
    df: pd.DataFrame,
    out_path: str,
    title: str = "Référentiel allergènes",
    overrides_df: Optional[pd.DataFrame] = None,
) -> str:
    """Écrit un référentiel (Plat + colonnes allergènes) dans un classeur au format 'Allergène ...'.

    - Référence: plats en colonne B, X en C:R.
    - Optionnel: feuille 'Overrides' (Plat + colonnes allergènes en -1/+1) pour mémoriser tes corrections.
    """
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Référence"

    # Titre
    ws["A1"] = title

    # En-têtes (ligne 3) : A3 libellé, B3 vide, C3..R3 allergènes
    ws["A3"] = "Référence"
    ws["B3"] = None
    for i, a in enumerate(ALLERGEN_COLUMNS):
        col = ALLERGEN_LETTERS[i]
        ws[f"{col}3"] = a

    # Lignes de données à partir de 4
    if df is not None and not df.empty:
        df2 = df.copy()
        # assure colonnes attendues
        if "Plat" not in df2.columns:
            raise ValueError("Le DataFrame de référence doit contenir une colonne 'Plat'.")
        for a in ALLERGEN_COLUMNS:
            if a not in df2.columns:
                df2[a] = ""

        # Dé-doublonnage/tri stable
        df2["__k"] = df2["Plat"].apply(normalize_key)
        df2 = df2[df2["__k"].astype(bool)].copy()
        df2 = df2.sort_values("__k").drop_duplicates("__k", keep="first")

        r = 4
        for _, row in df2.iterrows():
            dish = normalize_space(str(row["Plat"] or "")).strip()
            if not dish:
                continue
            ws[f"B{r}"] = dish
            for i, a in enumerate(ALLERGEN_COLUMNS):
                col = ALLERGEN_LETTERS[i]
                ws[f"{col}{r}"] = "X" if _truthy(row.get(a, "")) else ""
            r += 1

    # Feuille Overrides (optionnelle)
    if overrides_df is not None and not overrides_df.empty:
        ovs = wb.create_sheet("Overrides")
        ovs["A1"] = "Plat"
        for j, a in enumerate(ALLERGEN_COLUMNS, start=2):
            ovs.cell(row=1, column=j).value = a

        ov = overrides_df.copy()
        if "Plat" not in ov.columns:
            return out_path
        for a in ALLERGEN_COLUMNS:
            if a not in ov.columns:
                ov[a] = 0
        ov["__k"] = ov["Plat"].apply(normalize_key)
        ov = ov[ov["__k"].astype(bool)].copy()
        ov = ov.sort_values("__k").drop_duplicates("__k", keep="last")

        r = 2
        for _, row in ov.iterrows():
            dish = normalize_space(str(row["Plat"] or "")).strip()
            if not dish:
                continue
            ovs.cell(row=r, column=1).value = dish
            for j, a in enumerate(ALLERGEN_COLUMNS, start=2):
                v = row.get(a, 0)
                try:
                    iv = int(v)
                except Exception:
                    iv = 0
                ovs.cell(row=r, column=j).value = iv if iv in (-1, 1) else ""
            r += 1

    wb.save(out_path)
    return out_path



def merge_reference(master_path: Optional[str], new_df: pd.DataFrame, out_path: str) -> str:
    """Fusionne new_df dans master (OR) et réécrit le référentiel au format classeur (plats en colonne B, X en C:R).

    En plus, si le master contient une feuille 'Overrides', on la conserve et on l'alimente :
    - si le master (effectif, overrides appliqués) avait 'X' mais que le classeur corrigé ne l'a plus,
      on mémorise un override -1 (interdire de pré-remplir la prochaine fois).
    """

    def _load_master_bundle(p: str) -> tuple[pd.DataFrame, pd.DataFrame]:
        # 1) Essai format 'plat en colonnes' (ancien) : pas d'overrides
        try:
            mdf = pd.read_excel(p)
            mdf.columns = [normalize_space(str(c or "")).strip() for c in mdf.columns]
            if "Plat" in mdf.columns:
                base = mdf[["Plat"] + [c for c in ALLERGEN_COLUMNS if c in mdf.columns]].copy()
                # assure toutes colonnes
                for a in ALLERGEN_COLUMNS:
                    if a not in base.columns:
                        base[a] = ""
                ov = pd.DataFrame(columns=["Plat"] + ALLERGEN_COLUMNS)
                return base[["Plat"] + ALLERGEN_COLUMNS], ov
        except Exception:
            pass

        # 2) Classeur au format 'Allergène ...' + overrides
        base = extract_reference_from_filled_allergen_workbook(p)
        ov = read_overrides_sheet(p)
        return base, ov

    if master_path and Path(master_path).exists():
        master_df, overrides_df = _load_master_bundle(master_path)
    else:
        master_df = pd.DataFrame(columns=["Plat"] + ALLERGEN_COLUMNS)
        overrides_df = pd.DataFrame(columns=["Plat"] + ALLERGEN_COLUMNS)

    # Normaliser colonnes du new_df
    new_df = new_df.copy() if new_df is not None else pd.DataFrame(columns=["Plat"] + ALLERGEN_COLUMNS)
    if "Plat" not in new_df.columns:
        new_df["Plat"] = ""
    for a in ALLERGEN_COLUMNS:
        if a not in new_df.columns:
            new_df[a] = ""

    # 1) Apprentissage des suppressions -> Overrides (-1)
    effective_master_df = apply_overrides_df(master_df, overrides_df)
    overrides_df = update_overrides_from_corrections(effective_master_df, new_df, overrides_df)

    # 2) Fusion OR dans le référentiel principal (on ne "désapprend" pas ici)
    if master_df is None or master_df.empty:
        merged_df = new_df.copy()
    else:
        master = master_df.copy()
        # assure colonnes
        if "Plat" not in master.columns:
            master["Plat"] = ""
        for a in ALLERGEN_COLUMNS:
            if a not in master.columns:
                master[a] = ""

        master["__k"] = master["Plat"].apply(normalize_key)
        new_df["__k"] = new_df["Plat"].apply(normalize_key)

        master_idx = {k: i for i, k in enumerate(master["__k"].tolist()) if k}

        for _, row in new_df.iterrows():
            k = row.get("__k", "")
            if not k:
                continue
            if k in master_idx:
                i = master_idx[k]
                for a in ALLERGEN_COLUMNS:
                    if _truthy(master.at[i, a]) or _truthy(row.get(a, "")):
                        master.at[i, a] = "X"
            else:
                master_idx[k] = len(master)
                master.loc[len(master)] = {**{c: row.get(c, "") for c in ["Plat"] + ALLERGEN_COLUMNS}, "__k": k}

        merged_df = master.drop(columns=["__k"], errors="ignore")

    # Réécrit au format attendu, en conservant la feuille Overrides
    return write_reference_workbook(merged_df[["Plat"] + ALLERGEN_COLUMNS], out_path, overrides_df=overrides_df)

def learn_from_filled_allergen_workbook(
    filled_path: str,
    master_path: Optional[str],
    out_master_path: str,
) -> str:
    """Point d'entrée: lit le classeur rempli et met à jour le référentiel maître."""
    new_df = extract_reference_from_filled_allergen_workbook(filled_path)
    return merge_reference(master_path, new_df, out_master_path)

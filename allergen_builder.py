from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

REGIMES = [
    "Standards",
    "Végétariens",
    "Hypocaloriques",
    "Sans lactose",
    "Spéciaux av lactose"
]

ALLERGENES = [
    "Céréales/gluten",
    "Crustacés",
    "Mollusques",
    "Poisson",
    "Œuf",
    "Arachide",
    "Soja",
    "Lactose"
]


def generate_allergen_sheet(date_str, repas_type):
    wb = Workbook()
    ws = wb.active
    ws.title = "Allergènes"

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = f"{repas_type} {date_str}"
    ws["A1"].font = Font(size=14, bold=True)

    ws["A3"] = repas_type
    ws["A3"].font = Font(bold=True)

    for col, allergene in enumerate(ALLERGENES, start=3):
        cell = ws.cell(row=3, column=col)
        cell.value = allergene
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    current_row = 4

    for regime in REGIMES:
        ws.cell(row=current_row, column=1).value = regime
        ws.cell(row=current_row, column=1).font = Font(bold=True)

        for i in range(4):
            ws.cell(row=current_row, column=2).value = ""
            for col in range(3, 3 + len(ALLERGENES)):
                ws.cell(row=current_row, column=col).border = border
            current_row += 1

    current_row += 1
    ws.cell(row=current_row, column=1).value = "*G6PD (élargi) : artichauts, asperges, cannelle, lentilles, champignons, figues, genièvre, haricots, petits pois, pois chiches, quinquina, verveine, carottes, tomates"

    current_row += 2
    ws.cell(row=current_row, column=1).value = "Origine des viandes"
    ws.cell(row=current_row + 1, column=1).value = "Viande - Plat"
    ws.cell(row=current_row + 1, column=3).value = "Lieu de naissance"
    ws.cell(row=current_row + 1, column=5).value = "Lieu d’élevage"
    ws.cell(row=current_row + 1, column=7).value = "Lieu d’abattage"

    return wb

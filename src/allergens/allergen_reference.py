from __future__ import annotations

from dataclasses import dataclass
from typing import Dict, Set, Tuple, Optional, Iterable
import difflib

import pandas as pd

from .config import ALLERGEN_COLUMNS
from .utils import normalize_key, normalize_space

@dataclass
class AllergenRef:
    # key normalisée du plat -> set d'allergènes (libellés exacts ALLERGEN_COLUMNS)
    map_key_to_allergens: Dict[str, Set[str]]

    def lookup(self, dish_name: str) -> Tuple[Set[str], Optional[str]]:
        """Retourne (allergènes, clé_trouvée)."""
        k = normalize_key(dish_name)
        if not k:
            return set(), None
        if k in self.map_key_to_allergens:
            return set(self.map_key_to_allergens[k]), k

        # fuzzy match (tolérant aux petites différences)
        keys = list(self.map_key_to_allergens.keys())
        cand = difflib.get_close_matches(k, keys, n=1, cutoff=0.88)
        if cand:
            ck = cand[0]
            return set(self.map_key_to_allergens[ck]), ck
        return set(), None



def _truthy(v) -> bool:
    s = normalize_space(str(v or "")).strip().lower()
    return s in ("x", "1", "oui", "vrai", "true", "yes", "y")

def _extract_reference_from_allergen_workbook(path: str) -> pd.DataFrame:
    """Lit un classeur au format 'Allergène ...' (plats en colonne B, X en C:R, zone B4:R27 sur chaque feuille).
    Retourne un DataFrame: Plat + ALLERGEN_COLUMNS (valeurs 'X' ou '').
    """
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    rows = []
    letters = list("CDEFGHIJKLMNOPQR")  # 16 colonnes
    for ws in wb.worksheets:
        for r in range(4, 28):
            dish = normalize_space(str(ws[f"B{r}"].value or "")).strip()
            if not dish or dish == "—":
                continue
            rec = {"Plat": dish}
            for i, a in enumerate(ALLERGEN_COLUMNS):
                col = letters[i]
                rec[a] = "X" if _truthy(ws[f"{col}{r}"].value) else ""
            rows.append(rec)

    if not rows:
        return pd.DataFrame(columns=["Plat"] + ALLERGEN_COLUMNS)

    df = pd.DataFrame(rows)
    df["__k"] = df["Plat"].apply(normalize_key)
    agg = []
    for k, g in df.groupby("__k", dropna=False):
        if not k:
            continue
        base = {"Plat": g.iloc[0]["Plat"]}
        for a in ALLERGEN_COLUMNS:
            base[a] = "X" if any(_truthy(x) for x in g[a].tolist()) else ""
        agg.append(base)
    out = pd.DataFrame(agg)
    return out[["Plat"] + ALLERGEN_COLUMNS]

def _read_overrides_sheet(path: str, sheet_name: str = "Overrides") -> pd.DataFrame:
    """Lit la feuille Overrides si elle existe.
    Format attendu (ligne 1 = en-têtes): Plat + colonnes allergènes.
    Valeurs: +1 (forcer), -1 (interdire), ou vide.
    """
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    ws = None
    for name in wb.sheetnames:
        if normalize_key(name) == normalize_key(sheet_name):
            ws = wb[name]
            break
    if ws is None:
        return pd.DataFrame(columns=["Plat"] + ALLERGEN_COLUMNS)

    # lire en-têtes
    header = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        header.append(normalize_space(str(v or "")).strip())
    # trouver colonne plat
    plat_idx = None
    for i, h in enumerate(header):
        nk = normalize_key(h)
        if nk in ("plat", "plats", "denree", "produit", "libelle", "intitule"):
            plat_idx = i + 1
            break
    if plat_idx is None:
        plat_idx = 1

    # mapper allergènes
    norm_to_real = {normalize_key(x): x for x in ALLERGEN_COLUMNS}
    col_to_allergen = {}
    for i, h in enumerate(header):
        nk = normalize_key(h)
        if nk in norm_to_real:
            col_to_allergen[i + 1] = norm_to_real[nk]

    rows = []
    for r in range(2, ws.max_row + 1):
        dish = normalize_space(str(ws.cell(row=r, column=plat_idx).value or "")).strip()
        if not dish or dish.lower() == "nan":
            continue
        rec = {"Plat": dish}
        for a in ALLERGEN_COLUMNS:
            rec[a] = 0
        has_any = False
        for col_idx, allergen_name in col_to_allergen.items():
            v = ws.cell(row=r, column=col_idx).value
            if v is None or str(v).strip() == "":
                continue
            try:
                iv = int(float(str(v).strip()))
            except Exception:
                continue
            if iv not in (-1, 1):
                continue
            rec[allergen_name] = iv
            has_any = True
        if has_any:
            rows.append(rec)

    if not rows:
        return pd.DataFrame(columns=["Plat"] + ALLERGEN_COLUMNS)
    df = pd.DataFrame(rows)
    df["__k"] = df["Plat"].apply(normalize_key)
    # dédoublonnage: on garde la dernière occurrence non nulle par allergène
    out_rows = []
    for k, g in df.groupby("__k", dropna=False):
        if not k:
            continue
        base = {"Plat": g.iloc[-1]["Plat"]}
        for a in ALLERGEN_COLUMNS:
            # dernier non-zero sinon 0
            vals = [int(x) for x in g[a].tolist() if int(x) in (-1, 1)]
            base[a] = vals[-1] if vals else 0
        out_rows.append(base)
    out = pd.DataFrame(out_rows)
    return out[["Plat"] + ALLERGEN_COLUMNS]

def _apply_overrides(ref_df: pd.DataFrame, overrides_df: pd.DataFrame) -> pd.DataFrame:
    if ref_df is None or ref_df.empty or overrides_df is None or overrides_df.empty:
        return ref_df
    ref = ref_df.copy()
    ref["__k"] = ref["Plat"].apply(normalize_key)
    ov = overrides_df.copy()
    ov["__k"] = ov["Plat"].apply(normalize_key)
    ov_map = {k: ov[ov["__k"] == k].iloc[0] for k in ov["__k"].tolist() if k}
    for i, k in enumerate(ref["__k"].tolist()):
        if not k or k not in ov_map:
            continue
        row = ov_map[k]
        for a in ALLERGEN_COLUMNS:
            v = int(row.get(a, 0) or 0)
            if v == 1:
                ref.at[i, a] = "X"
            elif v == -1:
                ref.at[i, a] = ""
    return ref.drop(columns=["__k"], errors="ignore")

def load_allergen_reference(path: str) -> AllergenRef:
    """Lit un référentiel allergènes Excel.

    Supporte 2 formats :
      1) Tableau classique (colonnes) : une colonne Plat + colonnes allergènes
      2) Classeur au format 'Allergène ...' (plats en B4:B27 et X en C:R27, 1 feuille par service)
         + optionnel : une feuille 'Overrides' (Plat + colonnes allergènes en -1/+1) pour mémoriser tes corrections.

    Le parsing menu et la structure du tableau généré ne changent pas : l'intelligence est appliquée via Overrides.
    """
    if path.lower().endswith('.ods'):
        raise ValueError(
            "Le référentiel allergènes doit être un fichier Excel (.xlsx). "
            "Merci de convertir ton .ods en .xlsx (Fichier > Enregistrer sous)."
        )

    ref_df = None

    # 1) Essai format tableau classique
    try:
        df = pd.read_excel(path)
        if df is not None and not df.empty:
            cols = {c: normalize_key(str(c)) for c in df.columns}
            plat_col = None
            for c, nk in cols.items():
                if nk in ('plat', 'plats', 'denree', 'denree produit', 'produit', 'libelle', 'intitule'):
                    plat_col = c
                    break
            if plat_col is None:
                plat_col = df.columns[0]

            norm_to_real = {normalize_key(x): x for x in ALLERGEN_COLUMNS}
            col_map = {}
            for c in df.columns:
                nk = normalize_key(str(c))
                if nk in norm_to_real:
                    col_map[c] = norm_to_real[nk]
            if len(col_map) < len(ALLERGEN_COLUMNS):
                for c in df.columns:
                    nk = normalize_key(str(c))
                    for tgt in ALLERGEN_COLUMNS:
                        tnk = normalize_key(tgt)
                        if tnk and (tnk in nk or nk in tnk) and c not in col_map:
                            col_map[c] = tgt

            # si on a au moins 1 allergène reconnu, on considère que c'est le bon format
            if col_map:
                rows = []
                for _, row in df.iterrows():
                    dish = normalize_space(str(row.get(plat_col, '') or ''))
                    if not dish or dish.lower() == 'nan':
                        continue
                    rec = {'Plat': dish}
                    for a in ALLERGEN_COLUMNS:
                        rec[a] = ''
                    for c, allergen_name in col_map.items():
                        if _truthy(row.get(c, '')):
                            rec[allergen_name] = 'X'
                    rows.append(rec)
                if rows:
                    ref_df = pd.DataFrame(rows)[['Plat'] + ALLERGEN_COLUMNS]
    except Exception:
        ref_df = None

    # 2) Sinon, format 'Allergène ...'
    if ref_df is None:
        ref_df = _extract_reference_from_allergen_workbook(path)

    # Overrides (optionnel)
    overrides_df = _read_overrides_sheet(path)
    ref_df = _apply_overrides(ref_df, overrides_df)

    # Construire map_key -> set(allergènes)
    map_key_to_allergens: Dict[str, Set[str]] = {}
    if ref_df is not None and not ref_df.empty:
        ref_df = ref_df.copy()
        ref_df['__k'] = ref_df['Plat'].apply(normalize_key)
        for _, row in ref_df.iterrows():
            k = row.get('__k', '')
            if not k:
                continue
            aset = map_key_to_allergens.get(k, set())
            for a in ALLERGEN_COLUMNS:
                if _truthy(row.get(a, '')):
                    aset.add(a)
            map_key_to_allergens[k] = aset

    return AllergenRef(map_key_to_allergens=map_key_to_allergens)

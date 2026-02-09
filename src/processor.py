from __future__ import annotations

# Pare-feu: évite NameError si un bloc se retrouve au niveau module
headers = {}

from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple
import re
import unicodedata
import datetime as dt

import pandas as pd

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# PDF (bons de livraison)
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

DAY_NAMES = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"]


# --- Définitions défensives pour éviter des NameError à l'import ---
# Certains environnements (ou merges/patchs) ont pu laisser du code exécuté au chargement du module.
# On initialise donc des variables globales utilisées par sécurité; elles seront recalculées dans les fonctions.
headers: Dict[str, int] = {}
col_eff = col_coef = col_unit = col_sup = col_qty = None
col_pu = col_pt = col_wu = col_wt = None
ws_bc = None


def clean_text(x) -> str:
    """Normalize cell content into a clean string."""
    if x is None:
        return ""
    s = str(x)
    s = s.replace("\u2026", "...")  # ellipsis
    s = s.replace("*", "")
    s = re.sub(r"\s+", " ", s).strip()
    s = s.strip(" -;,:")
    return s


def normalize_regime_label(regime: str) -> str:
    """Normalise les libellés de régimes pour éviter les confusions (casse/accents ignorés).

    - ss / sans / SANS ... -> "SANS"
    - spéciaux / speciaux / spécial / special ... -> "SPÉCIAL"
    - sinon: libellé original nettoyé (clean_text)
    """
    raw = clean_text(regime)
    if not raw:
        return ""
    s = raw.strip().lower()
    s = "".join(ch for ch in unicodedata.normalize("NFKD", s) if not unicodedata.combining(ch))
    s = re.sub(r"[^a-z0-9 ]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()

    if re.search(r"\b(ss|sans)\b", s):
        return "SANS"
    if re.search(r"\b(hypo|hypocal)\b", s):
        return "Hypocaloriques"
    if "special" in s or re.search(r"\b(speciaux|special|speciale|speciales)\b", s):
        return "SPÉCIAL"

    return raw


def _parse_date(x, default_year: Optional[int] = None) -> Optional[dt.date]:
    if x is None:
        return None

    if isinstance(x, dt.datetime):
        return x.date()
    if isinstance(x, dt.date):
        return x

    # Excel serial date (int/float) -> uniquement si plausible (évite 18278 = 1950)
    if isinstance(x, (int, float)):
        v = float(x)
        # plage "raisonnable" (≈ 2015–2035) pour éviter des décennies de colonnes
        if not (41000 <= v <= 51000):
            return None
        try:
            base = dt.date(1899, 12, 30)
            return base + dt.timedelta(days=int(v))
        except Exception:
            return None

    if isinstance(x, str):
        s = x.strip()
        if not s:
            return None

        s = s.replace("-", "/").replace(".", "/")
        m = re.match(r"^(\d{1,2})/(\d{1,2})(?:/(\d{2,4}))?$", s)
        if m:
            d = int(m.group(1))
            mo = int(m.group(2))
            y = m.group(3)
            if y is None:
                y = default_year if default_year is not None else dt.date.today().year
            else:
                y = int(y)
                if y < 100:
                    y += 2000
            try:
                return dt.date(int(y), mo, d)
            except Exception:
                return None

    return None


def is_date_cell(x) -> bool:
    return _parse_date(x) is not None


def _to_number(x) -> float:
    """Robust numeric conversion."""
    if x is None:
        return 0.0
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip()
    if not s:
        return 0.0
    # handle commas
    s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return 0.0


def _norm(s: str) -> str:
    s = (s or "").lower()
    s = (
        s.replace("é", "e")
        .replace("è", "e")
        .replace("ê", "e")
        .replace("î", "i")
        .replace("ï", "i")
        .replace("ô", "o")
        .replace("à", "a")
        .replace("ç", "c")
    )
    s = re.sub(r"[^a-z0-9 ]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _find_sheet_name(sheetnames: list[str], wanted: str, aliases: Optional[list[str]] = None) -> Optional[str]:
    """Find a worksheet name in a tolerant way.

    Users sometimes rename Excel tabs (e.g. add week numbers or extra words), which can
    break strict `sheet_name` checks. This helper tries to match the requested sheet using:
    - exact match
    - normalized match (case/accents/punctuation/extra spaces)
    - token-subset match (e.g. wanted="planning fab" matches "Planning fab semaine 07")
    - optional aliases
    """
    aliases = aliases or []

    if not sheetnames:
        return None

    # 1) Exact
    if wanted in sheetnames:
        return wanted

    wn = _norm(wanted)
    wn_compact = wn.replace(" ", "")
    wtoks = set(wn.split())

    # 2) Normalized exact / compact
    for name in sheetnames:
        nn = _norm(name)
        if nn == wn or nn.replace(" ", "") == wn_compact:
            return name

    # 3) Token subset
    if wtoks:
        for name in sheetnames:
            ntoks = set(_norm(name).split())
            if wtoks.issubset(ntoks):
                return name

    # 4) Aliases (same logic)
    for alias in aliases:
        if alias in sheetnames:
            return alias
        an = _norm(alias)
        an_compact = an.replace(" ", "")
        atoks = set(an.split())
        for name in sheetnames:
            nn = _norm(name)
            if nn == an or nn.replace(" ", "") == an_compact:
                return name
        if atoks:
            for name in sheetnames:
                ntoks = set(_norm(name).split())
                if atoks.issubset(ntoks):
                    return name

    return None


def parse_planning_fabrication(path: str, sheet_name: str = "PLANNING FAB") -> Dict[str, pd.DataFrame]:
    """
    Planning fabrication -> {"dejeuner": df, "diner": df}
    (fonction conservée telle quelle dans ton projet; tronquée ici si tu avais une autre version)
    """
    wb_val = openpyxl.load_workbook(path, data_only=True)   # valeurs figées
    wb_fx  = openpyxl.load_workbook(path, data_only=False)  # formules
    real_sheet = sheet_name
    if real_sheet not in wb_fx.sheetnames:
        real_sheet = _find_sheet_name(wb_fx.sheetnames, sheet_name, aliases=["PLANNING FABRICATION", "PLANNING FAB."])
    if not real_sheet or real_sheet not in wb_fx.sheetnames:
        raise ValueError(
            f"Feuille '{sheet_name}' introuvable (même en recherche tolérante). Feuilles dispo: {wb_fx.sheetnames}"
        )
    ws_val = wb_val[real_sheet]
    ws_fx  = wb_fx[real_sheet]

    targets = {_norm(d) for d in DAY_NAMES}

    _ref_re = re.compile(
        r"^=\s*(?:\+)?\s*(?:'([^']+)'\s*|([^'!]+))!(\$?[A-Z]{1,3}\$?\d{1,7})\s*$"
    )
    _ref_token_re = re.compile(
        r"(?:'([^']+)'\s*|([^'!+\-*/(),;]+))!(\$?[A-Z]{1,3}\$?\d{1,7})"
    )

    def _norm_cell(x) -> str:
        return _norm(clean_text(x))

    def _strip_dollars(addr: str) -> str:
        return addr.replace("$", "")

    def _resolve_ref(sheet: str, addr: str, depth: int = 0):
        if depth > 25:
            return None
        if sheet not in wb_fx.sheetnames:
            return None
        addr0 = _strip_dollars(addr)
        wsV = wb_val[sheet] if sheet in wb_val.sheetnames else None
        wsF = wb_fx[sheet]

        if wsV is not None:
            v = wsV[addr0].value
            if v is not None:
                return v

        v2 = wsF[addr0].value
        if isinstance(v2, str) and v2.startswith("="):
            return _eval_formula(v2, depth + 1)
        return v2

    def _eval_formula(formula: str, depth: int = 0):
        if depth > 25:
            return None
        if not isinstance(formula, str) or not formula.startswith("="):
            return formula

        f = formula.strip()

        m = _ref_re.match(f)
        if m:
            sheet = m.group(1) or m.group(2)
            addr = m.group(3)
            return _resolve_ref(sheet.strip(), addr, depth + 1)

        f_up = f.upper().replace(" ", "")
        if f_up.startswith("=SOMME(") or f_up.startswith("=SUM("):
            inside = f[f.find("(") + 1: f.rfind(")")]
            parts = re.split(r"[;,]", inside)
            total = 0.0
            for part in parts:
                part = part.strip()
                if not part:
                    continue
                mt = _ref_token_re.search(part)
                if mt:
                    sh = mt.group(1) or mt.group(2)
                    ad = mt.group(3)
                    total += _to_number(_resolve_ref(sh.strip(), ad, depth + 1))
                else:
                    total += _to_number(part)
            return total

        if any(op in f for op in ["+", "-"]) and "(" not in f and ")" not in f:
            expr = f[1:]
            tokens = re.split(r"(\+|\-)", expr)
            total = 0.0
            sign = +1.0
            first = True
            for t in tokens:
                t = t.strip()
                if not t:
                    continue
                if t == "+":
                    sign = +1.0
                    continue
                if t == "-":
                    sign = -1.0
                    continue

                mt = _ref_token_re.search(t)
                if mt:
                    sh = mt.group(1) or mt.group(2)
                    ad = mt.group(3)
                    val = _to_number(_resolve_ref(sh.strip(), ad, depth + 1))
                else:
                    val = _to_number(t)

                if first:
                    total = val
                    first = False
                else:
                    total = total + sign * val
            return total

        return None

    def get_value(r: int, c: int):
        if not c:
            return None
        v = ws_val.cell(r, c).value
        if v is not None:
            return v
        fx = ws_fx.cell(r, c).value
        if isinstance(fx, str) and fx.startswith("="):
            return _eval_formula(fx, 0)
        return fx

    dejeuner_row = None
    diner_row = None
    for r in range(1, ws_fx.max_row + 1):
        v = ws_fx.cell(r, 1).value
        if isinstance(v, str):
            vv = v.upper()
            if "PLANNING" in vv and "FABRICATION" in vv:
                # Accept multiple variants used in the Excel files:
                # - "DÉJEUNER" / "DEJEUN"
                # - abbreviation "DEJ" (often in titles like "—  DEJ —")
                # - "DÎNER" / "DINER"
                # - abbreviation "DIN" (often in titles like "—  DIN —")
                if ("DEJEUN" in vv) or re.search(r"\bDEJ\b", vv):
                    dejeuner_row = r
                if ("DINER" in vv) or ("DÎNER" in vv) or re.search(r"\bDIN\b", vv):
                    diner_row = r
    if dejeuner_row is None or diner_row is None:
        raise ValueError("Impossible de localiser les sections DÉJEUNER / DÎNER (colonne A).")

    def _looks_like_days_header(r: int) -> bool:
        hits = 0
        for c in range(1, ws_fx.max_column + 1):
            if _norm_cell(ws_fx.cell(r, c).value) in targets:
                hits += 1
        return hits >= 3

    def _find_header_row_near(title_row: int) -> Optional[int]:
        for r in range(title_row + 1, min(ws_fx.max_row, title_row + 10) + 1):
            if _looks_like_days_header(r):
                return r
        return None

    def _get_day_columns(header_row: int) -> Dict[str, int]:
        cols: Dict[str, int] = {}
        for c in range(1, ws_fx.max_column + 1):
            v = _norm_cell(ws_fx.cell(header_row, c).value)
            for d in DAY_NAMES:
                if v == _norm(d):
                    cols[d] = c
        return cols

    def _find_total_col(header_row: int) -> Optional[int]:
        for c in range(1, ws_fx.max_column + 1):
            v = _norm_cell(ws_fx.cell(header_row, c).value)
            if v == "total" or "total" in v:
                return c
        return None

    def read_block(title_row: int, end_row: int) -> pd.DataFrame:
        header_r = _find_header_row_near(title_row)

        if header_r is None:
            day_cols = {d: 3 + i for i, d in enumerate(DAY_NAMES)}
            total_col = 10
            data_start = title_row + 1
        else:
            day_cols = _get_day_columns(header_r)
            total_col = _find_total_col(header_r)
            data_start = header_r + 1

        rows = []
        current_site = None

        for r in range(data_start, end_row + 1):
            rowA = ws_fx.cell(r, 1).value

            if isinstance(rowA, str):
                up = rowA.upper()
                if "PLANNING" in up and "FABRICATION" in up:
                    break

            row_vals = [ws_fx.cell(r, c).value for c in range(1, 12)]
            if all(v is None or (isinstance(v, str) and not v.strip()) for v in row_vals):
                continue

            if isinstance(rowA, str) and rowA.strip().upper() == "TOTAL":
                break
            if isinstance(rowA, str) and _norm(rowA.strip()) == "site":
                continue

            site = ws_fx.cell(r, 1).value
            regime = ws_fx.cell(r, 2).value

            if site is not None and clean_text(site):
                current_site = clean_text(site)

            regime_txt = normalize_regime_label(regime)
            if not regime_txt:
                continue

            vals = []
            for d in DAY_NAMES:
                raw = get_value(r, day_cols.get(d, 0))
                vals.append(int(_to_number(raw)))

            if total_col:
                total = _to_number(get_value(r, total_col))
            else:
                total = float(sum(vals))

            rows.append([current_site or "", regime_txt, *vals, total])

        cols = ["Site", "Regime", *DAY_NAMES, "Total"]
        df = pd.DataFrame(rows, columns=cols)
        if not df.empty:
            df["Regime"] = df["Regime"].astype(str).str.replace(r"\s+", " ", regex=True).str.strip()
        return df

    df_dej = read_block(dejeuner_row, diner_row - 1)
    df_din = read_block(diner_row, ws_fx.max_row)
    return {"dejeuner": df_dej, "diner": df_din}


def parse_planning_mixe_lisse(path: str, sheet_name: str = "Planning mixe lisse ") -> Dict[str, pd.DataFrame]:
    """
    Lit la feuille "Planning mixe lisse" (déjeuner + dîner) et retourne:
    {"dejeuner": df, "diner": df} avec colonnes Site, Regime (Mixé/Lissé), Lundi..Dimanche.
    Gère les formules Excel simples: références, SUM(), expressions + / -.
    """
    wb_val = openpyxl.load_workbook(path, data_only=True)
    wb_fx  = openpyxl.load_workbook(path, data_only=False)
    real_sheet = sheet_name
    if real_sheet not in wb_fx.sheetnames:
        real_sheet = _find_sheet_name(
            wb_fx.sheetnames,
            sheet_name,
            aliases=[
                "PLANNING MIXE LISSE",
                "PLANNING MIXE/LISSE",
                "PLANNING MIXE",
                "MIXE LISSE",
                "MIXE/LISSE",
            ],
        )

    if not real_sheet or real_sheet not in wb_fx.sheetnames:
        # keep old behaviour: if sheet missing, just return empty frames
        return {
            "dejeuner": pd.DataFrame(columns=["Site", "Regime"] + DAY_NAMES),
            "diner": pd.DataFrame(columns=["Site", "Regime"] + DAY_NAMES),
        }

    ws_val = wb_val[real_sheet]
    ws_fx  = wb_fx[real_sheet]

    ref_re = re.compile(r"^(?:'([^']+)'|([^'!]+))!?\$?([A-Z]{1,3})\$?(\d{1,7})$")
    sheetref_re = re.compile(r"^=\s*\+?\s*(?:'([^']+)'\s*|([^'!]+))!?\$?([A-Z]{1,3})\$?(\d{1,7})\s*$")
    cell_re = re.compile(r"^\$?([A-Z]{1,3})\$?(\d{1,7})$")

    def addr(col: str, row: int) -> str:
        return f"{col}{row}"

    def _get(sheet: str, a: str):
        if sheet in wb_val.sheetnames:
            v = wb_val[sheet][a].value
            if v is not None:
                return v
        return wb_fx[sheet][a].value

    def _eval(sheet: str, a: str, depth: int = 0):
        if depth > 20:
            return None
        v = _get(sheet, a)

        if v is None or isinstance(v, (int, float, dt.date, dt.datetime)):
            return v

        if isinstance(v, str) and v.startswith("="):
            s = v.strip()
            m = sheetref_re.match(s)
            if m:
                sh = m.group(1) or m.group(2)
                col = m.group(3)
                row = int(m.group(4))
                return _eval(sh, addr(col, row), depth + 1)

            sm = re.match(r"^=\s*SUM\(([^)]+)\)\s*$", s, re.I)
            if sm:
                rng = sm.group(1).strip()
                if ":" in rng:
                    a1, a2 = [x.strip().replace("$","") for x in rng.split(":", 1)]
                    m1 = cell_re.match(a1); m2 = cell_re.match(a2)
                    if m1 and m2:
                        c1, r1 = m1.group(1), int(m1.group(2))
                        c2, r2 = m2.group(1), int(m2.group(2))
                        if c1 == c2:
                            total = 0.0
                            for rr in range(min(r1,r2), max(r1,r2)+1):
                                total += _to_number(_eval(sheet, addr(c1, rr), depth + 1))
                            return total
                parts = [p.strip() for p in re.split(r"[;,]", rng) if p.strip()]
                total = 0.0
                for p in parts:
                    p2 = p.replace("$","")
                    if ":" in p2:
                        continue
                    m1 = cell_re.match(p2)
                    if m1:
                        total += _to_number(_eval(sheet, addr(m1.group(1), int(m1.group(2))), depth + 1))
                return total

            expr = s.lstrip("=").strip()
            expr = expr.replace(";", "+")
            tokens = re.split(r"(\+|\-)", expr)
            total = 0.0
            sign = +1.0
            for t in tokens:
                t = t.strip()
                if not t:
                    continue
                if t == "+":
                    sign = +1.0
                    continue
                if t == "-":
                    sign = -1.0
                    continue
                t_clean = t.replace("$","")
                mm = ref_re.match(t_clean)
                if mm:
                    sh = mm.group(1) or mm.group(2) or sheet
                    col = mm.group(3)
                    row = int(mm.group(4))
                    total += sign * _to_number(_eval(sh, addr(col, row), depth + 1))
                    continue
                mc = cell_re.match(t_clean)
                if mc:
                    total += sign * _to_number(_eval(sheet, addr(mc.group(1), int(mc.group(2))), depth + 1))
                    continue
                total += sign * _to_number(t_clean)
            return total

        return v

    def _read_block(start_row: int, header_row: int) -> pd.DataFrame:
        day_cols = {DAY_NAMES[i]: 3+i for i in range(7)}
        rows = []
        cur_site = None
        r = start_row
        while r <= ws_fx.max_row:
            a = clean_text(ws_fx.cell(r, 1).value)
            if a.upper() == "TOTAL":
                break
            if not a and not clean_text(ws_fx.cell(r, 2).value):
                r += 1
                continue

            if a:
                cur_site = a
            reg = clean_text(ws_fx.cell(r, 2).value)
            if not reg or reg.lower() not in {"mixe", "mixé", "lisse", "lissé"}:
                r += 1
                continue

            reg_out = "Mixé" if "mix" in _norm(reg) else "Lissé"
            vals = []
            for d, cidx in day_cols.items():
                col_letter = ws_fx.cell(header_row, cidx).column_letter
                v = _eval(sheet_name, addr(col_letter, r))
                vals.append(int(_to_number(v)))

            rows.append([cur_site or "", reg_out, *vals])
            r += 1

        df = pd.DataFrame(rows, columns=["Site","Regime"] + DAY_NAMES)
        return df

    df_dej = _read_block(start_row=3, header_row=2)
    df_din = _read_block(start_row=19, header_row=18)
    return {"dejeuner": df_dej, "diner": df_din}


@dataclass
class MenuItem:
    date: dt.date
    repas: str  # "Déjeuner" or "Dîner"
    categorie: str  # Entrée, Plat, Laitage, Dessert
    regime: str
    produit: str


def parse_menu(path: str, sheet_name: str = "Feuil2") -> List[MenuItem]:
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Feuille '{sheet_name}' introuvable. Feuilles dispo: {wb.sheetnames}")
    ws = wb[sheet_name]

    group = {c: clean_text(ws.cell(2, c).value) for c in range(2, ws.max_column + 1)}
    header = {c: clean_text(ws.cell(3, c).value) for c in range(2, ws.max_column + 1)}

    regime_by_col: Dict[int, str] = {}
    for c in range(2, ws.max_column + 1):
        h = header.get(c, "")
        g = group.get(c, "")
        if not h and not g:
            continue

        if not h:
            label = g
        elif not g:
            label = h
        elif g and g.upper() not in h.upper():
            label = f"{g} - {h}"
        else:
            label = h

        label = label.replace("STANDART", "STANDARD")
        label = re.sub(r"\s+", " ", label).strip()
        regime_by_col[c] = label

    items: List[MenuItem] = []

    def _norm_kw(s: str) -> str:
        s = clean_text(s).lower()
        s = "".join(ch for ch in unicodedata.normalize("NFKD", s) if not unicodedata.combining(ch))
        return s

    _DESSERT_KW = ("compote", "fruit", "gateau", "gateau", "tarte", "flan", "creme", "crème", "mousse", "riz au lait", "ile flottante")
    _DAIRY_KW = ("fromage", "yaourt", "yogourt", "fromage blanc", "petit suisse", "camembert", "emmental", "kiri", "tartare", "babybel", "gouda", "boursin")

    def _row_score(row: int, kws: tuple[str, ...]) -> int:
        score = 0
        for c in regime_by_col.keys():
            v = clean_text(ws.cell(row, c).value)
            if not v:
                continue
            t = _norm_kw(v)
            if any(k in t for k in kws):
                score += 1
        return score

    def detect_block_height(start_row: int) -> int:
        candidates = [(6, 4, 5), (5, 3, 4)]
        best_h = 6
        best_score = -1
        for h, off_lait, off_des in candidates:
            laitage_row = start_row + off_lait
            dessert_row = start_row + off_des
            if dessert_row > ws.max_row:
                continue
            score = _row_score(laitage_row, _DAIRY_KW) + _row_score(dessert_row, _DESSERT_KW)
            if _row_score(dessert_row, _DESSERT_KW) > 0:
                score += 2
            if score > best_score:
                best_score = score
                best_h = h
        return best_h

    def read_block(start_row: int, date_val: dt.date, repas: str, block_height: int):
        entree_row = start_row
        if block_height == 5:
            plat_rows = [start_row + 1, start_row + 2]
            laitage_row = start_row + 3
            dessert_row = start_row + 4
        else:
            plat_rows = [start_row + 1, start_row + 2, start_row + 3]
            laitage_row = start_row + 4
            dessert_row = start_row + 5

        for c, regime in regime_by_col.items():
            entree = clean_text(ws.cell(entree_row, c).value)
            plat_parts = [clean_text(ws.cell(rr, c).value) for rr in plat_rows]
            plat_parts = [p for p in plat_parts if p]
            laitage = clean_text(ws.cell(laitage_row, c).value)
            dessert = clean_text(ws.cell(dessert_row, c).value)

            if entree:
                items.append(MenuItem(date_val, repas, "Entrée", regime, entree))
            for p in plat_parts:
                items.append(MenuItem(date_val, repas, "Plat", regime, p))
            if laitage:
                items.append(MenuItem(date_val, repas, "Laitage", regime, laitage))
            if dessert:
                items.append(MenuItem(date_val, repas, "Dessert", regime, dessert))

    date_rows: List[tuple[int, dt.date]] = []
    anchor_year: Optional[int] = None
    for r in range(4, ws.max_row + 1):
        raw = ws.cell(r, 1).value
        d = _parse_date(raw, default_year=anchor_year)
        if d is not None:
            anchor_year = d.year
            date_rows.append((r, d))

    seen = set()
    uniq: List[tuple[int, dt.date]] = []
    for rr, dd in date_rows:
        if dd in seen:
            continue
        seen.add(dd)
        uniq.append((rr, dd))

    for rr, dd in uniq:
        h_dej = detect_block_height(rr)
        read_block(rr, dd, "Déjeuner", h_dej)

        rr_din = rr + h_dej
        if rr_din <= ws.max_row:
            h_din = detect_block_height(rr_din)
            read_block(rr_din, dd, "Dîner", h_din)

    return items


# =============================
# Bons de livraison (PDF)
# =============================

def clean_text_delivery(x) -> str:
    if x is None:
        return ""
    s = str(x)
    s = s.replace("\u2026", "...")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def parse_menu_delivery(path: str, sheet_name: str = "Feuil2") -> Dict[tuple, List[str]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Feuille '{sheet_name}' introuvable. Feuilles dispo: {wb.sheetnames}")
    ws = wb[sheet_name]

    group = {c: clean_text_delivery(ws.cell(2, c).value) for c in range(2, ws.max_column + 1)}
    header = {c: clean_text_delivery(ws.cell(3, c).value) for c in range(2, ws.max_column + 1)}

    regime_by_col: Dict[int, str] = {}
    for c in range(2, ws.max_column + 1):
        h = header.get(c, "")
        if not h:
            continue
        g = group.get(c, "")
        if g and g.upper() not in h.upper():
            label = f"{g} - {h}"
        else:
            label = h
        label = label.replace("STANDART", "STANDARD")
        label = re.sub(r"\s+", " ", label).strip()
        regime_by_col[c] = label

    out: Dict[tuple, List[str]] = {}

    def _merged_value_any(row: int, col: int) -> str:
        cell = ws.cell(row, col)
        if cell.value not in (None, ""):
            return clean_text_delivery(cell.value)
        for rng in ws.merged_cells.ranges:
            if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
                v = ws.cell(rng.min_row, rng.min_col).value
                return clean_text_delivery(v)
        return ""

    def _merged_value_strict(row: int, col: int) -> str:
        cell = ws.cell(row, col)
        if cell.value not in (None, ""):
            return clean_text_delivery(cell.value)
        for rng in ws.merged_cells.ranges:
            if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
                if rng.min_col != col:
                    return ""
                v = ws.cell(rng.min_row, rng.min_col).value
                return clean_text_delivery(v)
        return ""

    _DESSERT_KW = ("compote", "fruit", "gateau", "tarte", "flan", "creme", "crème", "mousse", "riz au lait", "ile flottante")
    _DAIRY_KW = ("fromage", "yaourt", "yogourt", "fromage blanc", "petit suisse", "camembert", "emmental", "kiri", "tartare", "babybel", "gouda", "boursin")

    def _norm_kw(s: str) -> str:
        s = clean_text_delivery(s).lower()
        s = "".join(ch for ch in unicodedata.normalize("NFKD", s) if not unicodedata.combining(ch))
        return s

    def _row_score(row: int, kws: tuple[str, ...]) -> int:
        score = 0
        for c in regime_by_col.keys():
            v = _merged_value_any(row, c)
            if not v:
                continue
            t = _norm_kw(v)
            if any(k in t for k in kws):
                score += 1
        return score

    def detect_block_height(start_row: int) -> int:
        candidates = [(6, 4, 5), (5, 3, 4)]
        best_h = 6
        best_score = -1
        for h, off_lait, off_des in candidates:
            laitage_row = start_row + off_lait
            dessert_row = start_row + off_des
            if dessert_row > ws.max_row:
                continue
            score = _row_score(laitage_row, _DAIRY_KW) + _row_score(dessert_row, _DESSERT_KW)
            if score > best_score:
                best_score = score
                best_h = h
        return best_h

    def read_block(start_row: int, date_val: dt.date, repas: str, block_height: int):
        if block_height == 5:
            entree_row = start_row + 0
            plat1_row = start_row + 1
            plat2_rows = [start_row + 2]
            laitage_row = start_row + 3
            dessert_row = start_row + 4
        else:
            entree_row = start_row + 0
            plat1_row = start_row + 1
            plat2_rows = [start_row + 2, start_row + 3]
            laitage_row = start_row + 4
            dessert_row = start_row + 5

        for c, regime in regime_by_col.items():
            entree = _merged_value_strict(entree_row, c)
            plat1 = _merged_value_strict(plat1_row, c)
            plat2_parts = [_merged_value_strict(rr, c) for rr in plat2_rows]
            plat2_parts = [p for p in plat2_parts if p]
            plat2 = " / ".join(plat2_parts)

            laitage = _merged_value_any(laitage_row, c)
            dessert = _merged_value_any(dessert_row, c)

            lines = [entree, plat1, plat2, laitage, dessert]
            lines = [ln if ln else "—" for ln in lines]
            out[(date_val, repas, regime)] = lines

    date_rows: List[tuple[int, dt.date]] = []
    anchor_year: Optional[int] = None
    for r in range(4, ws.max_row + 1):
        raw = ws.cell(r, 1).value
        d = _parse_date(raw, default_year=anchor_year)
        if d is not None:
            anchor_year = d.year
            date_rows.append((r, d))

    seen = set()
    uniq: List[tuple[int, dt.date]] = []
    for rr, dd in date_rows:
        if dd in seen:
            continue
        seen.add(dd)
        uniq.append((rr, dd))

    for rr, dd in uniq:
        h_dej = detect_block_height(rr)
        read_block(rr, dd, "Déjeuner", h_dej)

        rr_din = rr + h_dej
        if rr_din <= ws.max_row:
            h_din = detect_block_height(rr_din)
            read_block(rr_din, dd, "Dîner", h_din)

    return out


def export_bons_livraison_pdf(
    planning: Dict[str, pd.DataFrame],
    menu_path: str,
    out_pdf_path: str,
    planning_path: Optional[str] = None,
    sheet_menu: str = "Feuil2",
    sites_exclus: Optional[List[str]] = None,
) -> None:
    sites_exclus = sites_exclus or ["24 ter", "24 simple", "IME TL"]
    excl_norm = {_norm(s) for s in sites_exclus}

    menu_lines = parse_menu_delivery(menu_path, sheet_name=sheet_menu)

    def _dayname(d: dt.date) -> str:
        return DAY_NAMES[d.weekday()]

    def _site_day_total(df: pd.DataFrame, site: str, day_name: str) -> int:
        if df is None or df.empty:
            return 0
        tmp = df.copy()
        tmp["Site"] = tmp["Site"].astype(str).str.strip().str.lower()
        s = (site or "").strip().lower()
        sub = tmp[tmp["Site"] == s]
        if sub.empty or day_name not in sub.columns:
            return 0
        return int(pd.to_numeric(sub[day_name], errors="coerce").fillna(0).sum())

    all_dates = sorted({d for d in (_parse_date(k[0]) for k in menu_lines.keys()) if d is not None})

    sites = set()
    for key in ["dejeuner", "diner"]:
        df = planning.get(key)
        if df is not None and not df.empty:
            sites |= set(df["Site"].astype(str).tolist())
    sites_list = [s for s in sites if _norm(s) not in excl_norm]

    preferred = ["FAM", "Bussière", "Bruyère", "FO", "FDP", "MAS", "ESAT", "FM", "René"]
    pref_norm = [_norm(x) for x in preferred]

    def site_rank(s: str) -> tuple:
        ns = _norm(s)
        for idx, pn in enumerate(pref_norm):
            if pn in {"fo", "fdp"}:
                if ns == "fo" or ns == "fdp" or ns.startswith("fo ") or ns.startswith("fdp ") or " fo" in ns or " fdp" in ns:
                    return (idx, s)
            if ns == pn or ns.startswith(pn + " ") or (" " + pn) in ns:
                return (idx, s)
        return (999, s)

    sites = sorted(sites_list, key=site_rank)

    SITE_LABELS = {
        "fo": "Foyer Du Près",
        "fdp": "Foyer Du Près",
        "fm": "Foyer Fernand Marlier",
    }

    def display_site_name(s: str) -> str:
        ns = _norm(s)
        for k, v in SITE_LABELS.items():
            if ns == k or ns.startswith(k + " ") or (" " + k) in ns:
                return v
        return s

    order = [
        "hypocalorique",
        "speciaux",
        "sans",
        "lisse",
        "mixe",
        "standard",
        "vegetarien",
    ]

    def regime_sort_key(reg: str):
        n = _norm(reg)
        for i, tok in enumerate(order):
            if tok in n:
                return (i, n)
        return (999, n)

    def is_mixe_lisse(reg: str) -> bool:
        n = _norm(reg)
        if 'mixe' in n or 'lisse' in n:
            return True
        if n in {'ml', 'm l'} or ' ml ' in f' {n} ':
            return True
        return False

    def jour_col(date_val: dt.date) -> str:
        return DAY_NAMES[date_val.weekday()]

    def count_for(df: pd.DataFrame, site: str, regime: str, date_val: dt.date) -> int:
        if df is None or df.empty:
            return 0
        col = jour_col(date_val)
        sub = df[(df["Site"] == site) & (df["Regime"] == regime)]
        if sub.empty:
            return 0
        return int(pd.to_numeric(sub[col], errors="coerce").fillna(0).sum())

    def regimes_for_site(date_val: dt.date, site: str) -> List[str]:
        regs = set()
        for key in ["dejeuner", "diner"]:
            df = planning.get(key)
            if df is None or df.empty:
                continue
            col = jour_col(date_val)
            sub = df[df["Site"] == site]
            if sub.empty:
                continue
            sub = sub[pd.to_numeric(sub[col], errors="coerce").fillna(0) > 0]
            regs |= set(sub["Regime"].astype(str).tolist())
        return sorted(regs, key=regime_sort_key)

    def fmt_date_fr(d: dt.date) -> str:
        jours = ["lundi","mardi","mercredi","jeudi","vendredi","samedi","dimanche"]
        mois = ["janvier","février","mars","avril","mai","juin","juillet","août","septembre","octobre","novembre","décembre"]
        return f"{jours[d.weekday()]} {d.day} {mois[d.month-1]} {d.year}"

    W, H = A4
    x0 = 34
    y_top = H - 18
    line_h = 12

    mix_planning = None
    if planning_path:
        try:
            mix_planning = parse_planning_mixe_lisse(planning_path)
        except Exception:
            mix_planning = None

    c = canvas.Canvas(out_pdf_path, pagesize=A4)

    def _wrap_text(text: str, font_name: str, font_size: int, max_width: float) -> List[str]:
        words = (text or "").split()
        if not words:
            return [""]
        out_lines: List[str] = []
        cur = words[0]
        for w in words[1:]:
            test = cur + " " + w
            if c.stringWidth(test, font_name, font_size) <= max_width:
                cur = test
            else:
                out_lines.append(cur)
                cur = w
        out_lines.append(cur)
        return out_lines

    def _draw_bullet(x: float, y: float, text: str, max_width: float, font_size: int = 9, leading: float = 11) -> float:
        bullet = "• "
        font_name = "Helvetica"
        c.setFont(font_name, font_size)
        first_indent = c.stringWidth(bullet, font_name, font_size)
        lines = _wrap_text(text, font_name, font_size, max_width - first_indent)
        if not lines:
            lines = [""]
        c.drawString(x, y, bullet + lines[0])
        y -= leading
        for ln in lines[1:]:
            c.drawString(x + first_indent, y, ln)
            y -= leading
        return y

    def draw_page(site_key: str, date_val: dt.date):
        site_display = display_site_name(site_key)
        c.setFont("Helvetica-Bold", 15)
        c.drawString(x0, y_top, "BON DE LIVRAISON")

        y = y_top - 24
        c.setFont("Helvetica-Bold", 9)
        c.drawString(x0, y, "Site : ")
        c.setFont("Helvetica", 9)
        c.drawString(x0 + 27.5, y, site_display)

        y_min = 60
        max_width = (W - x0) - (x0 + 30)

        y -= line_h
        c.setFont("Helvetica-Bold", 9)
        c.drawString(x0, y, "Date : ")
        c.setFont("Helvetica", 9)
        c.drawString(x0 + 27.5, y, fmt_date_fr(date_val))

        tournee = "Barquette" if _norm(site_key) == _norm("MAS") else "Camion"
        y -= line_h
        c.setFont("Helvetica-Bold", 9)
        c.drawString(x0, y, "Tournée : ")
        c.setFont("Helvetica", 9)
        c.drawString(x0 + 44, y, tournee)

        df_dej = planning.get("dejeuner")
        df_din = planning.get("diner")
        col = jour_col(date_val)
        tot_dej = 0
        tot_din = 0
        if df_dej is not None and not df_dej.empty:
            tot_dej = int(pd.to_numeric(df_dej[df_dej["Site"] == site_key][col], errors="coerce").fillna(0).sum())
        if df_din is not None and not df_din.empty:
            tot_din = int(pd.to_numeric(df_din[df_din["Site"] == site_key][col], errors="coerce").fillna(0).sum())

        mix_dej = lisse_dej = mix_din = lisse_din = 0
        if mix_planning is not None:
            try:
                mdej = mix_planning.get("dejeuner")
                mdin = mix_planning.get("diner")
                if mdej is not None and not mdej.empty:
                    sub = mdej[mdej["Site"].astype(str).str.strip().str.lower() == str(site_key).strip().lower()]
                    if not sub.empty:
                        v_l = sub[sub["Regime"].astype(str).str.lower().str.contains("liss")][col].sum()
                        v_m = sub[sub["Regime"].astype(str).str.lower().str.contains("mix")][col].sum()
                        lisse_dej = int(_to_number(v_l)); mix_dej = int(_to_number(v_m))
                if mdin is not None and not mdin.empty:
                    sub = mdin[mdin["Site"].astype(str).str.strip().str.lower() == str(site_key).strip().lower()]
                    if not sub.empty:
                        v_l = sub[sub["Regime"].astype(str).str.lower().str.contains("liss")][col].sum()
                        v_m = sub[sub["Regime"].astype(str).str.lower().str.contains("mix")][col].sum()
                        lisse_din = int(_to_number(v_l)); mix_din = int(_to_number(v_m))
            except Exception:
                pass

        tot_dej_all = tot_dej + mix_dej + lisse_dej
        tot_din_all = tot_din + mix_din + lisse_din

        y -= line_h * 1.6
        c.setFont("Helvetica-Bold", 9)
        c.drawString(x0, y, f"Total Déjeuner : {tot_dej_all}    Total Dîner : {tot_din_all}")

        if (mix_dej + lisse_dej + mix_din + lisse_din) > 0:
            y -= 12
            c.setFont("Helvetica-Bold", 9)
            c.drawString(x0, y, f"Détail Mixé/Lissé (inclus dans le total) — Déj: Mixé {mix_dej} / Lissé {lisse_dej}   |   Dîn: Mixé {mix_din} / Lissé {lisse_din}")
            c.setFont("Helvetica", 9)

        mix_dej2 = 0
        mix_din2 = 0
        if df_dej is not None and not df_dej.empty:
            subm = df_dej[(df_dej["Site"] == site_key) & (df_dej["Regime"].astype(str).apply(is_mixe_lisse))]
            if not subm.empty:
                mix_dej2 = int(pd.to_numeric(subm[col], errors="coerce").fillna(0).sum())
        if df_din is not None and not df_din.empty:
            subm = df_din[(df_din["Site"] == site_key) & (df_din["Regime"].astype(str).apply(is_mixe_lisse))]
            if not subm.empty:
                mix_din2 = int(pd.to_numeric(subm[col], errors="coerce").fillna(0).sum())

        if mix_dej2 > 0 or mix_din2 > 0:
            y -= 12
            c.setFont("Helvetica-Bold", 9)
            c.drawString(x0, y, f"Mixé/Lissé livré : Déj {mix_dej2} / Dîn {mix_din2}")
            c.setFont("Helvetica", 9)

        y -= 24
        box_h = 28
        gap = 20
        box_w = (W - 2 * x0 - gap) / 2
        box_top = y

        c.rect(x0, box_top - box_h, box_w, box_h, stroke=1, fill=0)
        c.rect(x0 + box_w + gap, box_top - box_h, box_w, box_h, stroke=1, fill=0)

        c.setFont("Helvetica-Bold", 9)
        c.drawString(x0 + 6, box_top - 12, "Température départ : ____ °C")
        c.drawString(x0 + box_w + gap + 6, box_top - 12, "Température réception : ____ °C")

        y = box_top - box_h - 10

        regs = regimes_for_site(date_val, site_key)
        available_h = max(0.0, y - y_min)

        def _bullet_wrapped_lines(text: str, font_size: float) -> int:
            bullet = "• "
            font_name = "Helvetica"
            first_indent = c.stringWidth(bullet, font_name, font_size)
            lines = _wrap_text(text, font_name, font_size, max_width - first_indent)
            return max(1, len(lines or []))

        def _estimate_needed_height(scale: float) -> float:
            fs = 9.0 * scale
            lead = 10.0 * scale
            gap12 = 12.0 * scale
            gap6 = 6.0 * scale

            needed = 0.0
            for reg in regs:
                dej_n = count_for(df_dej, site_key, reg, date_val)
                din_n = count_for(df_din, site_key, reg, date_val)
                if dej_n <= 0 and din_n <= 0:
                    continue

                needed += gap12
                needed += gap12 + gap12
                if is_mixe_lisse(reg):
                    needed += gap12
                else:
                    reg_norm = _norm(reg)
                    lines = menu_lines.get((date_val, "Déjeuner", reg))
                    if not lines:
                        target = set(reg_norm.split())
                        best = None
                        best_score = -1
                        for (d, repas, rlabel), lns in menu_lines.items():
                            if d != date_val or repas != "Déjeuner":
                                continue
                            score = len(target & set(_norm(rlabel).split()))
                            if score > best_score:
                                best_score = score
                                best = lns
                        lines = best
                    lines = [ln for ln in (lines or []) if clean_text_delivery(ln) and clean_text_delivery(ln) != "—"]
                    if not lines:
                        lines = ["—"]
                    for ln in lines:
                        needed += _bullet_wrapped_lines(ln, fs) * lead

                needed += gap6 + gap12 + gap12
                if is_mixe_lisse(reg):
                    needed += gap12
                else:
                    reg_norm = _norm(reg)
                    lines = menu_lines.get((date_val, "Dîner", reg))
                    if not lines:
                        target = set(reg_norm.split())
                        best = None
                        best_score = -1
                        for (d, repas, rlabel), lns in menu_lines.items():
                            if d != date_val or repas != "Dîner":
                                continue
                            score = len(target & set(_norm(rlabel).split()))
                            if score > best_score:
                                best_score = score
                                best = lns
                        lines = best
                    lines = [ln for ln in (lines or []) if clean_text_delivery(ln) and clean_text_delivery(ln) != "—"]
                    if not lines:
                        lines = ["—"]
                    for ln in lines:
                        needed += _bullet_wrapped_lines(ln, fs) * lead

                needed += gap6

            return needed

        needed_h = _estimate_needed_height(1.0)
        scale = 1.0
        if needed_h > available_h and needed_h > 0:
            scale = max(0.75, min(1.0, (available_h / needed_h) * 0.98))

        fs = 9.0 * scale
        lead = 10.0 * scale
        gap12 = 12.0 * scale
        gap6 = 6.0 * scale

        for reg in regs:
            dej_n = count_for(df_dej, site_key, reg, date_val)
            din_n = count_for(df_din, site_key, reg, date_val)
            if dej_n <= 0 and din_n <= 0:
                continue

            y -= gap12
            c.setFont("Helvetica-Bold", fs)
            c.drawString(x0, y, f"{reg} —  Déj {dej_n} / Dîn {din_n}")

            y -= gap12
            c.setFont("Helvetica-Bold", fs)
            c.drawString(x0 + 12, y, "Déjeuner")
            y -= gap12
            c.setFont("Helvetica", fs)

            reg_norm = _norm(reg)
            if is_mixe_lisse(reg):
                c.drawString(x0 + 30, y, f"• Quantité mixé/lissé à livrer : {dej_n}")
                y -= gap12
            else:
                lines = menu_lines.get((date_val, "Déjeuner", reg))
                if not lines:
                    target = set(reg_norm.split())
                    best = None
                    best_score = -1
                    for (d, repas, rlabel), lns in menu_lines.items():
                        if d != date_val or repas != "Déjeuner":
                            continue
                        score = len(target & set(_norm(rlabel).split()))
                        if score > best_score:
                            best_score = score
                            best = lns
                    lines = best
                lines = [ln for ln in (lines or []) if clean_text_delivery(ln) and clean_text_delivery(ln) != "—"]
                if not lines:
                    lines = ["—"]
                for ln in lines:
                    y = _draw_bullet(x0 + 30, y, ln, max_width, font_size=int(fs), leading=lead)

            y -= gap6
            c.setFont("Helvetica-Bold", fs)
            c.drawString(x0 + 12, y, "Dîner")
            y -= gap12
            c.setFont("Helvetica", fs)

            if is_mixe_lisse(reg):
                c.drawString(x0 + 30, y, f"• Quantité mixé/lissé à livrer : {din_n}")
                y -= gap12
            else:
                lines = menu_lines.get((date_val, "Dîner", reg))
                if not lines:
                    target = set(reg_norm.split())
                    best = None
                    best_score = -1
                    for (d, repas, rlabel), lns in menu_lines.items():
                        if d != date_val or repas != "Dîner":
                            continue
                        score = len(target & set(_norm(rlabel).split()))
                        if score > best_score:
                            best_score = score
                            best = lns
                    lines = best
                lines = [ln for ln in (lines or []) if clean_text_delivery(ln) and clean_text_delivery(ln) != "—"]
                if not lines:
                    lines = ["—"]
                for ln in lines:
                    y = _draw_bullet(x0 + 30, y, ln, max_width, font_size=int(fs), leading=lead)

            y -= gap6

        c.setLineWidth(1)
        c.line(x0, 46, W - x0, 46)
        c.setFont("Helvetica-Bold", 9)
        c.drawString(x0, 30, "Chauffeur (signature) : ____________________")
        c.drawString(W / 2 + 20, 30, "Réception (signature) : ____________________")

    for site in sites:
        for d in all_dates:
            day = _dayname(d)
            tot_dej = _site_day_total(planning.get("dejeuner"), site, day)
            tot_din = _site_day_total(planning.get("diner"), site, day)
            if (tot_dej + tot_din) <= 0:
                continue
            draw_page(site, d)
            c.showPage()

    c.save()


def make_production_summary(df_planning: pd.DataFrame) -> pd.DataFrame:
    if df_planning is None or df_planning.empty:
        return pd.DataFrame(columns=["Jour", "Regime", "Nb"])

    long = df_planning.melt(
        id_vars=["Site", "Regime"],
        value_vars=DAY_NAMES,
        var_name="Jour",
        value_name="Nb",
    )
    long["Nb"] = pd.to_numeric(long["Nb"], errors="coerce").fillna(0).astype(int)
    long["Regime"] = long["Regime"].apply(normalize_regime_label)
    out = long.groupby(["Jour", "Regime"], as_index=False)["Nb"].sum()
    out["Nb"] = out["Nb"].astype(int)
    return out


def make_production_pivot(df_planning: pd.DataFrame) -> pd.DataFrame:
    if df_planning is None or df_planning.empty:
        cols = ["Regime", *DAY_NAMES, "Total"]
        return pd.DataFrame(columns=cols)

    long = make_production_summary(df_planning)
    if long.empty:
        cols = ["Regime", *DAY_NAMES, "Total"]
        return pd.DataFrame(columns=cols)

    pivot = (
        long.pivot_table(index="Regime", columns="Jour", values="Nb", aggfunc="sum", fill_value=0)
        .reindex(columns=DAY_NAMES, fill_value=0)
        .reset_index()
    )
    pivot["Total"] = pivot[DAY_NAMES].sum(axis=1).astype(int)

    total_row = {"Regime": "TOTAL"}
    for j in DAY_NAMES:
        total_row[j] = int(pivot[j].sum())
    total_row["Total"] = int(pivot["Total"].sum())
    pivot = pd.concat([pivot, pd.DataFrame([total_row])], ignore_index=True)

    for j in DAY_NAMES + ["Total"]:
        pivot[j] = pd.to_numeric(pivot[j], errors="coerce").fillna(0).astype(int)

    return pivot


_SAUCE_QUALIFIERS = [
    "vinaigrette",
    "mayonnaise",
    "mayo",
    "ketchup",
    "barbecue",
    "bbq",
    "aigre douce",
    "aigre-douce",
    "curry",
    "tomate",
    "fromagere",
    "fromagère",
]


def normalize_produit_for_grouping(produit: str) -> str:
    s = clean_text(produit)
    if not s:
        return ""
    s = re.split(r"\s*\+\s*", s, maxsplit=1)[0].strip()
    s = re.sub(r"\([^)]*\)", "", s).strip()

    low = s.lower()
    low = re.sub(r"\bsauce\b\s+.*$", "", low).strip()
    for q in _SAUCE_QUALIFIERS:
        if re.search(rf"\b{re.escape(q)}\b", low) and len(low.split()) > 1:
            low = re.sub(rf"\b{re.escape(q)}\b.*$", "", low).strip()
            break

    out = low
    out = re.sub(r"\s+", " ", out).strip(" -;,:/")
    if not out:
        out = s
    return out[:1].upper() + out[1:]


def build_bon_commande(planning: Dict[str, pd.DataFrame], menu_items: List[MenuItem]) -> pd.DataFrame:
    def norm_reg(s: str) -> str:
        s = (s or "").lower()
        s = (
            s.replace("é", "e")
            .replace("è", "e")
            .replace("ê", "e")
            .replace("î", "i")
            .replace("ï", "i")
            .replace("ô", "o")
            .replace("à", "a")
            .replace("ç", "c")
        )
        s = re.sub(r"[^a-z0-9 ]+", " ", s)
        s = re.sub(r"\s+", " ", s).strip()
        return s

    records = []
    for repas_key, repas_label in [("dejeuner", "Déjeuner"), ("diner", "Dîner")]:
        df = planning.get(repas_key)
        if df is None or df.empty:
            continue

        df2 = df.copy()
        df2["Regime"] = df2["Regime"].apply(normalize_regime_label)
        agg = df2.groupby("Regime")[DAY_NAMES].sum(numeric_only=True)
        for jour in DAY_NAMES:
            for regime, nb in agg[jour].items():
                records.append(
                    {
                        "Repas": repas_label,
                        "Jour": jour,
                        "Regime_planning": regime,
                        "reg_key_planning": norm_reg(regime),
                        "Nb_personnes": int(_to_number(nb)),
                    }
                )

    counts = pd.DataFrame(records)
    if counts.empty:
        menu_df = pd.DataFrame(
            [
                {
                    "Date": it.date,
                    "Jour": DAY_NAMES[it.date.weekday()],
                    "Repas": it.repas,
                    "Typologie": it.categorie,
                    "Produit": it.produit,
                    "Effectif": 0,
                    "Coefficient": 1.0,
                }
                for it in menu_items
            ]
        )
        menu_df["Produit"] = menu_df["Produit"].astype(str)
        menu_df["Produit_base"] = menu_df["Produit"].apply(normalize_produit_for_grouping)
        menu_df["Coefficient"] = "1"
        menu_df["Unité"] = "unité"
        menu_df["Fournisseur"] = ""
        menu_df["Quantité"] = (menu_df["Effectif"] * 1.0).astype(int)
        grouped = (
            menu_df.groupby(["Repas", "Typologie", "Produit_base", "Coefficient"], as_index=False)
            .agg(
                {
                    "Jour": lambda s: ", ".join(sorted(set(s), key=lambda x: DAY_NAMES.index(x))),
                    "Produit": "first",
                    "Effectif": "sum",
                    "Quantité": "sum",
                }
            )
            .rename(columns={"Jour": "Jour(s)", "Produit_base": "Produit"})
        )
        grouped["Coefficient"] = "1"
        grouped["Unité"] = "unité"
        grouped["Fournisseur"] = ""
        grouped["Prix cible unitaire"] = ""
        grouped["Prix cible total"] = ""
        grouped["Poids unitaire (kg)"] = ""
        grouped["Poids total (kg)"] = ""
        return grouped[["Jour(s)", "Repas", "Typologie", "Produit", "Effectif", "Coefficient", "Unité", "Fournisseur", "Quantité", "Prix cible unitaire", "Prix cible total", "Poids unitaire (kg)", "Poids total (kg)"]].sort_values(["Repas", "Typologie", "Produit"]).reset_index(drop=True)

    planning_keys = counts[["Regime_planning", "reg_key_planning"]].drop_duplicates().to_dict("records")

    def best_match_planning_key(menu_key: str) -> Optional[str]:
        if not menu_key:
            return None
        mtoks = set(menu_key.split())
        best_key = None
        best_score = -1
        for rec in planning_keys:
            ptoks = set((rec["reg_key_planning"] or "").split())
            score = len(mtoks & ptoks)
            if score > best_score:
                best_score = score
                best_key = rec["reg_key_planning"]
        if best_score <= 0:
            return None
        return best_key

    menu_df = pd.DataFrame(
        [
            {
                "Date": it.date,
                "Jour": DAY_NAMES[it.date.weekday()],
                "Repas": it.repas,
                "Categorie": it.categorie,
                "Regime_menu": it.regime,
                "reg_key_menu": norm_reg(it.regime),
                "Produit": it.produit,
            }
            for it in menu_items
        ]
    )

    menu_df["reg_key_planning"] = menu_df["reg_key_menu"].apply(best_match_planning_key)

    merged = menu_df.merge(
        counts[["Repas", "Jour", "reg_key_planning", "Nb_personnes", "Regime_planning"]],
        on=["Repas", "Jour", "reg_key_planning"],
        how="left",
    )

    merged["Nb_personnes"] = merged["Nb_personnes"].fillna(0).astype(int)
    merged["Coefficient"] = "1"
    merged["Unité"] = "unité"
    merged["Fournisseur"] = ""

    base = merged[
        ["Date", "Jour", "Repas", "Categorie", "Produit", "Nb_personnes", "Coefficient", "Unité", "Fournisseur"]
    ].rename(
        columns={"Categorie": "Typologie", "Nb_personnes": "Effectif"}
    )

    base["Produit"] = base["Produit"].astype(str)
    base["Produit_base"] = base["Produit"].apply(normalize_produit_for_grouping)
    base["Quantité"] = (base["Effectif"] * 1.0).round().astype(int)

    grouped = (
        base.groupby(
            ["Repas", "Typologie", "Produit_base", "Coefficient", "Unité", "Fournisseur"],
            as_index=False,
        )
        .agg(
            {
                "Jour": lambda s: ", ".join(sorted(set(s), key=lambda x: DAY_NAMES.index(x))),
                "Effectif": "sum",
                "Quantité": "sum",
            }
        )
        .rename(columns={"Jour": "Jour(s)", "Produit_base": "Produit"})
    )

    if "Unité" not in grouped.columns:
        grouped["Unité"] = "unité"
    if "Fournisseur" not in grouped.columns:
        grouped["Fournisseur"] = ""
    grouped["Prix cible unitaire"] = ""

    grouped["Prix cible total"] = ""
    grouped["Poids unitaire (kg)"] = ""
    grouped["Poids total (kg)"] = ""

    grouped = grouped[["Jour(s)", "Repas", "Typologie", "Produit", "Effectif", "Coefficient", "Unité", "Fournisseur", "Quantité", "Prix cible unitaire", "Prix cible total", "Poids unitaire (kg)", "Poids total (kg)"]]
    return grouped.sort_values(["Repas", "Typologie", "Produit"]).reset_index(drop=True)


def export_excel(
    bon_commande: pd.DataFrame,
    prod_dej: pd.DataFrame,
    prod_din: pd.DataFrame,
    out_path: str,
    *,
    coefficients: Optional[List[Dict[str, object]]] = None,
    units: Optional[List[str]] = None,
    suppliers: Optional[List[Dict[str, str]]] = None,
) -> None:
    # (export_excel conservée proche de ta version; fixes: indentation + suppression doublon poids)
    def _clean_coeffs(raw: Optional[List[Dict[str, object]]]) -> List[Dict[str, object]]:
        out: List[Dict[str, object]] = []
        if not raw:
            raw = []
        for c in raw:
            name = str((c or {}).get("name", "")).strip()
            if not name:
                continue
            try:
                val = float((c or {}).get("value", 1.0))
            except Exception:
                val = 1.0
            out.append({"name": name, "value": val})
        if not any(x["name"] == "1" for x in out):
            out.insert(0, {"name": "1", "value": 1.0})
        seen = set()
        uniq = []
        for x in out:
            if x["name"] in seen:
                continue
            seen.add(x["name"])
            uniq.append(x)
        return uniq

    def _clean_units(raw: Optional[List[str]]) -> List[str]:
        # Contrainte: uniquement kg, L, unité
        return ["kg", "L", "unité"]

    def _clean_suppliers(raw: Optional[List[Dict[str, str]]]) -> List[Dict[str, str]]:
        if not raw:
            return []
        out = []
        for s in raw:
            name = str((s or {}).get("name", "")).strip()
            if not name:
                continue
            out.append(
                {
                    "name": name,
                    "customer_code": str((s or {}).get("customer_code", "") or ""),
                    "coord1": str((s or {}).get("coord1", "") or ""),
                    "coord2": str((s or {}).get("coord2", "") or ""),
                }
            )
        seen = set()
        uniq = []
        for x in out:
            if x["name"] in seen:
                continue
            seen.add(x["name"])
            uniq.append(x)
        return uniq

    coefficients = _clean_coeffs(coefficients)
    units = _clean_units(units)
    suppliers = _clean_suppliers(suppliers)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        bon_commande.to_excel(writer, sheet_name="Bon de commande", index=False)
        dej_piv = prod_dej
        din_piv = prod_din
        dej_piv.to_excel(writer, sheet_name="Déjeuner", index=False)
        din_piv.to_excel(writer, sheet_name="Dîner", index=False)

        wb = writer.book
        # --- Formules automatiques : Quantité uniquement (prix/poids calculés lors des bons fournisseurs) ---


        # --- Auto-formules sur la feuille "Bon de commande" ---
        # Objectif: pré-remplir Poids unitaire (kg) selon Unité (0,1 pour unité, 1 pour kg),
        # et calculer automatiquement les totaux (prix/poids). Les cellules restent modifiables
        # (l'utilisateur peut écraser la formule).
        try:
            ws_bc = wb["Bon de commande"]
            headers = {str(ws_bc.cell(row=1, column=c).value).strip(): c for c in range(1, ws_bc.max_column + 1)}
            col_unit = headers.get("Unité")
            col_qty = headers.get("Quantité")
            col_pu = headers.get("Prix cible unitaire")
            col_pt = headers.get("Prix cible total")
            col_wu = headers.get("Poids unitaire (kg)")
            col_wt = headers.get("Poids total (kg)")

            import math

            def _is_blank(v) -> bool:
                """Considère vide : None, chaîne vide, NaN."""
                if v is None:
                    return True
                if isinstance(v, str) and v.strip() == "":
                    return True
                if isinstance(v, (int, float)) and isinstance(v, float) and math.isnan(v):
                    return True
                return False

            if col_unit and col_qty and col_wu:
                for r in range(2, ws_bc.max_row + 1):
                    unit_cell = f"{get_column_letter(col_unit)}{r}"
                    qty_cell = f"{get_column_letter(col_qty)}{r}"

                    # Poids unitaire auto si cellule vide
                    wu_cell_obj = ws_bc.cell(row=r, column=col_wu)
                    if _is_blank(wu_cell_obj.value):
                        wu_cell = f"{get_column_letter(col_wu)}{r}"
                        wu_cell_obj.value = (
                            f'=IF(OR(LOWER(TRIM({unit_cell}))="kg",'
                            f'LOWER(TRIM({unit_cell}))="kilo",'
                            f'LOWER(TRIM({unit_cell}))="kilogramme",'
                            f'LOWER(TRIM({unit_cell}))="l",'
                            f'LOWER(TRIM({unit_cell}))="litre",'
                            f'LOWER(TRIM({unit_cell}))="litres"),1,0.1)'
                        )

                    # Poids total
                    if col_wt:
                        wt_obj = ws_bc.cell(row=r, column=col_wt)
                        if _is_blank(wt_obj.value):
                            wu_cell = f"{get_column_letter(col_wu)}{r}"
                            wt_obj.value = f"={qty_cell}*{wu_cell}"

                    # Prix cible total
                    if col_pu and col_pt:
                        pt_obj = ws_bc.cell(row=r, column=col_pt)
                        if _is_blank(pt_obj.value):
                            pu_cell = f"{get_column_letter(col_pu)}{r}"
                            pt_obj.value = f"={qty_cell}*{pu_cell}"
        except Exception:
            pass


        if "Listes" in wb.sheetnames:
            ws_list = wb["Listes"]
            ws_list.delete_rows(1, ws_list.max_row)
        else:
            ws_list = wb.create_sheet("Listes")

        ws_list["A1"].value = "Coefficient"
        ws_list["B1"].value = "Valeur"
        for i, c in enumerate(coefficients, start=2):
            ws_list.cell(row=i, column=1).value = str(c["name"]).strip()
            ws_list.cell(row=i, column=2).value = float(c["value"])

        ws_list["E1"].value = "Unités"
        for i, u in enumerate(units, start=2):
            ws_list.cell(row=i, column=5).value = str(u).strip()

        ws_list["G1"].value = "Fournisseur"
        ws_list["H1"].value = "Code client"
        ws_list["I1"].value = "Coordonnée 1"
        ws_list["J1"].value = "Coordonnée 2"
        for i, s in enumerate(suppliers, start=2):
            ws_list.cell(row=i, column=7).value = str(s["name"]).strip()
            ws_list.cell(row=i, column=8).value = str(s.get("customer_code", "") or "")
            ws_list.cell(row=i, column=9).value = str(s.get("coord1", "") or "")
            ws_list.cell(row=i, column=10).value = str(s.get("coord2", "") or "")

        ws_list.sheet_state = "hidden"

        ws_bc = wb["Bon de commande"]

        headers: dict[str, int] = {}
        for c in range(1, ws_bc.max_column + 1):
            v = ws_bc.cell(row=1, column=c).value
            if v is None:
                continue
            key = str(v).strip().lower()
            if key:
                headers[key] = c

        col_eff = headers.get("effectif")
        col_coef = headers.get("coefficient")
        col_unit = headers.get("unité") or headers.get("unite")
        col_sup = headers.get("fournisseur")
        col_qty = headers.get("quantité") or headers.get("quantite")

        n_coef = len(coefficients)
        n_units = len(units)
        n_sup = len(suppliers)

        coef_range = f"=Listes!$A$2:$A${1 + n_coef}" if n_coef else '"1"'
        unit_range = f"=Listes!$E$2:$E${1 + n_units}" if n_units else '"unité"'
        sup_range = f"=Listes!$G$2:$G${1 + n_sup}" if n_sup else '""'

        if col_coef:
            dv_coef = DataValidation(type="list", formula1=coef_range, allow_blank=True)
            ws_bc.add_data_validation(dv_coef)
            dv_coef.add(
                f"{openpyxl.utils.get_column_letter(col_coef)}2:"
                f"{openpyxl.utils.get_column_letter(col_coef)}{ws_bc.max_row}"
            )

        if col_unit:
            dv_unit = DataValidation(type="list", formula1=unit_range, allow_blank=True)
            ws_bc.add_data_validation(dv_unit)
            dv_unit.add(
                f"{openpyxl.utils.get_column_letter(col_unit)}2:"
                f"{openpyxl.utils.get_column_letter(col_unit)}{ws_bc.max_row}"
            )

        if col_sup:
            dv_sup = DataValidation(type="list", formula1=sup_range, allow_blank=True)
            ws_bc.add_data_validation(dv_sup)
            dv_sup.add(
                f"{openpyxl.utils.get_column_letter(col_sup)}2:"
                f"{openpyxl.utils.get_column_letter(col_sup)}{ws_bc.max_row}"
            )

        # ----------------- Formule Quantité -----------------
        if col_eff and col_coef and col_qty:
            eff_letter = openpyxl.utils.get_column_letter(col_eff)
            coef_letter = openpyxl.utils.get_column_letter(col_coef)
            lookup_table = f"Listes!$A$2:$B${1 + n_coef}"

            for r in range(2, ws_bc.max_row + 1):
                coef_key = f"TRIM(TEXT({coef_letter}{r},\"@\"))"
                ws_bc.cell(row=r, column=col_qty).value = (
                    f"=ROUND({eff_letter}{r}*IFERROR(VLOOKUP({coef_key},{lookup_table},2,FALSE),1),0)"
                )
                ws_bc.cell(row=r, column=col_qty).number_format = "0"

        # ----------------- Formules Prix/Poids (cibles) -----------------
        col_pu = headers.get("prix cible unitaire")
        col_pt = headers.get("prix cible total")
        col_wu = headers.get("poids unitaire (kg)") or headers.get("poids unitaire")
        col_wt = headers.get("poids total (kg)") or headers.get("poids total")

        if col_qty and col_pu and col_pt:
            qty_letter = openpyxl.utils.get_column_letter(col_qty)
            pu_letter = openpyxl.utils.get_column_letter(col_pu)
            for r in range(2, ws_bc.max_row + 1):
                ws_bc.cell(row=r, column=col_pt).value = (
                    f'=IF(OR({qty_letter}{r}="",{pu_letter}{r}=""),"",{qty_letter}{r}*{pu_letter}{r})'
                )
                ws_bc.cell(row=r, column=col_pt).number_format = '#,##0.00'

        if col_qty and col_wu and col_wt:
            qty_letter = openpyxl.utils.get_column_letter(col_qty)
            wu_letter = openpyxl.utils.get_column_letter(col_wu)
            for r in range(2, ws_bc.max_row + 1):
                ws_bc.cell(row=r, column=col_wt).value = (
                    f'=IF(OR({qty_letter}{r}="",{wu_letter}{r}=""),"",{qty_letter}{r}*{wu_letter}{r})'
                )
                ws_bc.cell(row=r, column=col_wt).number_format = '#,##0.000'

        # ----------------- Styles -----------------
        thin = Side(style="thin", color="9E9E9E")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill("solid", fgColor="EDEDED")
        header_font = Font(bold=True)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell_align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        band_fill = PatternFill("solid", fgColor="F7F7F7")

        for name in ["Bon de commande", "Déjeuner", "Dîner"]:
            ws = wb[name]

            header_row = 1
            if name in ("Déjeuner", "Dîner"):
                ws.insert_rows(1)
                max_col_tmp = ws.max_column
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col_tmp)
                tcell = ws.cell(row=1, column=1)
                tcell.value = f"BON DE COMMANDE – {name.upper()}"
                tcell.font = Font(bold=True, size=14)
                tcell.alignment = Alignment(horizontal="center", vertical="center")
                ws.row_dimensions[1].height = 28
                header_row = 2
                ws.freeze_panes = "B3"
                ws.page_setup.orientation = "landscape"
            else:
                ws.freeze_panes = "A2"

            max_row = ws.max_row
            max_col = ws.max_column

            for c in range(1, max_col + 1):
                cell = ws.cell(row=header_row, column=c)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = border
            ws.row_dimensions[header_row].height = 24

            for r in range(header_row + 1, max_row + 1):
                ws.row_dimensions[r].height = 18
                for c in range(1, max_col + 1):
                    cell = ws.cell(row=r, column=c)
                    if name in ("Déjeuner", "Dîner") and c >= 2:
                        cell.alignment = cell_align_center
                    else:
                        cell.alignment = cell_align
                    cell.border = border

                if r % 2 == 0:
                    for c in range(1, max_col + 1):
                        ws.cell(row=r, column=c).fill = band_fill

            if name in ("Déjeuner", "Dîner"):
                for r in range(header_row + 1, max_row + 1):
                    ws.cell(row=r, column=1).font = Font(bold=True)
                if max_row >= 2 and str(ws.cell(row=max_row, column=1).value).strip().upper() == "TOTAL":
                    for c in range(1, max_col + 1):
                        ws.cell(row=max_row, column=c).font = Font(bold=True)
                        ws.cell(row=max_row, column=c).fill = PatternFill("solid", fgColor="E0E0E0")

            ws.auto_filter.ref = f"A{header_row}:{openpyxl.utils.get_column_letter(max_col)}{max_row}"

            for c_idx in range(1, max_col + 1):
                col_letter = openpyxl.utils.get_column_letter(c_idx)
                max_len = 0
                for r_idx in range(1, min(max_row, 400) + 1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    if cell.value is None:
                        continue
                    max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 60)

            if name in ("Déjeuner", "Dîner"):
                ws.column_dimensions["A"].width = 34
                for idx, _day in enumerate(DAY_NAMES, start=2):
                    ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = 12
                ws.column_dimensions[openpyxl.utils.get_column_letter(2 + len(DAY_NAMES))].width = 12

            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0

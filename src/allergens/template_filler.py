
from __future__ import annotations

from datetime import date
from copy import copy
from typing import Dict, List, Tuple, Optional, Set
import os

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .config import (
    SERVICE_DEJ, SERVICE_DIN,
    REG_STANDARD, REG_VEGETARIEN, REG_VEGETALIEN, REG_HYPO, REG_SPEC_SANS, REG_SPEC_AVEC,
    ALLERGEN_COLUMNS
)
from .utils import normalize_space, normalize_key

# Positions fixes du template "Allergène déjeuner ..."
REGIME_START_ROW = {
    REG_STANDARD: 4,
    REG_VEGETARIEN: 8,
    REG_HYPO: 12,
    REG_VEGETALIEN: 16,
    REG_SPEC_SANS: 20,   # Sans lactose
    REG_SPEC_AVEC: 24,   # Spéciaux
}
COURSE_ROW_OFFSETS = {"entree": 0, "plat": 1, "fromage": 2, "dessert": 3}

# Dans le template, les allergènes sont en colonnes C -> R dans l'ordre ALLERGEN_COLUMNS
ALLERGEN_LETTERS = list("CDEFGHIJKLMNOPQR")
ALLERGEN_COLS = {ALLERGEN_COLUMNS[i]: ALLERGEN_LETTERS[i] for i in range(len(ALLERGEN_COLUMNS))}

MEAT_ENTRY_ROWS = [33, 35, 37]  # B33, B35, B37 (chacun sur 2 lignes fusionnées)

def copy_sheet(src_ws: Worksheet, dst_wb: Workbook, title: str) -> Worksheet:
    """Copie une feuille avec styles/dimensions/merges (format EXACT conservé)."""
    dst_ws = dst_wb.create_sheet(title=title)

    for col, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col].width = dim.width
        dst_ws.column_dimensions[col].hidden = dim.hidden
        dst_ws.column_dimensions[col].outlineLevel = dim.outlineLevel

    for r, dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[r].height = dim.height
        dst_ws.row_dimensions[r].hidden = dim.hidden
        dst_ws.row_dimensions[r].outlineLevel = dim.outlineLevel

    for row in src_ws.iter_rows():
        for cell in row:
            dst_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            # Copie sûre des styles (compatible openpyxl 3.1+)
            if cell.has_style:
                # On copie chaque composant ; évite les StyleProxy "unhashable" sur certaines versions
                dst_cell.number_format = cell.number_format
                dst_cell.alignment = copy(cell.alignment)
                dst_cell.font = copy(cell.font)
                dst_cell.fill = copy(cell.fill)
                dst_cell.border = copy(cell.border)
                # Protection volontairement ignorée : elle n'impacte pas l'affichage et peut casser la copie
            else:
                dst_cell.number_format = cell.number_format

    for merged_range in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merged_range))

    dst_ws.freeze_panes = src_ws.freeze_panes
    dst_ws.page_setup.orientation = src_ws.page_setup.orientation
    dst_ws.page_setup.paperSize = src_ws.page_setup.paperSize
    dst_ws.page_setup.fitToWidth = src_ws.page_setup.fitToWidth
    dst_ws.page_setup.fitToHeight = src_ws.page_setup.fitToHeight
    dst_ws.page_margins = src_ws.page_margins
    return dst_ws


def _clear_values_only(ws: Worksheet, cell_range: str) -> None:
    for row in ws[cell_range]:
        for cell in row:
            cell.value = None


def _set_service_title(ws: Worksheet, service: str, day: date) -> None:
    ddmm = day.strftime("%d/%m")
    if service == SERVICE_DEJ:
        ws["A1"].value = f"Déjeuner {ddmm}"
        ws["A3"].value = "Déjeuner  "
    else:
        ws["A1"].value = f"Dîner {ddmm}"
        ws["A3"].value = "Dîner  "


def _course_value(service_menu: Dict[str, str], key: str) -> str:
    v = normalize_space(str(service_menu.get(key, "—") or "—")).strip()
    if key == "plat":
        g = normalize_space(str(service_menu.get("garnitures", "—") or "—")).strip()
        if g and g != "—":
            if v == "—":
                v = g
            elif g.lower() not in v.lower():
                v = f"{v}, {g}"
    return v or "—"


def _fill_row(ws: Worksheet, row: int, dish: str, allergens: Optional[Set[str]]) -> None:
    ws[f"B{row}"].value = dish if dish else "—"
    for col_letter in ALLERGEN_COLS.values():
        ws[f"{col_letter}{row}"].value = None
    if not allergens:
        return
    for a in allergens:
        col = ALLERGEN_COLS.get(a)
        if col:
            ws[f"{col}{row}"].value = "X"


def _detect_meats(dish_texts: List[str]) -> List[str]:
    meats = []
    KW = [
        ("Bœuf", ["boeuf", "bœuf", "steak", "bourguignon", "pot au feu"]),
        ("Porc", ["porc", "jambon", "lardon", "tartiflette", "saucisse", "chipolata"]),
        ("Veau", ["veau", "blanquette"]),
        ("Poulet", ["poulet", "volaille"]),
        ("Dinde", ["dinde"]),
        ("Agneau", ["agneau", "mouton"]),
        ("Canard", ["canard"]),
        ("Lapin", ["lapin"]),
    ]
    for d in dish_texts:
        t = normalize_space(str(d or "")).lower()
        if not t or t == "—":
            continue
        for label, kws in KW:
            if any(k in t for k in kws):
                entry = f"{label} – {normalize_space(str(d)).strip()}"
                if entry not in meats:
                    meats.append(entry)
                break
        if len(meats) >= 3:
            break
    return meats[:3]


def _fill_meat_section(
    ws: Worksheet,
    meat_entries: List[str],
    meat_origin_ref: Optional[Dict[str, Tuple[str, str, str]]] = None,
) -> None:
    for r in MEAT_ENTRY_ROWS:
        ws[f"B{r}"].value = None
        ws[f"C{r}"].value = None
        ws[f"H{r}"].value = None
        ws[f"N{r}"].value = None
    for idx, entry in enumerate(meat_entries[:3]):
        r = MEAT_ENTRY_ROWS[idx]
        ws[f"B{r}"].value = entry

        if meat_origin_ref:
            k = normalize_key(entry)
            org = meat_origin_ref.get(k)
            if org:
                naissance, elevage, abattage = org
                if naissance:
                    ws[f"C{r}"].value = naissance
                if elevage:
                    ws[f"H{r}"].value = elevage
                if abattage:
                    ws[f"N{r}"].value = abattage


def fill_allergen_workbook(
    menus_by_day: Dict[date, Dict[str, Dict[str, Dict[str, str]]]],
    allergen_ref_key_to_allergens: Dict[str, Set[str]],
    template_dir: str,
    out_path: str,
    meat_origin_ref: Optional[Dict[str, Tuple[str, str, str]]] = None,
) -> Tuple[str, List[str]]:
    """Génère un classeur Excel au format EXACT du template, 1 feuille par service."""
    template_path = os.path.join(template_dir, "template_dejeuner.xlsx")
    wb_tpl = load_workbook(template_path)
    ws_tpl = wb_tpl.active

    out_wb = Workbook()
    out_wb.remove(out_wb.active)

    missing: List[str] = []

    for day in sorted(menus_by_day.keys()):
        day_block = menus_by_day[day]
        for service in (SERVICE_DEJ, SERVICE_DIN):
            title = ("Déj " if service == SERVICE_DEJ else "Din ") + day.strftime("%a %d-%m")
            ws = copy_sheet(ws_tpl, out_wb, title=title)
            _set_service_title(ws, service, day)

            _clear_values_only(ws, "B4:B27")
            _clear_values_only(ws, "C4:R27")
            _fill_meat_section(ws, [], meat_origin_ref)

            meat_candidates: List[str] = []

            for regime, start_row in REGIME_START_ROW.items():
                service_menu = day_block.get(regime, {}).get(service, {}) if day_block.get(regime) else {}

                entree = _course_value(service_menu, "entree")
                plat = _course_value(service_menu, "plat")
                fromage = _course_value(service_menu, "fromage")
                dessert = _course_value(service_menu, "dessert")

                for key, val in [("entree", entree), ("plat", plat), ("fromage", fromage), ("dessert", dessert)]:
                    row = start_row + COURSE_ROW_OFFSETS[key]
                    dish = normalize_space(val).strip() if val else "—"
                    allergens = None
                    if dish and dish != "—":
                        k = normalize_key(dish)
                        allergens = allergen_ref_key_to_allergens.get(k)
                        if allergens is None:
                            missing.append(dish)
                    _fill_row(ws, row, dish, allergens)

                meat_candidates.append(plat)

            meats = _detect_meats(meat_candidates)
            _fill_meat_section(ws, meats, meat_origin_ref)

    out_wb.save(out_path)

    # dedup missing order
    seen=set()
    missing2=[]
    for x in missing:
        k=normalize_key(x)
        if k and k not in seen:
            seen.add(k); missing2.append(x)
    return out_path, missing2

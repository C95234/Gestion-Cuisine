from __future__ import annotations

from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas


@dataclass
class SupplierInfo:
    name: str
    customer_code: str = ""
    coord1: str = ""
    coord2: str = ""


def _supplier_lookup(suppliers: List[Dict[str, str]]) -> Dict[str, SupplierInfo]:
    out: Dict[str, SupplierInfo] = {}
    for s in suppliers or []:
        name = str(s.get("name", "") or "").strip()
        if not name:
            continue
        out[name] = SupplierInfo(
            name=name,
            customer_code=str(s.get("customer_code", "") or ""),
            coord1=str(s.get("coord1", "") or ""),
            coord2=str(s.get("coord2", "") or ""),
        )
    return out


def group_lines_for_order(df: pd.DataFrame) -> pd.DataFrame:
    """Prépare un bon fournisseur lisible.

    - Produit : on prend 'Libellé' si présent, sinon 'Produit'
    - Regroupe par Produit + Unité
    - Somme des Quantités
    - Calcule Prix cible total (= Qté * PU) et Poids logistique total
      * 0,1 par 'unité'
      * 1 par 'kg' et 'L'
    """
    cols = ["Produit", "Unité", "Quantité", "Prix cible unitaire", "Prix cible total", "Poids total (kg)"]
    if df is None or df.empty:
        return pd.DataFrame(columns=cols)

    d = df.copy()

    # Produit (priorité Libellé si fourni)
    if "Libellé" in d.columns and "Produit" in d.columns:
        d["Produit"] = d["Libellé"].where(d["Libellé"].notna() & (d["Libellé"].astype(str).str.strip() != ""), d["Produit"])
    elif "Libellé" in d.columns and "Produit" not in d.columns:
        d["Produit"] = d["Libellé"]
    elif "Produit" not in d.columns:
        d["Produit"] = ""

    # Unité
    unit_col = None
    for cand in ["Unité", "Unite", "Unité", "U", "Unites"]:
        if cand in d.columns:
            unit_col = cand
            break
    d["Unité"] = d[unit_col] if unit_col else ""

    # Normalisation
    d["Produit"] = d["Produit"].fillna("").astype(str).str.strip()
    d["Unité"] = d["Unité"].fillna("").astype(str).str.strip()

    d["Quantité"] = pd.to_numeric(d.get("Quantité", 0), errors="coerce").fillna(0)

    pu = pd.to_numeric(d.get("Prix cible unitaire", pd.NA), errors="coerce")
    d["_pu"] = pu
    # Prix total uniquement si le PU est renseigné (sinon: vide)
    d["_prix_total"] = (d["Quantité"] * d["_pu"]).where(d["_pu"].notna())

    # Poids logistique total (kg)
    # Règle attendue :
    # - 1 kg par "kg" / "L" (litre)
    # - 0,1 kg par "unité" (et par défaut)
    # Si l'Excel contient déjà "Poids unitaire (kg)" ou "Poids total (kg)", on les respecte si > 0.
    poids_total_col = "Poids total (kg)" if "Poids total (kg)" in d.columns else None
    poids_unit_col = "Poids unitaire (kg)" if "Poids unitaire (kg)" in d.columns else None

    pt = pd.to_numeric(d.get(poids_total_col, 0) if poids_total_col else 0, errors="coerce").fillna(0)
    wu = pd.to_numeric(d.get(poids_unit_col, 0) if poids_unit_col else 0, errors="coerce").fillna(0)

    unit_norm = d["Unité"].astype(str).str.strip().str.lower()
    wu_auto = unit_norm.apply(
        lambda u: 1.0 if u in ("kg", "kilo", "kilogramme", "l", "litre", "litres") else 0.1
    )

    wu_final = wu.where(wu > 0, wu_auto)
    d["_poids_total"] = pt.where(pt > 0, d["Quantité"] * wu_final)
    key_cols = ["Produit", "Unité"]
    grouped = (
        d.groupby(key_cols, as_index=False)
        .agg(
            {
                "Quantité": "sum",
                "_prix_total": lambda s: s.sum(min_count=1),
                "_poids_total": "sum",
                "_pu": lambda s: s.dropna().unique().tolist(),
            }
        )
        .sort_values(key_cols)
        .reset_index(drop=True)
    )

    def _one_or_blank(v):
        """Retourne le PU s’il est unique dans le groupe, sinon vide."""
        uniq = []
        for x in (v or []):
            try:
                fx = float(x)
            except Exception:
                continue
            if pd.isna(fx) or fx == 0.0:
                continue
            uniq.append(fx)

        # unique (stable)
        seen = set()
        uniq2 = []
        for x in uniq:
            if x not in seen:
                seen.add(x)
                uniq2.append(x)

        if len(uniq2) == 1:
            return uniq2[0]
        return ""

    grouped["Prix cible unitaire"] = grouped["_pu"].apply(_one_or_blank)
    grouped["Prix cible total"] = grouped["_prix_total"].round(2)
    grouped["Prix cible total"] = grouped["Prix cible total"].where(grouped["Prix cible total"].notna(), "")
    grouped["Poids total (kg)"] = grouped["_poids_total"].round(3)

    grouped = grouped.drop(columns=["_pu", "_prix_total", "_poids_total"])

    for c in cols:
        if c not in grouped.columns:
            grouped[c] = ""
    return grouped[cols]

def split_grouped_into_lots(
    grouped: pd.DataFrame,
    *,
    delivery_dates: Optional[List[str]] = None,
    max_weight_kg: float = 600.0,
) -> List[Tuple[pd.DataFrame, str]]:
    """Split un bon fournisseur en lots (DataFrame) en respectant :
    - limite de poids `max_weight_kg`
    - priorisation des créneaux de livraison :
        * si 2 dates sont fournies : on remplit le lot 1 (date 1) jusqu'au seuil, puis le reste passe sur la date 2
        * si plus de dates : chaque nouveau lot prend la date suivante, puis on reste sur la dernière
    Retourne une liste de tuples (lot_df, delivery_date).
    """
    if grouped is None or grouped.empty:
        return []

    dates = [str(d).strip() for d in (delivery_dates or []) if str(d).strip()]
    # index de date utilisé par lot
    date_idx = 0

    lots: List[Tuple[pd.DataFrame, str]] = []
    current_rows: List[Dict[str, object]] = []
    current_w = 0.0

    def current_date() -> str:
        if not dates:
            return ""
        idx = min(max(date_idx, 0), len(dates) - 1)
        return dates[idx]

    for _, row in grouped.iterrows():
        w = float(row.get("Poids total (kg)", 0) or 0.0)

        # si on dépasse le seuil : on clôt le lot courant
        if current_rows and (current_w + w) > float(max_weight_kg):
            lots.append((pd.DataFrame(current_rows), current_date()))
            current_rows = []
            current_w = 0.0

            # lot suivant => date suivante (si dispo), sinon on reste sur la dernière
            if dates and date_idx < len(dates) - 1:
                date_idx += 1

        current_rows.append(row.to_dict())
        current_w += w

    if current_rows:
        lots.append((pd.DataFrame(current_rows), current_date()))

    return lots


def export_orders_per_supplier_excel(
    bon_df: pd.DataFrame,
    out_path: str,
    *,
    suppliers: Optional[List[Dict[str, str]]] = None,
    delivery_dates_by_supplier: Optional[Dict[str, List[str]]] = None,
    max_weight_kg: float = 600.0,
) -> None:
    """Crée un classeur Excel avec 1 feuille par bon fournisseur.

    - 1 fournisseur par bon
    - split automatique par poids cumulé (max_weight_kg)
    - si 2 dates sont fournies pour un fournisseur : on remplit d'abord le lot 1 (date 1) puis le reste bascule sur la date 2
    - colonnes affichées : Produit | Unité | Quantité | Prix cible unitaire | Prix cible total
    - ligne TOTAL (€) en bas
    """
    suppliers = suppliers or []
    sup_map = _supplier_lookup(suppliers)
    delivery_dates_by_supplier = delivery_dates_by_supplier or {}

    df = bon_df.copy() if bon_df is not None else pd.DataFrame()
    if df.empty:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Aucun fournisseur"
        ws["A1"].value = "Aucune ligne dans le bon de commande"
        wb.save(out_path)
        return

    if "Fournisseur" not in df.columns:
        df["Fournisseur"] = ""
    df["Fournisseur"] = df["Fournisseur"].fillna("").astype(str).str.strip()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    title_fill = PatternFill("solid", fgColor="2F5597")
    title_font = Font(bold=True, size=14, color="FFFFFF")
    subtitle_font = Font(bold=False, size=10, color="333333")

    header_fill = PatternFill("solid", fgColor="F2F2F2")
    header_font = Font(bold=True, color="1F1F1F")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    body_align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    body_align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    money_fmt = '#,##0.00'

    def _safe_sheet_name(name: str) -> str:
        name = (name or "Fournisseur").strip()[:28]
        return name or "Fournisseur"

    def _safe_sheet_name_with_suffix(base: str, suffix: str) -> str:
        """Nom d'onglet Excel compatible (31 caractères max, sans caractères interdits)."""
        base = (base or "Fournisseur").strip()
        suffix = (suffix or "").strip()
        # Caractères interdits par Excel: : \ / ? * [ ]
        for ch in [":", "\\", "/", "?", "*", "[", "]"]:
            base = base.replace(ch, "-")
            suffix = suffix.replace(ch, "-")
        if not suffix:
            name = base
        else:
            name = f"{base} - {suffix}"
        name = name.strip() or "Fournisseur"
        return name[:31]

    suppliers_in_df = [s for s in df["Fournisseur"].unique().tolist() if str(s).strip()]
    for supplier_name in sorted(suppliers_in_df):
        df_sup = df[df["Fournisseur"] == supplier_name].copy()

        # regroupe et recalcule poids/prix
        grouped = group_lines_for_order(df_sup)
        if grouped.empty:
            continue

        # split en lots + affectation de date par lot
        dates = delivery_dates_by_supplier.get(supplier_name) or []
        lots = split_grouped_into_lots(grouped, delivery_dates=dates, max_weight_kg=max_weight_kg)
        if not lots:
            continue

        for lot_idx, (lot_df, delivery_date) in enumerate(lots, start=1):
            title = _safe_sheet_name(supplier_name)
            # Si plusieurs lots : on préfère afficher la date de livraison plutôt que "(1/2)".
            if delivery_date:
                sheet_name = _safe_sheet_name_with_suffix(title, f"Liv {delivery_date}")
            else:
                sheet_name = title if len(lots) == 1 else _safe_sheet_name_with_suffix(title, f"Lot {lot_idx}/{len(lots)}")
            ws = wb.create_sheet(sheet_name)

            # En-tête
            ws.merge_cells("A1:E1")
            if delivery_date:
                ws["A1"].value = f"BON DE COMMANDE — {supplier_name} — Livraison : {delivery_date}"
            else:
                ws["A1"].value = f"BON DE COMMANDE — {supplier_name}" + (f" (Lot {lot_idx}/{len(lots)})" if len(lots) > 1 else "")
            ws["A1"].font = title_font
            ws["A1"].fill = title_fill
            ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[1].height = 24

            info = sup_map.get(supplier_name, SupplierInfo(name=supplier_name))
            ws["A2"].value = f"Code client: {info.customer_code}" if info.customer_code else ""
            ws["A3"].value = info.coord1 or ""
            ws["A4"].value = info.coord2 or ""

            if delivery_date:
                ws["D2"].value = "Livraison :"
                ws["D2"].font = header_font
                ws["D3"].value = str(delivery_date)

            for cell in ["A2", "A3", "A4", "D2", "D3"]:
                ws[cell].font = subtitle_font
                ws[cell].alignment = Alignment(horizontal="left", vertical="center")

            start_row = 6
            headers = ["Produit", "Unité", "Quantité", "Prix cible unitaire", "Prix cible total"]
            for c, h in enumerate(headers, start=1):
                cell = ws.cell(row=start_row, column=c, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_align
                cell.border = border

            # Lignes
            for r_i, (_, r) in enumerate(lot_df.iterrows(), start=1):
                rr = start_row + r_i
                ws.cell(rr, 1, r.get("Produit", "")).alignment = body_align_left
                ws.cell(rr, 2, r.get("Unité", "")).alignment = body_align_left
                ws.cell(rr, 3, r.get("Quantité", 0)).alignment = body_align_right
                ws.cell(rr, 4, r.get("Prix cible unitaire", "")).alignment = body_align_right
                ws.cell(rr, 5, r.get("Prix cible total", 0)).alignment = body_align_right

                ws.cell(rr, 4).number_format = money_fmt
                ws.cell(rr, 5).number_format = money_fmt

                for c in range(1, 6):
                    ws.cell(rr, c).border = border

            # Total
            total_row = start_row + len(lot_df) + 1
            ws.cell(total_row, 4, "TOTAL").font = header_font
            ws.cell(total_row, 4).alignment = body_align_right
            sum_cell = ws.cell(total_row, 5)
            s = pd.to_numeric(lot_df.get("Prix cible total", pd.Series(dtype=float)), errors="coerce")
            sum_val = float(s.sum()) if s.notna().any() else None
            sum_cell.value = sum_val
            sum_cell.number_format = money_fmt
            sum_cell.font = header_font
            for c in range(1, 6):
                ws.cell(total_row, c).border = border
                ws.cell(total_row, c).fill = header_fill

            # Largeurs
            ws.column_dimensions["A"].width = 38
            ws.column_dimensions["B"].width = 8
            ws.column_dimensions["C"].width = 12
            ws.column_dimensions["D"].width = 18
            ws.column_dimensions["E"].width = 18

    if not wb.sheetnames:
        ws = wb.create_sheet("Aucun fournisseur")
        ws["A1"].value = "Aucune ligne avec fournisseur renseigné"

    wb.save(out_path)

def export_orders_per_supplier_pdf(
    bon_df: pd.DataFrame,
    out_path: str,
    *,
    suppliers: Optional[List[Dict[str, str]]] = None,
    delivery_dates_by_supplier: Optional[Dict[str, List[str]]] = None,
    max_weight_kg: float = 600.0,
) -> None:
    """PDF : 1 page par bon fournisseur (split à `max_weight_kg`).

    Affiche : Produit | Unité | Quantité | Prix cible unitaire | Prix cible total + TOTAL.
    """
    suppliers = suppliers or []
    sup_map = _supplier_lookup(suppliers)
    delivery_dates_by_supplier = delivery_dates_by_supplier or {}

    df = bon_df.copy() if bon_df is not None else pd.DataFrame()
    c = canvas.Canvas(out_path, pagesize=A4)

    if df.empty:
        c.drawString(72, 800, "Aucune ligne dans le bon de commande")
        c.save()
        return

    if "Fournisseur" not in df.columns:
        df["Fournisseur"] = ""
    df["Fournisseur"] = df["Fournisseur"].fillna("").astype(str).str.strip()

    suppliers_in_df = [s for s in df["Fournisseur"].unique().tolist() if str(s).strip()]
    if not suppliers_in_df:
        c.drawString(72, 800, "Aucune ligne avec fournisseur renseigné")
        c.save()
        return

    for supplier_name in sorted(suppliers_in_df):
        df_sup = df[df["Fournisseur"] == supplier_name].copy()
        grouped = group_lines_for_order(df_sup)
        if grouped.empty:
            continue

        dates = delivery_dates_by_supplier.get(supplier_name) or []
        lots = split_grouped_into_lots(grouped, delivery_dates=dates, max_weight_kg=max_weight_kg)
        if not lots:
            continue

        info = sup_map.get(supplier_name, SupplierInfo(name=supplier_name))

        for lot_idx, (lot_df, delivery_date) in enumerate(lots, start=1):
            # Titre : on remplace "(1/2)" par la date de livraison si elle est fournie.
            if delivery_date:
                title = f"BON DE COMMANDE — {supplier_name} — Livraison : {delivery_date}"
            else:
                title = f"BON DE COMMANDE — {supplier_name}" + (f" (Lot {lot_idx}/{len(lots)})" if len(lots) > 1 else "")

            # Marges / largeur page (A4 portrait: 595 x 842)
            page_w = 595
            left = 40
            right = 40
            usable_w = page_w - left - right

            # Colonnes (un peu plus aérées)
            x_prod = left
            x_unit = left + 265
            x_qty_r = left + 355
            x_pu_r = left + 450
            x_total_r = page_w - right

            def draw_page_header() -> float:
                """Dessine l'en-tête + le header de tableau, et renvoie la position Y de départ des lignes."""
                # Bandeau titre (avec marges, pas en plein bord)
                c.setFillColorRGB(0.18, 0.34, 0.59)
                c.rect(left, 802, usable_w, 34, fill=1, stroke=0)
                c.setFillColorRGB(1, 1, 1)
                c.setFont("Helvetica-Bold", 15)
                c.drawString(left, 812, title)

                # Infos fournisseur
                c.setFillColorRGB(0, 0, 0)
                c.setFont("Helvetica", 9)
                y_info = 785
                if info.customer_code:
                    c.drawString(left, y_info, f"Code client : {info.customer_code}")
                    y_info -= 12
                if info.coord1:
                    c.drawString(left, y_info, info.coord1)
                    y_info -= 12
                if info.coord2:
                    c.drawString(left, y_info, info.coord2)
                    y_info -= 12

                # Tableau
                y_table = 740
                # Bande grisée pour l'entête du tableau (avec marges cohérentes)
                c.setFillColorRGB(0.9, 0.9, 0.9)
                # Encadrement léger
                c.setStrokeColorRGB(0.75, 0.75, 0.75)
                c.rect(left, y_table - 5, usable_w, 18, fill=1, stroke=1)
                c.setFillColorRGB(0, 0, 0)
                c.setFont("Helvetica-Bold", 9)
                c.drawString(x_prod, y_table, "Produit")
                c.drawString(x_unit, y_table, "Unité")
                c.drawRightString(x_qty_r, y_table, "Quantité")
                c.drawRightString(x_pu_r, y_table, "PU €")
                c.drawRightString(x_total_r, y_table, "Total €")
                return y_table - 14

            y_table = draw_page_header()
            c.setFont("Helvetica", 9)
            total = 0.0
            has_price = False
            for _, r in lot_df.iterrows():
                prod = str(r.get("Produit", "") or "")
                unit = str(r.get("Unité", "") or "")
                qty = float(r.get("Quantité", 0) or 0)
                pu = r.get("Prix cible unitaire", "")
                pt_raw = r.get("Prix cible total", "")
                pt_num = pd.to_numeric(pt_raw, errors="coerce")
                if pd.notna(pt_num):
                    pt = float(pt_num)
                    total += pt
                    has_price = True
                else:
                    pt = ""

                # découpe simple produit sur 2 lignes si trop long
                if len(prod) > 60:
                    p1, p2 = prod[:60], prod[60:120]
                    c.drawString(x_prod, y_table, p1); y_table -= 12
                    c.drawString(x_prod, y_table, p2)
                else:
                    c.drawString(x_prod, y_table, prod)

                c.drawString(x_unit, y_table, unit[:8])
                c.drawRightString(x_qty_r, y_table, f"{qty:g}")
                c.drawRightString(x_pu_r, y_table, f"{float(pu):.2f}" if str(pu).strip() != "" else "")
                c.drawRightString(x_total_r, y_table, f"{float(pt):.2f}" if str(pt).strip() != "" and pd.notna(pd.to_numeric(pt, errors="coerce")) else "")

                y_table -= 12
                if y_table < 80:
                    c.showPage()
                    y_table = draw_page_header()

            # Total mis en valeur
            y_total_box = max(y_table - 18, 55)
            c.setFillColorRGB(0.85, 0.85, 0.85)
            # Encadré total aligné avec les marges + un peu plus de padding
            total_box_w = 230
            total_box_h = 22
            total_box_x = (595 - right) - total_box_w
            c.setStrokeColorRGB(0.75, 0.75, 0.75)
            c.rect(total_box_x, y_total_box, total_box_w, total_box_h, fill=1, stroke=1)
            c.setFillColorRGB(0, 0, 0)
            c.setFont("Helvetica-Bold", 11)
            y_total_text = y_total_box + 7
            c.drawRightString((595 - right) - 55, y_total_text, "TOTAL COMMANDE :")
            c.drawRightString(595 - right, y_total_text, f"{total:.2f} €" if has_price else "—")

            c.showPage()

    c.save()

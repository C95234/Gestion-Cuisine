from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from typing import Optional
import re
import unicodedata

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .utils import normalize_space, strip_asterisks
from .config import (
    REG_STANDARD, REG_VEGETARIEN, REG_VEGETALIEN,
    REG_HYPO, REG_SPEC_AVEC, REG_SPEC_SANS,
    SERVICE_DEJ, SERVICE_DIN
)

# ============================================================
# CONFIG EN-TÊTES
# ============================================================

HEADER_HINTS = {
    REG_STANDARD: ["viande/poisson/œuf", "viande/poisson/oeuf", "viande", "poisson", "oeuf"],
    REG_VEGETARIEN: ["végétarien", "vegetarien"],
    REG_VEGETALIEN: ["végétalien", "vegetalien"],
    REG_HYPO: ["hypocalorique"],
    REG_SPEC_AVEC: ["avec lactose"],
    REG_SPEC_SANS: ["sans lactose"],
}

DAY_RE = re.compile(
    r"^(lundi|mardi|mercredi|jeudi|vendredi|samedi|dimanche)\s+(\d{1,2})\s+(\w+)\s+(\d{4})",
    re.IGNORECASE
)

MONTHS_FR = {
    "janvier": 1, "février": 2, "fevrier": 2, "mars": 3, "avril": 4, "mai": 5, "juin": 6,
    "juillet": 7, "août": 8, "aout": 8, "septembre": 9, "octobre": 10,
    "novembre": 11, "décembre": 12, "decembre": 12,
}

# ============================================================
# OUTILS EXCEL
# ============================================================

def _is_merged_top_left(ws, row: int, col: int) -> bool:
    cell = ws.cell(row=row, column=col)
    if cell.coordinate not in ws.merged_cells:
        return True
    for r in ws.merged_cells.ranges:
        if cell.coordinate in r:
            return r.min_row == row and r.min_col == col
    return True


def _merged_value(ws: Worksheet, row: int, col: int):
    cell = ws.cell(row=row, column=col)
    if cell.value not in (None, ""):
        return cell.value
    for merged in ws.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
            return ws.cell(row=merged.min_row, column=merged.min_col).value
    return None


# ============================================================
# DÉTECTION DES COLONNES DE RÉGIMES (CORRIGÉ)
# ============================================================

def _find_header_row_and_cols(ws: Worksheet) -> tuple[int, dict[str, int]]:
    """
    Détecte la ligne d'en-tête des régimes.
    Fix critique : "Lactose" seul = Spécial AVEC lactose
    """

    def strip_accents(s: str) -> str:
        s = unicodedata.normalize("NFD", s)
        return "".join(c for c in s if unicodedata.category(c) != "Mn")

    def norm(s: str) -> str:
        s = normalize_space(str(s or ""))
        s = strip_asterisks(s)
        s = strip_accents(s).lower()
        return re.sub(r"\s+", " ", s)

    best_row = None
    best_cols = {}
    best_score = -1

    for r in range(1, min(ws.max_row, 200) + 1):
        row_vals = []
        for c in range(1, min(ws.max_column, 50) + 1):
            v = norm(_merged_value(ws, r, c))
            vg = norm(_merged_value(ws, r - 1, c)) if r > 1 else ""
            # Certains fichiers mettent des regroupements (ex: HYPOCALORIQUE, MENU STANDARD) sur la ligne au-dessus.
            if not v and vg:
                v = vg
            elif v and vg and vg not in v:
                v = (vg + " " + v).strip()
            row_vals.append(v)

        joined = " ".join(v for v in row_vals if v)

        if "viande" not in joined or "vegetar" not in joined:
            continue

        cols = {}
        for c, v in enumerate(row_vals, start=1):
            if not v:
                continue

            if "sans lactose" in v:
                cols.setdefault(REG_SPEC_SANS, c)
                continue

            if "lactose" in v:
                cols.setdefault(REG_SPEC_AVEC, c)
                continue

            for reg, hints in HEADER_HINTS.items():
                if reg == REG_SPEC_AVEC:
                    continue
                if any(norm(h) in v for h in hints):
                    cols.setdefault(reg, c)

        required = {REG_STANDARD, REG_VEGETARIEN, REG_VEGETALIEN}
        if not required.issubset(cols):
            continue

        score = len(cols)
        if score > best_score:
            best_score = score
            best_row = r
            best_cols = cols

    if best_row is None:
        raise ValueError("Impossible de trouver l'en-tête des colonnes de régimes.")

    return best_row, best_cols


# ============================================================
# PARSING DATE
# ============================================================

def _parse_day_cell(value) -> Optional[date]:
    from datetime import datetime, date as ddate
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, ddate):
        return value

    text = normalize_space(str(value or "")).lower()
    m = DAY_RE.match(text)
    if not m:
        return None

    _, d, month, y = m.groups()
    return date(int(y), MONTHS_FR.get(month, 0), int(d))


def _clean_cell(v) -> str:
    return strip_asterisks(normalize_space(str(v or "")))


# ============================================================
# NORMALISATION TEXTE
# ============================================================

def _norm2(s: str) -> str:
    s = strip_asterisks(normalize_space(str(s or ""))).lower()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", s)


# ============================================================
# MOTS-CLÉS MÉTIER
# ============================================================

_DESSERT_KW = (
    "compote", "fruit", "gateau", "gâteau", "tarte", "flan",
    "mousse", "riz au lait", "éclair", "eclair"
)

_DAIRY_KW = (
    "fromage", "yaourt", "yogourt", "fromage blanc", "tomme",
    "camembert", "emmental", "kiri", "tartare", "babybel", "gouda", "boursin", "coulomiers", "petit suisse", "brie", "comte"
)

_ENTREE_KW = (
    "salade", "crud", "soupe", "carotte", "tomate", "lentilles",
    "œuf dur", "oeuf dur", "vinaigrette"
)


def _is_dessert(s: str) -> bool:
    return any(k in _norm2(s) for k in _DESSERT_KW)


def _is_dairy(s: str) -> bool:
    return any(k in _norm2(s) for k in _DAIRY_KW)


def _looks_like_entree(s: str) -> bool:
    return any(k in _norm2(s) for k in _ENTREE_KW)


# ============================================================
# CONSTRUCTION D'UN SERVICE
# ============================================================

def _build_service_positional(rows: list[str], service: str, is_vegan: bool) -> dict[str, str]:
    out = {"entree": "—", "plat": "—", "garnitures": "—", "fromage": "—", "dessert": "—"}

    for x in rows:
        if out["dessert"] == "—" and _is_dessert(x):
            out["dessert"] = x
        elif out["fromage"] == "—" and _is_dairy(x):
            out["fromage"] = x
        elif out["entree"] == "—":
            out["entree"] = x
        elif out["plat"] == "—":
            out["plat"] = x
        elif out["garnitures"] == "—":
            out["garnitures"] = x

    return out


# ============================================================
# LECTURE PRINCIPALE
# ============================================================

def read_menus(excel_path: str) -> dict[date, dict[str, dict[str, dict[str, str]]]]:
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active

    header_row, cols = _find_header_row_and_cols(ws)

    starts = []
    for r in range(header_row + 1, ws.max_row + 1):
        if not _is_merged_top_left(ws, r, 1):
            continue
        d = _parse_day_cell(_merged_value(ws, r, 1))
        if d:
            starts.append((r, d))

    menus = {}

    for i, (sr, d) in enumerate(starts):
        er = starts[i + 1][0] - 1 if i + 1 < len(starts) else ws.max_row
        day_menu = {}

        # --- Détection de la séparation Déjeuner / Dîner ---
        # On évite le découpage "moitié-moitié" qui casse dès qu'il y a une ligne de plat en moins/plus.
        def _row_texts(r: int) -> list[str]:
            texts = []
            for c in cols.values():
                v = _clean_cell(_merged_value(ws, r, c))
                if v:
                    texts.append(v)
            return texts

        best_split = None
        best_score = -1
        for r in range(sr, er):
            above = _row_texts(r)
            below = _row_texts(r + 1)
            if not above or not below:
                continue
            dessert_score = sum(1 for t in above if _is_dessert(t))
            entree_score = sum(1 for t in below if _looks_like_entree(t))
            dairy_score = sum(1 for t in above if _is_dairy(t))
            score = dessert_score * 3 + entree_score * 3 + dairy_score
            if score > best_score:
                best_score = score
                best_split = r

        # fallback sûr
        if best_split is None:
            best_split = sr + max(0, (er - sr) // 2)

        dej_range = range(sr, best_split + 1)
        din_range = range(best_split + 1, er + 1)


        for reg, col in cols.items():
            cells_dej = [
                _clean_cell(_merged_value(ws, r, col))
                for r in dej_range
                if _clean_cell(_merged_value(ws, r, col))
            ]
            cells_din = [
                _clean_cell(_merged_value(ws, r, col))
                for r in din_range
                if _clean_cell(_merged_value(ws, r, col))
            ]

            day_menu[reg] = {
                SERVICE_DEJ: _build_service_positional(cells_dej, SERVICE_DEJ, reg == REG_VEGETALIEN),
                SERVICE_DIN: _build_service_positional(cells_din, SERVICE_DIN, reg == REG_VEGETALIEN),
            }

        menus[d] = day_menu

    return menus

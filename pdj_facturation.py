from __future__ import annotations

import io
import os
import re
from dataclasses import dataclass
from typing import Dict, List, Tuple, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# PDF -> image
import fitz  # pymupdf
from PIL import Image, ImageOps, ImageEnhance
# OCR (requis pour PDF scanné)
import pytesseract


# Sur Streamlit Cloud, tesseract est à /usr/bin/tesseract si tu as packages.txt
# En local Windows/Mac, ça dépend (PATH). On ne force QUE si le chemin existe.
try:
    if os.path.exists("/usr/bin/tesseract"):
        pytesseract.pytesseract.tesseract_cmd = "/usr/bin/tesseract"
except Exception:
    pass


# --- Configuration modèle économat (colonnes sites) ---
SITE_COL_NAMES = [
    "FAM TL",
    "Bussière",
    "Bruyères",
    "FO",
    "MAS",
    "ESAT",
    "FM",
    "René Lalouette",
    "Internat",
]

# --- Mots-clés -> site (doit correspondre à une colonne du modèle) ---
# Objectif: détection robuste sans modifier l'UI.
# Tu peux enrichir librement.
SITE_KEYWORDS: Dict[str, str] = {
    # MAS Toulouse-Lautrec
    "toulouse lautrec": "MAS",
    "lautrec": "MAS",
    "mas": "MAS",

    # Internat / sites internes (ex: Rosa Bonheur 24 TER)
    "internat": "Internat",
    "24 ter": "Internat",
    "24t": "Internat",
    "rosa bonheur": "Internat",
    "léonard de vinci": "Internat",
    "leonard de vinci": "Internat",
}

# Quelques normalisations de libellés (tu peux enrichir)
ALIASES = {
    "yaourt nature": "Yaourt",
    "lait demi": "Lait 1/2 écrémé",
    "lait demi-écrémé": "Lait 1/2 écrémé",
    "lait 1/2": "Lait 1/2 écrémé",
    "jus de pomme": "Jus Pomme",
    "jus pomme": "Jus Pomme",
    "jus d'orange": "Jus Orange",
    "jus orange": "Jus Orange",
    "jus de raisin": "Jus Raisin",
    "jus raisin": "Jus Raisin",
    "tablettes chocolat": "Chocolat",
    "nesquick": "Chocolat",
    "sucre en poudre": "Sucre",
    "sucre en poudre (sachets)": "Sucre",
}

# Stop-words / bruit OCR
BAD_TOKENS = {
    "quantité",
    "livré",
    "reçu",
    "commandé",
    "commande",
    "total",
    "produit",
    "unité",
    "prix",
}


def _safe_float(x) -> Optional[float]:
    """Convertit en float si possible, sinon None (gère virgule)."""
    if x is None:
        return None
    try:
        if isinstance(x, (int, float)) and pd.notna(x):
            return float(x)
    except Exception:
        pass
    s = str(x).strip()
    if not s or s.lower() in {"nan", "none"}:
        return None
    m = re.fullmatch(r"\s*([0-9]+(?:[.,][0-9]+)?)\s*", s)
    if not m:
        return None
    try:
        return float(m.group(1).replace(",", "."))
    except Exception:
        return None


@dataclass
class ParsedOrder:
    filename: str
    site: Optional[str]
    # liste (produit_normalisé, quantité_commandée)
    items: List[Tuple[str, float]]


# ----------------- Utilitaire lecture fichier (UploadedFile OU chemin str) -----------------

def _read_any_file(f) -> Tuple[str, bytes]:
    """
    Accepte:
    - UploadedFile Streamlit (read + name)
    - chemin str vers un fichier
    - file-like (read) sans name
    Retour: (filename, bytes)
    """
    # Cas: chemin string
    if isinstance(f, str):
        filename = os.path.basename(f)
        with open(f, "rb") as fh:
            return filename, fh.read()

    # Cas: UploadedFile / file-like
    filename = getattr(f, "name", "bon")
    if hasattr(f, "read"):
        return filename, f.read()

    raise TypeError(f"Type de fichier non supporté: {type(f)}")


# ----------------- Détection site -----------------

def _detect_site_from_text(text: str, filename: str = "") -> Optional[str]:
    t = (text or "").lower() + " " + (filename or "").lower()
    t = re.sub(r"\s+", " ", t).strip()

    # 1) Dictionnaire de mots-clés -> site
    for kw, site in SITE_KEYWORDS.items():
        if kw and kw.lower() in t:
            return site

    # 2) Heuristiques supplémentaires (titres type "IME ..." / "MAS ...")
    # Ex: "BON DE COMMANDE PETIT DEJEUNER MAS TOULOUSE LAUTREC"
    m = re.search(r"\b(mas|ime)\s+([a-z0-9\-\"' ]{3,})", t, flags=re.I)
    if m:
        bloc = (m.group(0) or "").lower()
        for kw, site in SITE_KEYWORDS.items():
            if kw and kw.lower() in bloc:
                return site

    return None


# ----------------- Normalisation produit -----------------

def _normalize_product_name(raw: str) -> str:
    s = (raw or "").strip()
    s = re.sub(r"\s+", " ", s)

    low = s.lower()

    for k, v in ALIASES.items():
        if k in low:
            return v

    if "lait" in low and ("1/2" in low or "demi" in low):
        return "Lait 1/2 écrémé"

    return s


# ----------------- Parsing PDF (scanné) -----------------

def _pdf_to_images(pdf_bytes: bytes, dpi_scale: float = 2.0) -> List[Image.Image]:
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    images: List[Image.Image] = []
    for page in doc:
        mat = fitz.Matrix(dpi_scale, dpi_scale)
        pix = page.get_pixmap(matrix=mat)
        img = Image.open(io.BytesIO(pix.tobytes("png")))
        images.append(img)
    return images


def _ocr_text_from_pdf_bytes(pdf_bytes: bytes) -> str:
    images = _pdf_to_images(pdf_bytes, dpi_scale=2.0)
    out = []
    for img in images:
        out.append(pytesseract.image_to_string(img, lang="fra"))
    return "\n".join(out)



def _ocr_header_text_from_pdf_bytes(pdf_bytes: bytes) -> str:
    """OCR ciblé sur l'entête du 1er page (détection du site)."""
    try:
        images = _pdf_to_images(pdf_bytes, dpi_scale=3.0)
        if not images:
            return ""
        img = images[0]
        w, h = img.size
        header = img.crop((0, 0, w, int(h * 0.22)))

        gray = ImageOps.grayscale(header)
        gray = ImageEnhance.Contrast(gray).enhance(2.5)
        gray = ImageEnhance.Sharpness(gray).enhance(2.0)
        bw = gray.point(lambda x: 0 if x < 180 else 255, "1")

        return pytesseract.image_to_string(bw, lang="fra", config="--psm 6")
    except Exception:
        return ""


def _parse_items_from_ocr_text(text: str) -> List[Tuple[str, float]]:
    """
    Parse robuste pour bons PDJ scannés (tableau).

    - Supporte "Produit" sur une ligne puis quantité sur la suivante
    - Corrige quelques erreurs OCR fréquentes (GO->30, U->4)
    - Ignore les poids/grammages (ex: 250g, 5kg) pour éviter les faux positifs
    """
    items: List[Tuple[str, float]] = []
    lines = [l.strip() for l in (text or "").splitlines() if l.strip()]

    last_product: Optional[str] = None
    pending_qty: Optional[float] = None  # quand l'OCR décale la quantité

    def _to_qty(token: str) -> Optional[float]:
        t = (token or "").strip()
        if not t:
            return None

        # Corrections OCR observées sur tes bons
        if t.upper() == "GO":  # "30" mal lu
            return 30.0
        if t.upper() == "U":   # "4" mal lu
            return 4.0

        if re.fullmatch(r"\d+(?:[.,]\d+)?", t):
            try:
                return float(t.replace(",", "."))
            except Exception:
                return None
        return None

    def _looks_like_weight(line_low: str) -> bool:
        # vrai poids: doit contenir un nombre + (g|kg)
        return bool(re.search(r"\b\d+\s*(?:kg|g)\b", line_low))

    def _is_bad_product(name_low: str) -> bool:
        # lignes souvent non commandées / parasites sur tes PDJ
        return any(x in name_low for x in ["sel", "poivre"])

    for l in lines:
        low = l.lower()

        # ignorer bruit
        if any(tok in low for tok in BAD_TOKENS):
            continue

        # ignorer les lignes de poids
        if _looks_like_weight(low):
            # on ne prend PAS de nombre depuis ces lignes
            if not re.search(r"\d", l) and len(l) > 2:
                last_product = l
            continue

        # Cas A: ligne = juste une quantité (ex: "30" sous le produit)
        qty_only = _to_qty(l)
        if qty_only is not None:
            if last_product and not _is_bad_product(last_product.lower()):
                if 0 < qty_only <= 200:
                    items.append((_normalize_product_name(last_product), float(qty_only)))
                last_product = None
                pending_qty = None
            else:
                # on garde la quantité en attente (décalage OCR)
                pending_qty = float(qty_only)
                last_product = None
            continue

        # Cas B: "Produit <token>"
        m = re.match(r"^(.*?)(?:\s+)([A-Za-z]{1,2}|\d+(?:[.,]\d+)?)\s*$", l)
        if m:
            name = m.group(1).strip(" -:\t")
            token = m.group(2).strip()
            qty = _to_qty(token)
            if name and qty is not None and 0 < qty <= 200:
                nlow = name.lower()
                if "date" not in nlow and "commande" not in nlow:
                    norm = _normalize_product_name(name)
                    # évite un faux positif fréquent: la date "26/27/30" attachée aux condiments
                    if norm in {"Ketchup", "Mayonnaise", "Sel", "Poivre"} and 20 <= qty <= 31:
                        continue
                    items.append((norm, float(qty)))
                last_product = None
                pending_qty = None
                continue

        # Cas C: ligne texte (produit)
        if not re.search(r"\d", l):
            # si on avait une quantité en attente, on tente de l'appliquer ici
            if pending_qty is not None:
                norm = _normalize_product_name(l)
                if 0 < pending_qty <= 200:
                    items.append((norm, float(pending_qty)))
                pending_qty = None
                last_product = None
                continue

            if len(l) > 2 and not any(x in low for x in ["date", "commande", "livraison"]):
                last_product = l
            continue

        # sinon ignore

    return items


def _parse_excel(file_bytes: bytes, filename: str) -> ParsedOrder:
    """Parse Excel.

    Supporte 2 formats:
    1) Tableau structuré avec en-têtes (Produit / Quantité...)
    2) Modèle 'COMMANDE DES PETITS-DEJEUNERS' (sans en-têtes pandas):
       - repère la section 'Ingrédients à commander'
       - produit en col0, quantité en col1, commentaire en col2
    """

    bio = io.BytesIO(file_bytes)
    # On lit d'abord en mode standard (avec en-tête).
    try:
        df = pd.read_excel(bio)
    except Exception:
        df = pd.DataFrame()

    header_text = " ".join([str(c) for c in getattr(df, "columns", [])])
    site = _detect_site_from_text(header_text, filename)

    items: List[Tuple[str, float]] = []

    col_prod = None
    for cand in ["Produit", "Désignation", "Designation", "Article", "Libellé", "Libelle"]:
        if cand in df.columns:
            col_prod = cand
            break

    col_qty = None
    for cand in ["Qté", "Qte", "Quantité", "Quantite", "Commande", "Commandé", "Commandee"]:
        if cand in df.columns:
            col_qty = cand
            break

    if col_prod and col_qty and not df.empty:
        for _, r in df.iterrows():
            prod_raw = str(r.get(col_prod, "")).strip()
            if not prod_raw or prod_raw.lower() in BAD_TOKENS:
                continue
            try:
                qty = float(str(r.get(col_qty, "0")).replace(",", "."))
            except Exception:
                continue
            if qty <= 0:
                continue
            items.append((_normalize_product_name(prod_raw), qty))
    else:
        # Fallback 1: format 'COMMANDE DES PETITS-DEJEUNERS' (souvent sans en-têtes)
        bio2 = io.BytesIO(file_bytes)
        df2 = pd.read_excel(bio2, header=None)

        header_blob = " ".join([str(x) for x in df2.head(25).stack().dropna().tolist()])
        site = site or _detect_site_from_text(header_blob, filename)

        # trouve la ligne 'Ingrédients à commander'
        start_row = None
        for i in range(len(df2)):
            row = " ".join([str(x) for x in df2.iloc[i].fillna("").tolist()])
            if "ingrédients" in row.lower() and "commander" in row.lower():
                start_row = i + 1
                break
        if start_row is None:
            start_row = 0

        empty_streak = 0
        for i in range(start_row, len(df2)):
            prod = df2.iloc[i, 0] if df2.shape[1] > 0 else None
            qty = df2.iloc[i, 1] if df2.shape[1] > 1 else None

            prod_s = "" if pd.isna(prod) else str(prod).strip()
            qty_n = _safe_float(qty)

            if not prod_s and qty_n is None:
                empty_streak += 1
                if empty_streak >= 5:
                    break
                continue
            empty_streak = 0

            if prod_s.lower().startswith("autres"):
                break
            if not prod_s:
                continue

            if qty_n is None or qty_n <= 0:
                # pas de quantité -> on ignore (mais tu peux enrichir si besoin)
                continue

            items.append((_normalize_product_name(prod_s), float(qty_n)))

        # Fallback 2: colonnes numériques (rare)
        if not items and not df.empty:
            for c in df.columns:
                if isinstance(c, str) and c.strip():
                    s = df[c]
                    if pd.api.types.is_numeric_dtype(s):
                        qty = float(s.fillna(0).sum())
                        if qty > 0:
                            items.append((_normalize_product_name(c), qty))

    return ParsedOrder(filename=filename, site=site, items=items)


# ----------------- API attendue par app.py -----------------

def parse_pdj_order_file(uploaded_file) -> ParsedOrder:
    """
    Fonction attendue par ton app.py.
    Accepte UploadedFile Streamlit OU chemin str.
    """
    name, file_bytes = _read_any_file(uploaded_file)

    if name.lower().endswith(".pdf"):
        text = _ocr_text_from_pdf_bytes(file_bytes)

        # Détection site: OCR complet puis OCR ciblé entête (plus fiable sur scans pâles)
        site = _detect_site_from_text(text, name)
        if not site:
            header_text = _ocr_header_text_from_pdf_bytes(file_bytes)
            site = _detect_site_from_text(header_text, name)

        items = _parse_items_from_ocr_text(text)
        return ParsedOrder(filename=name, site=site, items=items)

    if name.lower().endswith((".xls", ".xlsx", ".xlsm")):
        return _parse_excel(file_bytes, name)

    return ParsedOrder(filename=name, site=None, items=[])


def update_facturation_economat(
    modele_xlsx,
    order_files: List,
    mois: Optional[str] = None,
) -> bytes:
    """
    Fonction attendue par ton app.py.
    - modele_xlsx: UploadedFile Streamlit (xlsx) OU chemin str
    - order_files: liste de fichiers uploadés (pdf/xls/xlsx) OU chemins str
    - mois: 'YYYY-MM' optionnel
    Retour: bytes du fichier xlsx généré (pour st.download_button)
    """

    # 1) Charger modèle en openpyxl (conserver styles)
    if isinstance(modele_xlsx, str):
        wb = load_workbook(modele_xlsx)
    else:
        _, model_bytes = _read_any_file(modele_xlsx)
        wb = load_workbook(io.BytesIO(model_bytes))

    ws = wb.active

    # 2) Trouver ligne en-tête (celle qui contient "Produit")
    header_row = None
    for r in range(1, 50):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and v.strip().lower() == "produit":
            header_row = r
            break
    if header_row is None:
        header_row = 4

    # 3) Identifier colonnes
    col_by_name: Dict[str, int] = {}
    max_col = ws.max_column
    for c in range(1, max_col + 1):
        val = ws.cell(header_row, c).value
        if isinstance(val, str):
            col_by_name[val.strip()] = c

    col_prix = col_by_name.get("Prix unitaire (€)", 3)
    col_produit = col_by_name.get("Produit", 1)

    site_cols: Dict[str, int] = {}
    for s in SITE_COL_NAMES:
        if s in col_by_name:
            site_cols[s] = col_by_name[s]

    col_total_produit = col_by_name.get("Total produit (€)")
    if not col_total_produit:
        col_total_produit = ws.max_column

    # 4) Dictionnaire produit -> ligne
    product_row: Dict[str, int] = {}
    first_data_row = header_row + 1

    total_site_row = None
    for r in range(first_data_row, ws.max_row + 1):
        v = ws.cell(r, col_produit).value
        if isinstance(v, str) and "total site" in v.lower():
            total_site_row = r
            break

    last_product_row = (total_site_row - 1) if total_site_row else ws.max_row

    for r in range(first_data_row, last_product_row + 1):
        v = ws.cell(r, col_produit).value
        if isinstance(v, str) and v.strip():
            product_row[v.strip().lower()] = r

    # 5) Parser tous les bons + agréger (site, produit) = somme
    parsed: List[ParsedOrder] = [parse_pdj_order_file(f) for f in order_files]

    agg: Dict[Tuple[str, str], float] = {}
    for po in parsed:
        if not po.site:
            continue
        for prod, qty in po.items:
            if qty <= 0:
                continue
            key = (po.site, prod)
            agg[key] = agg.get(key, 0.0) + float(qty)

    def find_row_for_product(prod_name: str) -> Optional[int]:
        low = prod_name.strip().lower()
        if low in product_row:
            return product_row[low]
        for k, r in product_row.items():
            if low in k or k in low:
                return r
        return None

    # 6) Écrire les quantités (valeurs)
    for (site, prod), qty in agg.items():
        if site not in site_cols:
            continue
        r = find_row_for_product(_normalize_product_name(prod))
        if not r:
            # produit non trouvé => ignoré (pour ne pas casser le modèle)
            continue
        c = site_cols[site]
        cur = ws.cell(r, c).value
        try:
            cur_num = float(str(cur).replace(",", ".")) if cur not in (None, "") else 0.0
        except Exception:
            cur_num = 0.0
        ws.cell(r, c).value = cur_num + qty

    # 7) Écrire le mois si demandé
    if mois:
        for r in range(1, 15):
            for c in range(1, 6):
                v = ws.cell(r, c).value
                if isinstance(v, str) and "mois" in v.lower():
                    ws.cell(r, c + 1).value = mois
                    break

    # 8) Formules totaux (modifiable)
    site_col_indices = [site_cols[s] for s in SITE_COL_NAMES if s in site_cols]
    if site_col_indices:
        first_site_col = min(site_col_indices)
        last_site_col = max(site_col_indices)

        for r in range(first_data_row, last_product_row + 1):
            v = ws.cell(r, col_produit).value
            if not isinstance(v, str) or not v.strip():
                continue
            prix_cell = f"{get_column_letter(col_prix)}{r}"
            sum_range = (
                f"{get_column_letter(first_site_col)}{r}:{get_column_letter(last_site_col)}{r}"
            )
            ws.cell(r, col_total_produit).value = f"={prix_cell}*SUM({sum_range})"

        if total_site_row:
            prix_range = (
                f"{get_column_letter(col_prix)}{first_data_row}:{get_column_letter(col_prix)}{last_product_row}"
            )
            for s, c in site_cols.items():
                qty_range = (
                    f"{get_column_letter(c)}{first_data_row}:{get_column_letter(c)}{last_product_row}"
                )
                ws.cell(total_site_row, c).value = f"=SUMPRODUCT({prix_range},{qty_range})"

            total_range = (
                f"{get_column_letter(col_total_produit)}{first_data_row}:{get_column_letter(col_total_produit)}{last_product_row}"
            )
            ws.cell(total_site_row, col_total_produit).value = f"=SUM({total_range})"

    # 9) Retour bytes
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

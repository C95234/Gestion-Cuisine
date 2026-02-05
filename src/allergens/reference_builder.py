from __future__ import annotations

from typing import List, Set
import re

from openpyxl import load_workbook, Workbook

from .config import ALLERGEN_COLUMNS

_STOPWORDS = {
    "dejeuner","déjeuner","diner","dîner","menu","semaine","jour","lundi","mardi","mercredi",
    "jeudi","vendredi","samedi","dimanche","entree","entrée","entrees","entrées","plat","plats",
    "accompagnement","accompagnements","fromage","dessert","desserts","standard","standards",
    "vegetarien","végétarien","vegetariens","végétariens","vegetalien","végétalien","vegetaliens",
    "végétaliens","hypocalorique","hypocaloriques","sans lactose","avec lactose","speciaux","spéciaux",
    "regime","régime","regimes","régimes","quantite","quantité","grammage","g","kg"
}

def _clean_text(s: str) -> str:
    s = str(s or "").strip()
    s = re.sub(r"\s+", " ", s)
    s = s.strip("•-–—*· ")
    return s

def extract_unique_items_from_menu(menu_path: str) -> List[str]:
    """Extraction générique : parcourt toutes les cellules texte du classeur
    et renvoie une liste dédupliquée de libellés potentiels (plats/denrées).
    """
    wb = load_workbook(menu_path, data_only=True)
    found: Set[str] = set()

    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if v is None:
                    continue
                if isinstance(v, (int, float)):
                    continue
                txt = _clean_text(str(v))
                if not txt:
                    continue
                low = txt.lower()
                # Filtre basique : évite les titres / cellules très courtes
                if len(low) < 3:
                    continue
                if low in _STOPWORDS:
                    continue
                # ignore "10 kg", "250 g"
                if re.fullmatch(r"\d+(?:[\.,]\d+)?\s*(kg|g|l|cl|ml)", low):
                    continue
                # ignore dates
                if re.fullmatch(r"\d{1,2}[/-]\d{1,2}(?:[/-]\d{2,4})?", low):
                    continue

                found.add(txt)

    # tri stable (alpha) pour faciliter la complétion
    return sorted(found, key=lambda x: x.lower())

def build_reference_template(menu_path: str, out_path: str) -> str:
    """Crée un fichier Excel 'référentiel allergènes' à compléter, à partir du menu."""
    items = extract_unique_items_from_menu(menu_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Référentiel"

    headers = ["Plat"] + ALLERGEN_COLUMNS
    ws.append(headers)
    for it in items:
        ws.append([it] + [""] * len(ALLERGEN_COLUMNS))

    # un peu de lisibilité
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 45
    for i in range(2, len(headers)+1):
        ws.column_dimensions[chr(64+i)].width = 18  # ok jusqu'à Z; suffisant ici

    wb.save(out_path)
    return out_path
